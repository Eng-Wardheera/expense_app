import calendar
from collections import defaultdict
import datetime
from decimal import Decimal, InvalidOperation
import io
import json
from flask_mail import Message   # ✅ CORRECT
from openpyxl import Workbook, load_workbook

import math
import os
import random
import re
import secrets
import traceback
import uuid
import pytz
from bson import ObjectId
import cloudinary
from flask import Blueprint, abort, current_app, flash, jsonify, make_response, redirect, render_template, request, session, url_for
from flask_login import current_user, login_required, login_user, logout_user
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename
from app import ALLOWED_EXTENSIONS, EAT, google, now_eat
from app.extensions import mongo, mail
from datetime import datetime, timedelta
from xhtml2pdf import pisa
from flask import Response
import dns.resolver  # Ku dar kor faylkaaga
from flask import send_file

from app.modal import Account, Category, Saving, SavingTransaction, Transaction, User, UserRole


bp = Blueprint('main', __name__)

#------------------------------------------
#---- Function: 1 | Func Allowed Files  ---
#------------------------------------------
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
 
def create_guest_session(mongo):
    if not session.get("guest_token"):

        token = secrets.token_hex(24)

        session["guest_token"] = token

        mongo.db.sessions.insert_one({
            "session_token": token,
            "user_id": None,   # guest
            "ip": request.remote_addr,
            "device": request.user_agent.string,
            "created_at": datetime.utcnow(),
            "expires_at": None,
            "routes": []   # store visited pages
        })



# 1. Index route: Wuxuu soo bandhigayaa page-ka iyo data-da projects-ka
from bson import ObjectId

@bp.route("/")
def index():

    slides = [
            {
                "title": "Manage Your Expenses",
                "description": "Track your income, expenses, savings and budgets easily in one place.",
                "image": "frontend/images/image.webp"
            },
            {
                "title": "Control Your Money",
                "description": "Monitor every transaction and stay in control of your finances.",
                "image": "frontend/images/image5.jpg"
            },
            {
                "title": "Reach Your Saving Goals",
                "description": "Create saving plans and watch your balance grow over time.",
                "image": "frontend/images/image3.jpg"
            },
            {
                "title": "Smart Financial Reports",
                "description": "Generate detailed reports and analyze your financial performance instantly.",
                "image": "frontend/images/image4.jpg"
            }
        ]

    features = [
        {
            "icon": "fa-wallet",
            "title": "Accounts",
            "text": "Manage multiple bank and cash accounts."
        },
        {
            "icon": "fa-money-bill-wave",
            "title": "Expenses",
            "text": "Record every expense instantly."
        },
        {
            "icon": "fa-piggy-bank",
            "title": "Savings",
            "text": "Track savings goals."
        },
        {
            "icon": "fa-chart-line",
            "title": "Reports",
            "text": "Powerful financial reports."
        }
    ]

    # Total Users (excluding Super Admin)
    total_users = mongo.db.users.count_documents({
        "role": {
            "$in": ["admin", "user"]
        },
        "status": True
    })

    # Total Admins
    total_admins = mongo.db.users.count_documents({
        "role": "admin",
        "status": True
    })

    # Total Normal Users
    total_members = mongo.db.users.count_documents({
        "role": "user",
        "status": True
    })

    # Latest Registered Users
    latest_users = list(
        mongo.db.users.find(
            {
                "role": {
                    "$in": ["admin", "user"]
                }
            },
            {
                "password": 0
            }
        )
        .sort("created_at", -1)
        .limit(8)
    )

    # Total Transactions
    total_transactions = mongo.db.transactions.count_documents({})

    # Total Accounts (Companies = Accounts)
    total_accounts = mongo.db.accounts.count_documents({})

    # Total Savings
    total_savings = mongo.db.savings.count_documents({})

    # Total Saving Transactions
    total_saving_transactions = mongo.db.saving_transactions.count_documents({})

    return render_template(
        "frontend/home/index.html",
        slides=slides,
        features=features,

        total_users=total_users,
        total_admins=total_admins,
        total_members=total_members,
        latest_users=latest_users,

        total_transactions=total_transactions,
        total_accounts=total_accounts,
        total_savings=total_savings,
        total_saving_transactions=total_saving_transactions,

        login_url="https://maareye.vercel.app/login",
        app_url="https://appsgeyser.io/19942993/Maareeye Expense"
    )





@bp.route('/check-username', methods=['POST'])
def check_username():
    username = request.json.get('username')
    user = mongo.db.users.find_one({"username": username})
    
    if user:
        # Soo saar 3 magac oo kale
        suggestions = [f"{username}{random.randint(10,99)}" for _ in range(3)]
        return jsonify({"taken": True, "suggestions": suggestions})
    
    return jsonify({"taken": False})


def is_valid_email_domain(email):
    try:
        domain = email.split('@')[1]
        # Waxaan hubineynaa in domain-ku leeyahay MX record (Mail Exchange)
        records = dns.resolver.resolve(domain, 'MX')
        return True if records else False
    except:
        return False

@bp.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        username = request.form.get('username')
        fullname = request.form.get('fullname')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('password_confirmation')
        
         # 1. Hubi in format-ku sax yahay (Regex)
        if not re.match(r"[^@]+@[^@]+\.[^@]+", email):
            flash("Fadlan geli email sax ah!", "danger")
            return redirect(url_for('main.register'))

        # 2. Hubi in domain-ku dhab ahaan u jiro (MX check)
        if not is_valid_email_domain(email):
            flash("Email-kan domain-kiisu ma jiro (Email does not exist)!", "danger")
            return redirect(url_for('main.register'))
        
        # 3. Hubi haddii user-ku horey u jiray
        if mongo.db.users.find_one({"email": email}):
            flash("Email-kan horey ayaa loo isticmaalay!", "danger")
            return redirect(url_for('main.register'))

        # 4. Hubi username-ka inuu database-ka ku jiro mar kale
        if mongo.db.users.find_one({"username": username}):
            flash("Username-kan horey ayaa loo qaatay, fadlan mid kale dooro!", "danger")
            return redirect(url_for('main.register'))
        
        # 5. Hubi xoogga password-ka (8 xaraf, 1 xaraf weyn, 1 lambar, 1 calaamad)
        if not re.match(r"^(?=.*[A-Za-z])(?=.*\d)(?=.*[@$!%*#?&])[A-Za-z\d@$!%*#?&]{8,}$", password):
            flash("Password-ku waa inuu ka koobnaadaa ugu yaraan 8 xaraf, lambar, iyo calaamad!", "danger")
            return redirect(url_for('main.register'))
        
        # 6. Hubi haddii passwords-ku isku mid yihiin
        if password != confirm_password:
            flash("Passwords-ka isma laha!", "danger")
            return redirect(url_for('main.register'))

        # 7. Role Logic
        user_count = mongo.db.users.count_documents({})
        role = UserRole.superadmin.value if user_count == 0 else UserRole.admin.value

        # 8. Save
        new_user = {
            "fullname": fullname,
            "username": username,
            "email": email,
            "password": generate_password_hash(password),
            "role": role,
            "status": True,
            "created_at": datetime.utcnow()
        }
        mongo.db.users.insert_one(new_user)
        
        flash("Diiwaangelinta way guulaysatay!", "success")
        return redirect(url_for('main.login'))

    # Wadada saxda ah ee faylkaaga:
    return render_template("backend/auth/auth-register.html")


@bp.route('/login', methods=['GET', 'POST'])
def login():
    # Haddi uu user-ku horay u soo galay, u dir dashboard-ka
    if current_user.is_authenticated:
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        remember = True if request.form.get('remembr_me') else False

        # 1. Ka raadi user-ka database-ka
        user_data = mongo.db.users.find_one({"email": email})

        # 2. Hubi haddii password-ku sax yahay
        if user_data and check_password_hash(user_data.get('password'), password):
            # Samee User object
            user = User(user_data) 
            
            # 3. Login u samee
            login_user(user, remember=remember)
            
            flash("Si guul leh ayaad u gashay dashboard-ka!", "success")
            return redirect(url_for('main.dashboard')) 
        else:
            flash("Email ama Password khaldan!", "danger")
            # Waxaan u beddelay 'auth.login' si uu ugu laabto isla boggaas
            return redirect(url_for('main.login')) 

    return render_template("backend/auth/auth-login.html")


@bp.app_errorhandler(403)
def forbidden(error):
    return render_template('frontend/errors/403.html'), 403

@bp.route("/login/google")
def login_google():
    redirect_uri = url_for("main.google_callback", _external=True)
    print("REDIRECT URI:", redirect_uri)
    return google.authorize_redirect(redirect_uri)



@bp.route("/google/callback")
def google_callback():
    token = google.authorize_access_token()
    user_info = token.get("userinfo")
    email = user_info.get("email")

    # 1. Check if the user exists in your database
    raw_user = mongo.db.users.find_one({"email": email})

    # 2. If the user does not exist, block the login
    if not raw_user:
        flash("You do not have an account. Please register first.", "danger")
        return redirect(url_for("main.login"))

    # 3. Optional: Check if the account was registered via Google previously
    # This prevents users from trying to log in with Google to an email 
    # that was registered via standard email/password (if you prefer).
    if raw_user.get("auth_provider") != "google":
        # You could also choose to update their profile here instead of blocking
        pass

    # 4. Proceed with Login
    user_obj = User(raw_user)
    login_user(user_obj, remember=True)
    
    flash("Successfully logged in with Google!", "success")
    return redirect(url_for("main.dashboard"))


@bp.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():

    if request.method == 'POST':
        email = request.form.get('email')

        if not email:
            flash("Email waa mandatory!", "danger")
            return redirect(url_for('main.forgot_password'))

        user = mongo.db.users.find_one({"email": email})

        flash("If this email is registered, you will receive password reset instructions shortly.", "info")

        if not user:
            return redirect(url_for('main.forgot_password'))

        # Generate OTP
        otp_code = str(random.randint(100000, 999999))

        # ✅ SAVE IN SESSION (IMPORTANT FIX)
        session['forgot_password_email'] = email
        session['forgot_password_otp'] = otp_code
        session['otp_created_at'] = datetime.now(pytz.timezone("Africa/Nairobi")).isoformat()

        try:
            send_otp_email(
                user_email=email,
                otp_code=otp_code,
                username=user.get("username")
            )

            flash("OTP sent to your email. Please check your inbox.", "success")

        except Exception as e:
            flash(f"Failed to send OTP email. ({str(e)})", "danger")

        return redirect(url_for('main.forgot_password_verify_otp'))

    return render_template('backend/auth/auth-reset-creative.html')


def send_otp_email(user_email, otp_code, username, reset_link=None,
                   sender_name=None, sender_email=None):

    try:
        # Defaults
        group_name = "My App"
        system_name = "My System"
        email_support = "support@example.com"

        # Current year
        current_year = datetime.now(pytz.timezone("Africa/Nairobi")).year

        # Sender info (safe fallback)
        sender_email = sender_email or current_app.config.get("MAIL_USERNAME")
        sender_name = sender_name or group_name
        sender_full = f"{sender_name} <{sender_email}>"

        # =====================================
        # 🔥 AUTO RESET LINK (IMPORTANT FIX)
        # =====================================
        if not reset_link:
            reset_link = url_for(
                'main.forgot_password_verify_otp',
                otp=otp_code,
                email=user_email,
                _external=True
            )

        # Create message
        msg = Message(
            subject=f"{group_name} - OTP Verification Code",
            sender=sender_full,
            recipients=[user_email]
        )

        # HTML body
        msg.html = render_template(
            "backend/auth/auth-sms-verify.html",
            otp_code=otp_code,
            username=username,
            email=user_email,
            reset_link=reset_link,
            current_year=current_year,
            group_name=group_name,
            system_name=system_name,
            email_support=email_support
        )

        # Send email
        mail.send(msg)

        print("✅ OTP email sent successfully")

    except Exception as e:
        print(f"❌ Error sending OTP email: {str(e)}")
        raise

@bp.route('/forgot-password/verify-otp', methods=['GET', 'POST'])
def forgot_password_verify_otp():

    if current_user.is_authenticated:
        flash("You are already logged in.", "info")
        return redirect(url_for("main.dashboard"))

    # ===============================
    # 🔥 AUTO-FILL FROM EMAIL LINK
    # ===============================
    url_otp = request.args.get('otp')
    url_email = request.args.get('email')

    if url_email:
        session['forgot_password_email'] = url_email

    if url_otp:
        session['forgot_password_otp'] = url_otp

    # ===============================
    # SESSION CHECK
    # ===============================
    email = session.get('forgot_password_email')
    saved_otp = session.get('forgot_password_otp')
    otp_created_at = session.get('otp_created_at')

    if not email or not saved_otp:
        flash("Session expired. Please start the password reset again.", "error")
        return redirect(url_for('main.forgot_password'))

    # ===============================
    # POST VERIFY OTP
    # ===============================
    if request.method == 'POST':
        input_otp = request.form.get('otp_code')

        if not input_otp:
            flash("OTP is required!", "error")
            return redirect(url_for('main.forgot_password_verify_otp'))

        # OTP expiry check
        if otp_created_at:
            try:
                otp_time = datetime.fromisoformat(otp_created_at)
            except:
                session.clear()
                flash("Session error. Please request OTP again.", "error")
                return redirect(url_for('main.forgot_password'))

            current_time = datetime.now(pytz.timezone("Africa/Nairobi"))

            if otp_time.tzinfo is None:
                otp_time = pytz.timezone("Africa/Nairobi").localize(otp_time)

            if current_time - otp_time > timedelta(minutes=5):
                session.clear()
                flash("OTP expired. Please request a new password reset.", "error")
                return redirect(url_for('main.forgot_password'))

        # OTP validation
        if str(input_otp).strip() == str(saved_otp).strip():

            user = mongo.db.users.find_one({"email": email})

            if not user:
                flash("User not found for this email.", "error")
                return redirect(url_for('main.forgot_password'))

            session['forgot_password_verified_email'] = email

            session.pop('forgot_password_otp', None)
            session.pop('otp_created_at', None)

            flash("OTP verified successfully. Please change your password.", "success")
            return redirect(url_for('main.forgot_password_change_password'))

        else:
            flash("Invalid OTP. Please try again.", "error")
            return redirect(url_for('main.forgot_password_verify_otp'))

    # ===============================
    # GET PAGE RENDER (AUTO OTP PASS)
    # ===============================
    return render_template(
        'backend/auth/auth-verify-creative.html',
        email=email,
        auto_otp=url_otp
    )


@bp.route('/forgot-password/change-password', methods=['GET', 'POST'])
def forgot_password_change_password():

    # =========================
    # SESSION CHECK
    # =========================
    email = session.get('forgot_password_verified_email')

    if not email:
        flash("Session expired. Please start the password reset again.", "error")
        return redirect(url_for('main.forgot_password'))

    # =========================
    # POST (CHANGE PASSWORD)
    # =========================
    if request.method == 'POST':

        new_password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')

        if not new_password or not confirm_password:
            flash("All fields are required!", "error")
            return redirect(url_for('main.forgot_password_change_password'))

        if new_password != confirm_password:
            flash("Passwords do not match!", "error")
            return redirect(url_for('main.forgot_password_change_password'))

        if len(new_password) < 6:
            flash("Password must be at least 6 characters!", "error")
            return redirect(url_for('main.forgot_password_change_password'))

        # =========================
        # HASH PASSWORD
        # =========================
        from werkzeug.security import generate_password_hash

        hashed_password = generate_password_hash(new_password)

        # =========================
        # UPDATE MONGO USER
        # =========================
        mongo.db.users.update_one(
            {"email": email},
            {"$set": {
                "password": hashed_password
            }}
        )

        # =========================
        # CLEAN SESSION
        # =========================
        session.pop('forgot_password_email', None)
        session.pop('forgot_password_verified_email', None)
        session.pop('forgot_password_otp', None)
        session.pop('otp_created_at', None)

        flash("Password changed successfully. Please login.", "success")
        return redirect(url_for('main.login'))

    # =========================
    # GET PAGE
    # =========================
    return render_template('backend/auth/auth-change-password.html', email=email)



@bp.route("/dashboard")
@login_required
def dashboard():

    # =========================
    # ROLE PROTECTION
    # =========================
    if current_user.role not in ["superadmin", "admin"]:
        abort(403)

    # =========================
    # USER FILTER LOGIC
    # =========================
    try:
        user_id = ObjectId(current_user.id)
    except:
        user_id = current_user.id

    if current_user.role == "superadmin":
        user_filter = {}
    else:
        user_filter = {"user_id": user_id}

    # =====================================
    # DEFAULT CHART VARIABLES
    # =====================================

    chart_labels = []
    chart_income = []
    chart_expense = []
    chart_profit = []

    account_cards = []

    # =========================
    # DATA
    # =========================
    accounts = list(mongo.db.accounts.find({**user_filter, "status": True}))
    transactions = list(mongo.db.transactions.find(user_filter))
    savings = list(mongo.db.savings.find(user_filter))
    categories = list(mongo.db.categories.find({**user_filter, "status": True}))

    recent_users = list(
        mongo.db.users.find(user_filter, {"password": 0})
        .sort("created_at", -1)
        .limit(5)
    )

    leads = {
        "New": 0,
        "Working": 0,
        "Others": 0
    }

    for u in recent_users:
        status = u.get("auth_status", "New")

        if status == "login":
            leads["Working"] += 1
        elif status == "logout":
            leads["New"] += 1
        else:
            leads["Others"] += 1

    
    # =========================
    # SAFE FLOAT
    # =========================
    def safe_float(v):
        try:
            return float(v)
        except:
            return 0.0

    # =========================
    # CATEGORY CHART (FIXED LOGIC)
    # =========================
    category_totals = defaultdict(float)

    
    for t in transactions:
        # ❌ ONLY EXPENSE
        if t.get("transaction_type") != "expense":
            continue

        amount = float(t.get("amount", 0))
        category = t.get("category")

        if not category:
            category = "Unknown"

        category_totals[category] += amount


    # REMOVE ZERO VALUES (IMPORTANT FIX)
    category_totals = {
        k: v for k, v in category_totals.items() if v > 0
    }

    category_labels = list(category_totals.keys())
    category_values = list(category_totals.values())

    # =========================
    # TOTALS
    # =========================
    total_balance = sum(safe_float(a.get("balance")) for a in accounts)

    total_income = sum(
        safe_float(t.get("amount"))
        for t in transactions
        if t.get("transaction_type") == "income"
    )

    total_expense = sum(
        safe_float(t.get("amount"))
        for t in transactions
        if t.get("transaction_type") == "expense"
    )

    total_savings = sum(safe_float(s.get("current_balance")) for s in savings)

    # =========================
    # RECENT
    # =========================
    recent_transactions = sorted(
        transactions,
        key=lambda x: x.get("date") or datetime.utcnow(),
        reverse=True
    )[:10]

    active_savings = [s for s in savings if s.get("status") == "active"]

    dashboard = {
        "balance": total_balance,
        "income": total_income,
        "expense": total_expense,
        "savings": total_savings,
        "accounts": len(accounts),
        "categories": len(categories),
        "transactions": len(transactions),
    }

    if current_user.role == "superadmin":
        tx_filter = {}
    else:
        try:
            user_id = ObjectId(current_user.id)
        except:
            user_id = current_user.id

        tx_filter = {"user_id": user_id}

    latest_transactions = list(
        mongo.db.transactions.find(tx_filter)
        .sort("created_at", -1)
        .limit(10)
    )

    if current_user.role == "superadmin":
        st_filter = {}
    else:
        try:
            user_id = ObjectId(current_user.id)
        except:
            user_id = current_user.id

        st_filter = {"user_id": user_id}

    latest_saving_transactions = list(
        mongo.db.saving_transactions.find(st_filter)
        .sort("created_at", -1)
        .limit(10)
    )

    # Ku dar magaca Saving iyo Account
    for tx in latest_saving_transactions:

        # Saving Name
        saving = mongo.db.savings.find_one({
            "_id": ObjectId(tx["saving_id"])
        }) if tx.get("saving_id") else None

        tx["saving_name"] = saving["title"] if saving else "N/A"

        # Account Name
        account = mongo.db.accounts.find_one({
            "_id": ObjectId(tx["account_id"])
        }) if tx.get("account_id") else None

        tx["account_name"] = account["name"] if account else "N/A"
        

    # =====================================
    # SAVING PROGRESS
    # =====================================

    saving_progress=[]


    for saving in savings:


        target=float(
            saving.get("target_amount",0)
        )


        current=float(
            saving.get("current_balance",0)
        )


        percentage=0


        if target > 0:

            percentage=round(
                (current/target)*100,
                2
            )


        saving_progress.append({

            "title":saving.get("title"),

            "target":target,

            "current":current,

            "percentage":percentage

        })



    # ==========================================
    # SAVING GOALS REPORT
    # ==========================================

    saving_goals = []

    saving_warnings = []


    savings = list(
        mongo.db.savings.find({
            "$or": [
                {
                    "user_id": str(current_user.id)
                },
                {
                    "user_id": ObjectId(current_user.id)
                }
            ]
        })
    )



    for saving in savings:


        target = float(
            saving.get(
                "target_amount",
                0
            )
        )


        saved = float(
            saving.get(
                "current_balance",
                0
            )
        )


        remaining = target - saved


        progress = 0

        if target > 0:

            progress = round(
                (saved / target) * 100,
                2
            )



        # ===============================
        # STATUS
        # ===============================

        if progress >= 100:

            status = "Completed"

        elif progress >= 50:

            status = "Good Progress"

        elif progress > 0:

            status = "In Progress"

        else:

            status = "Not Started"



        # ===============================
        # SAVING WARNINGS
        # ===============================

        if saved > target:


            saving_warnings.append({

                "type": "success",

                "title": saving.get(
                    "title",
                    "Saving"
                ),

                "message":
                f"Saving exceeded target by ${saved-target:,.2f}"

            })



        elif progress >= 90 and progress < 100:


            saving_warnings.append({

                "type": "warning",

                "title": saving.get(
                    "title",
                    "Saving"
                ),

                "message":
                f"Almost completed. Only ${remaining:,.2f} remaining."

            })



        elif progress < 25 and target > 0:


            saving_warnings.append({

                "type": "danger",

                "title": saving.get(
                    "title",
                    "Saving"
                ),

                "message":
                f"Saving progress is low ({progress}%)."

            })



        # ===============================
        # REPORT DATA
        # ===============================

        saving_goals.append({

            "_id": str(
                saving.get("_id")
            ),

            "title":
            saving.get(
                "title",
                "Saving Goal"
            ),


            "target":
            target,


            "saved":
            saved,


            "remaining":
            max(
                remaining,
                0
            ),


            "progress":
            progress,


            "status":
            status,


            "is_over_target":
            saved > target

        })


    # ==========================================
    # SAVING DEPOSIT REPORT
    # ==========================================

    saving_deposit_report = []


    saving_transactions = list(
        mongo.db.saving_transactions.find({
            "$or": [
                {
                    "user_id": str(current_user.id)
                },
                {
                    "user_id": ObjectId(current_user.id)
                }
            ],
            "status": True
        })
    )



    for goal in saving_goals:

        saving_id = goal["_id"]

        total_deposit = 0
        total_withdrawal = 0
        deposit_count = 0


        for st in saving_transactions:

            if str(st.get("saving_id")) != str(saving_id):
                continue


            amount = float(
                st.get("amount",0)
            )


            if st.get("transaction_type") == "deposit":

                total_deposit += amount
                deposit_count += 1


            elif st.get("transaction_type") == "withdrawal":

                total_withdrawal += amount



        growth = (
            total_deposit -
            total_withdrawal
        )


        saving_deposit_report.append({

            "title": goal["title"],

            "target": goal["target"],

            "current": goal["saved"],

            "total_deposit": total_deposit,

            "total_withdrawal": total_withdrawal,

            "growth": growth,

            "deposit_count": deposit_count,

            "remaining": goal["remaining"],

            "progress": goal["progress"]

        })



    # =====================================
    # ACCOUNT CARDS
    # =====================================

        account_cards = []


        for account in accounts:

            account_cards.append({

                "name": account.get("name"),

                "type": account.get("type"),

                "balance": float(
                    account.get("balance",0)
                )

            })




# =====================================
# MONTHLY CHART DATA
# =====================================


    monthly_income = defaultdict(float)

    monthly_expense = defaultdict(float)






    for t in transactions:

        date = t.get("date")


        if not date:
            continue


        month = date.strftime("%b")


        amount = float(
            t.get("amount",0)
        )


        if t.get("transaction_type") == "income":

            monthly_income[month] += amount



        elif t.get("transaction_type") == "expense":

            monthly_expense[month] += amount




    chart_labels = sorted(

        list(

            set(monthly_income.keys())

            |

            set(monthly_expense.keys())

        )

    )



    chart_income = [

        round(
            monthly_income.get(month,0),
            2
        )

        for month in chart_labels

    ]



    chart_expense = [

        round(
            monthly_expense.get(month,0),
            2
        )

        for month in chart_labels

    ]



    chart_profit = [

        round(

            monthly_income.get(month,0)

            -

            monthly_expense.get(month,0),

            2

        )

        for month in chart_labels

    ]


    return render_template(
        "backend/home/dashboard.html",
        dashboard=dashboard,
        accounts=accounts,
        categories=categories,
        transactions=recent_transactions,
        savings=active_savings,
        user=current_user,
        category_labels=category_labels,
        category_values=category_values,
        recent_users=recent_users,
        leads=leads ,
        latest_transactions=latest_transactions,
        saving_transactions=latest_saving_transactions,

        saving_progress=saving_progress,

        
        chart_labels=chart_labels,

chart_income=chart_income,

chart_expense=chart_expense,

chart_profit=chart_profit,


account_cards=account_cards,



saving_goals=saving_goals,
saving_warnings=saving_warnings,

saving_deposit_report=saving_deposit_report,
    )


@bp.route("/saving-goals/view/<id>")
@login_required
def view_saving_goal(id):

    try:

        saving_id = ObjectId(id)

    except Exception:

        abort(404)



    user_id = str(current_user.id)



    # ==================================
    # GET SAVING
    # COLLECTION = savings
    # ==================================

    saving = mongo.db.savings.find_one({

        "_id": saving_id,

        "$or": [

            {
                "user_id": user_id
            },

            {
                "user_id": ObjectId(user_id)
            }

        ]

    })



    print("==============================")
    print("Saving ID:", saving_id)
    print("User:", user_id)
    print("Saving:", saving)
    print("==============================")



    if not saving:


        flash(
            "Saving goal not found.",
            "danger"
        )


        return redirect(
            url_for(
                "main.dashboard"
            )
        )




    # ==================================
    # CALCULATE SAVING
    # ==================================


    target = float(
        saving.get(
            "target_amount",
            0
        )
    )


    saved = float(
        saving.get(
            "current_balance",
            0
        )
    )



    remaining = max(
        target - saved,
        0
    )



    progress = 0


    if target > 0:

        progress = round(
            (saved / target) * 100,
            2
        )



    if progress >= 100:

        status = "Completed"


    elif progress >= 50:

        status = "Good Progress"


    elif progress > 0:

        status = "In Progress"


    else:

        status = "Not Started"




    # ==================================
    # TRANSACTIONS
    # ==================================


    transactions = list(

        mongo.db.saving_transactions.find({

            "$or":[

                {
                    "saving_id": saving_id
                },

                {
                    "saving_id": str(saving_id)
                }

            ],

            "status": True

        })

        .sort(
            "date",
            -1
        )

    )




    # ==================================
    # SUMMARY
    # ==================================


    total_deposit = 0

    total_withdrawal = 0



    for trx in transactions:


        amount = float(
            trx.get(
                "amount",
                0
            )
        )


        if trx.get(
            "transaction_type"
        ) == "deposit":


            total_deposit += amount



        elif trx.get(
            "transaction_type"
        ) == "withdrawal":


            total_withdrawal += amount





    growth = (
        total_deposit -
        total_withdrawal
    )




    # ==================================
    # TEMPLATE DATA
    # ==================================


    goal = {


        "_id":
        str(
            saving["_id"]
        ),


        "title":
        saving.get(
            "title",
            "Saving Goal"
        ),


        "description":
        saving.get(
            "description",
            ""
        ),


        "target":
        target,


        "saved":
        saved,


        "remaining":
        remaining,


        "progress":
        progress,


        "status":
        status,


        "maturity_date":
        saving.get(
            "maturity_date"
        ),


        "created_at":
        saving.get(
            "created_at"
        )

    }



    return render_template(

        "backend/pages/components/savings/view.html",
        goal=goal,
        transactions=transactions,
        total_deposit=total_deposit,
        total_withdrawal=total_withdrawal,
        growth=growth

    )






@bp.route("/profile")
@login_required
def profile():
    return render_template(
        "backend/pages/components/users/profile.html",
        user=current_user
    )


@bp.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "POST":

        old_password = request.form.get("old_password")
        new_password = request.form.get("new_password")
        confirm_password = request.form.get("confirm_password")

        # 1. check passwords match
        if new_password != confirm_password:
            flash("Passwords do not match", "danger")
            return redirect(url_for("main.change_password"))

        # 2. get user from DB
        user = mongo.db.users.find_one({"_id": ObjectId(current_user.id)})

        # 3. verify old password
        if not check_password_hash(user["password"], old_password):
            flash("Old password is incorrect", "danger")
            return redirect(url_for("main.change_password"))

        # 4. update password
        hashed_password = generate_password_hash(new_password)

        mongo.db.users.update_one(
            {"_id": ObjectId(current_user.id)},
            {"$set": {"password": hashed_password}}
        )

        flash("Password changed successfully", "success")
        return redirect(url_for("main.dashboard"))

    return render_template("backend/pages/components/users/change_password.html")

@bp.route("/account-settings", methods=["GET", "POST"])
@login_required
def account_settings():

    if request.method == "POST":

        data = {
            "fullname": request.form.get("fullname"),
            "username": request.form.get("username"),
            "phone": request.form.get("phone"),
            "country": request.form.get("country"),
            "state": request.form.get("state"),
            "city": request.form.get("city"),
            "address": request.form.get("address"),
            "bio": request.form.get("bio"),
            "updated_at": datetime.utcnow()
        }

        file = request.files.get("photo")

        if file and file.filename:

            upload_result = cloudinary.uploader.upload(file, folder="users")

            data["photo"] = upload_result["secure_url"]  # 🔥 IMPORTANT

        mongo.db.users.update_one(
            {"_id": ObjectId(current_user.id)},
            {"$set": data}
        )

        flash("Account updated successfully.", "success")
        return redirect(url_for("main.account_settings"))

    return render_template(
        "backend/pages/components/users/account_settings.html",
        user=current_user
    )



@bp.route('/add-user', methods=['GET', 'POST'])
@login_required
def add_user():

    if current_user.role not in ['superadmin', 'admin']:
        return abort(403)

    countries = [
        {"code": "SO", "name": "Somalia", "flag_url": "https://flagcdn.com/so.svg"},
        {"code": "KE", "name": "Kenya", "flag_url": "https://flagcdn.com/ke.svg"},
    ]

    if request.method == 'POST':

        fullname = request.form.get('fullname')
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')
        role = request.form.get('role') or "user"
        country = request.form.get('country')
        phone = request.form.get('phone')
        state = request.form.get('state')
        city = request.form.get('city')
        address = request.form.get('address')
        status = True if request.form.get('status') == '1' else False

        # ================= VALIDATION =================
        if not email or not username or not fullname:
            flash("Fadlan buuxi fields-ka muhiimka ah!", "danger")
            return redirect(url_for('main.add_user'))

        if password != confirm_password:
            flash("Passwords-ka isma laha!", "danger")
            return redirect(url_for('main.add_user'))

        if mongo.db.users.find_one({"email": email}):
            flash("Email-kan horey ayaa loo isticmaalay!", "danger")
            return redirect(url_for('main.add_user'))

        if mongo.db.users.find_one({"username": username}):
            flash("Username-kan horey ayaa loo isticmaalay!", "danger")
            return redirect(url_for('main.add_user'))

        # ================= PHOTO UPLOAD =================
        photo_path = None

        file = request.files.get('photo')

        if file and file.filename:

            project_root = os.path.abspath(os.getcwd())

            upload_dir = os.path.join(
                project_root,
                'static',
                'backend',
                'uploads',
                'users'
            )

            os.makedirs(upload_dir, exist_ok=True)

            filename = f"{uuid.uuid4().hex}_{secure_filename(file.filename)}"
            file_path = os.path.join(upload_dir, filename)

            file.save(file_path)

            photo_path = f"backend/uploads/users/{filename}"

        # ================= CREATE USER =================
        new_user = {
            "fullname": fullname,
            "username": username,
            "email": email,
            "password": generate_password_hash(password),
            "role": role,
            "country": country,
            "phone": phone,
            "state": state,
            "city": city,
            "address": address,
            "status": status,
            "photo": photo_path,
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow()
        }

        mongo.db.users.insert_one(new_user)

        flash(f"User {username} si guul leh ayaa loo diiwaangeliyey!", "success")
        return redirect(url_for('main.add_user'))

    return render_template(
        "backend/pages/components/users/add_user.html",
        countries=countries
    )


@bp.route('/edit-user/<user_id>', methods=['GET', 'POST'])
@login_required
def edit_user(user_id):

    if current_user.role not in ['superadmin', 'admin']:
        return abort(403)

    try:
        raw_user = mongo.db.users.find_one({"_id": ObjectId(user_id)})
    except Exception:
        flash("Invalid user ID!", "danger")
        return redirect(url_for('main.index'))

    if not raw_user:
        flash("User-ka lama helin!", "danger")
        return redirect(url_for('main.index'))

    user = User(raw_user)

    if request.method == 'POST':

        fullname = request.form.get('fullname')
        username = request.form.get('username')
        email = request.form.get('email')
        role = request.form.get('role')
        country = request.form.get('country')
        phone = request.form.get('phone')
        address = request.form.get('address')
        bio = request.form.get('bio')
        status = True if request.form.get('status') == '1' else False

        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')

        # ================= VALIDATION =================
        if mongo.db.users.find_one({
            "username": username,
            "_id": {"$ne": ObjectId(user_id)}
        }):
            flash("Username-kan horey ayaa loo isticmaalay!", "danger")
            return redirect(url_for('main.edit_user', user_id=user_id))

        if mongo.db.users.find_one({
            "email": email,
            "_id": {"$ne": ObjectId(user_id)}
        }):
            flash("Email-kan horey ayaa loo isticmaalay!", "danger")
            return redirect(url_for('main.edit_user', user_id=user_id))

        updated_data = {
            "fullname": fullname,
            "username": username,
            "email": email,
            "role": role,
            "country": country,
            "phone": phone,
            "address": address,
            "bio": bio,
            "status": status,
            "updated_at": datetime.utcnow()
        }

        # ================= PASSWORD =================
        if password:
            if password != confirm_password:
                flash("Passwords-ka isma laha!", "danger")
                return redirect(url_for('main.edit_user', user_id=user_id))

            updated_data["password"] = generate_password_hash(password)

        # ================= CLOUDINARY PHOTO =================
        file = request.files.get('photo')

        if file and file.filename:

            old_public_id = raw_user.get("photo_public_id")

            # delete old image
            if old_public_id:
                try:
                    cloudinary.uploader.destroy(old_public_id)
                except Exception:
                    pass

            # upload new image
            result = cloudinary.uploader.upload(
                file,
                folder="users"
            )

            updated_data["photo"] = result["secure_url"]
            updated_data["photo_public_id"] = result["public_id"]

        # ================= UPDATE DB =================
        mongo.db.users.update_one(
            {"_id": ObjectId(user_id)},
            {"$set": updated_data}
        )

        flash("User si guul leh ayaa loo cusbooneysiiyey!", "success")
        return redirect(url_for('main.edit_user', user_id=user_id))

    return render_template(
        "backend/pages/components/users/edit_user.html",
        user=user
    )


@bp.route('/delete-user/<user_id>', methods=['POST'])
@login_required
def delete_user(user_id):

    if current_user.role not in ['superadmin', 'admin']:
        return abort(403)

    try:
        user = mongo.db.users.find_one({
            "_id": ObjectId(user_id)
        })
    except Exception:
        flash("Invalid user ID!", "danger")
        return redirect(url_for('main.all_users'))

    if not user:
        flash("User-ka lama helin!", "danger")
        return redirect(url_for('main.all_users'))

    # ==========================================
    # DELETE USER PHOTO FROM CLOUDINARY
    # ==========================================

    photo_public_id = user.get("photo_public_id")

    if photo_public_id:
        try:
            cloudinary.uploader.destroy(photo_public_id)
        except Exception:
            pass

    # ==========================================
    # FIND ALL USER ORDERS
    # ==========================================

    orders = list(
        mongo.db.orders.find({
            "user_id": ObjectId(user_id)
        })
    )

    # ==========================================
    # RESTORE PRODUCT STOCK
    # ==========================================

    for order in orders:

        for item in order.get("items", []):

            try:
                mongo.db.products.update_one(
                    {
                        "_id": ObjectId(item["product_id"])
                    },
                    {
                        "$inc": {
                            "stock": int(item["qty"])
                        }
                    }
                )
            except Exception:
                pass

    # ==========================================
    # DELETE ALL CUSTOMER ORDERS
    # ==========================================

    mongo.db.orders.delete_many({
        "user_id": ObjectId(user_id)
    })

    # ==========================================
    # DELETE USER
    # ==========================================

    mongo.db.users.delete_one({
        "_id": ObjectId(user_id)
    })

    flash(
        "Customer, orders-kiisii iyo payments-kiisii si guul leh ayaa loo tirtiray!",
        "success"
    )

    return redirect(url_for('main.all_users'))



@bp.route('/all-users', methods=['GET'])
@login_required
def all_users():

    if current_user.role not in ['superadmin', 'admin']:
        return abort(403)

    if current_user.role == 'superadmin':
        # Superadmin sees everyone
        users_cursor = mongo.db.users.find().sort('created_at', -1)

    else:  # admin
        # Admin cannot see superadmins
        users_cursor = mongo.db.users.find(
            {"role": {"$ne": "superadmin"}}
        ).sort('created_at', -1)

    users = [User(user_data) for user_data in users_cursor]

    return render_template(
        'backend/pages/components/users/all_users.html',
        users=users
    )





# ==========================================
# CALCULATE HEALTHY WEIGHT
# ==========================================

def calculate_goal_weight(height_cm):

    height_m = height_cm / 100

    healthy_min = 18.5 * (height_m ** 2)

    healthy_max = 24.9 * (height_m ** 2)

    goal = (healthy_min + healthy_max) / 2


    return (
        round(goal, 1),
        round(healthy_min, 1),
        round(healthy_max, 1)
    )





# ==========================================
# BMI CALCULATOR
# ==========================================

def calculate_bmi(weight, height_cm):

    try:

        weight = float(weight)
        height_cm = float(height_cm)


        if weight <= 0 or height_cm <= 0:
            return 0


        # CM TO METER
        height_m = height_cm / 100


        bmi = weight / (height_m * height_m)


        return round(bmi, 2)


    except:

        return 0



def bmi_status(bmi):

    if bmi < 18.5:

        return "Under Weight"


    elif bmi < 25:

        return "Normal Weight"


    elif bmi < 30:

        return "Over Weight"


    else:

        return "Obesity"





@bp.route("/gym/checkin", methods=["GET"])
@login_required
def gym_checkin():

    user_id = ObjectId(current_user.id)

    today = datetime.utcnow().date()


    # =====================================
    # PROFILE
    # =====================================

    profile = mongo.db.gym_profile.find_one({

        "user_id": user_id

    })


    if not profile:

        flash(
            "Please create Gym Profile first.",
            "warning"
        )

        return redirect(
            url_for("main.gym_profile")
        )



    # =====================================
    # WEIGHT DATA
    # =====================================

    latest_weight = mongo.db.weight_progress.find_one(

        {
            "user_id": user_id
        },

        sort=[
            ("date",-1)
        ]

    )


    if latest_weight:

        current_weight = float(
            latest_weight.get(
                "weight",
                0
            )
        )

    else:

        current_weight = float(
            profile.get(
                "start_weight",
                0
            )
        )



    start_weight = float(

        profile.get(
            "start_weight",
            current_weight
        )

    )



    # =====================================
    # WEIGHT RESULT
    # =====================================

    if current_weight > start_weight:


        net_gain = round(
            current_weight - start_weight,
            2
        )

        net_loss = 0


    elif current_weight < start_weight:


        net_loss = round(
            start_weight-current_weight,
            2
        )

        net_gain = 0


    else:

        net_gain = 0
        net_loss = 0





    # =====================================
    # STREAK DATA
    # =====================================

    streak = mongo.db.gym_streak.find_one({

        "user_id":user_id

    })


    if not streak:

        streak = {

            "current_streak":0,

            "best_streak":0,

            "total_checkins":0,

            "start_date":None

        }





    # =====================================
    # CHECKIN HISTORY
    # =====================================


    checkins = list(

        mongo.db.gym_checkins.find({

            "user_id":user_id

        })

    )



    checked_dates = {}



    for item in checkins:


        check_date = item.get(
            "date"
        )


        if check_date:


            checked_dates[check_date.date()] = {

                "time":check_date.strftime(
                    "%H:%M"
                )

            }






    # =====================================
    # JOURNEY START DATE
    # =====================================


    journey_start = profile.get(
        "created_at"
    )



    if not journey_start:

        journey_start = streak.get(
            "start_date"
        )



    if not journey_start:

        journey_start = datetime.utcnow()



    journey_start = journey_start.date()






    # =====================================
    # DAILY ATTENDANCE REPORT
    # =====================================


    calendar_days = []


    present_days = 0

    absent_days = 0



    total_days = (

        today -

        journey_start

    ).days + 1




    for i in range(total_days):


        day = journey_start + timedelta(
            days=i
        )



        if day in checked_dates:


            checked = True

            status = "Checked"

            icon = "☑️"

            checkin_time = checked_dates[day]["time"]

            present_days += 1



        else:


            checked = False

            status = "Absent"

            icon = "⬜"

            checkin_time = "-"

            absent_days += 1





        calendar_days.append({


            "date":day.strftime(
                "%Y-%m-%d"
            ),


            "day_name":calendar.day_name[
                day.weekday()
            ],


            "checked":checked,


            "icon":icon,


            "status":status,


            "checkin_time":checkin_time


        })







    # =====================================
    # DAYS SINCE START
    # =====================================


    days_since_start = (

        today -

        journey_start

    ).days + 1





    # =====================================
    # MOTIVATION
    # =====================================


    current = streak.get(

        "current_streak",

        0

    )


    if current >= 60:


        message = "🔥 Legendary! 60+ days consistency!"



    elif current >= 30:


        message = "👑 Amazing! 30 days completed!"



    elif current >=14:


        message = "🚀 Two weeks strong! Keep pushing."



    elif current >=7:


        message = "🏆 One week completed!"



    elif current >=3:


        message = "🔥 Great consistency. Keep going!"



    else:


        message = "💪 Every workout counts. Stay focused."






    # =====================================
    # TEMPLATE
    # =====================================


    return render_template(

        "backend/pages/components/gym/checkin_report.html",


        profile=profile,


        calendar_days=calendar_days,


        journey_start=journey_start,


        total_days=total_days,


        present_days=present_days,


        absent_days=absent_days,


        current_streak=streak.get(
            "current_streak",
            0
        ),


        best_streak=streak.get(
            "best_streak",
            0
        ),


        total_checkins=streak.get(
            "total_checkins",
            0
        ),


        start_date=streak.get(
            "start_date"
        ),


        days_since_start=days_since_start,


        start_weight=start_weight,


        current_weight=current_weight,


        net_gain=net_gain,


        net_loss=net_loss,


        message=message

    )


@bp.route("/gym/checkin/manual", methods=["POST"])
@login_required
def gym_checkin_manual():

    user_id = ObjectId(current_user.id)


    selected_dates = request.form.getlist(
        "checkin_dates"
    )


    if not selected_dates:

        flash(
            "Please select at least one day.",
            "warning"
        )

        return redirect(
            url_for("main.gym_checkin")
        )



    # =====================================
    # GET PROFILE START DATE
    # =====================================

    profile = mongo.db.gym_profile.find_one({

        "user_id": user_id

    })


    if not profile:

        flash(
            "Please create Gym Profile first.",
            "warning"
        )

        return redirect(
            url_for("main.gym_profile")
        )



    start_date = profile.get(
        "created_at"
    )


    if not start_date:

        start_date = datetime.utcnow()



    start_date = start_date.date()



    weight = float(

        profile.get(
            "start_weight",
            0
        )

    )



    saved = 0



    # =====================================
    # SAVE SELECTED DAYS
    # =====================================

    for date_string in selected_dates:


        check_date = datetime.strptime(

            date_string,

            "%Y-%m-%d"

        )



        check_day = check_date.date()



        # ================================
        # BLOCK BEFORE PROFILE DATE
        # ================================

        if check_day < start_date:


            continue




        # ================================
        # DUPLICATE CHECK
        # ================================

        exists = mongo.db.gym_checkins.find_one({

            "user_id": user_id,


            "date": check_date

        })



        if not exists:


            mongo.db.gym_checkins.insert_one({

                "user_id": user_id,

                "date": check_date,

                "weight": weight,

                "created_at": datetime.utcnow()

            })


            saved += 1




    if saved > 0:


        flash(

            f"✅ {saved} gym attendance days saved.",

            "success"

        )


    else:


        flash(

            "No new attendance saved.",

            "info"

        )



    return redirect(

        url_for(
            "main.gym_checkin"
        )

    )



@bp.route("/gym/calendar")
@login_required
def gym_calendar():

    user_id = ObjectId(current_user.id)

    now = datetime.utcnow()

    today = now.date()

    # =====================================
    # PROFILE
    # =====================================

    profile = mongo.db.gym_profile.find_one({
        "user_id": user_id
    })

    if not profile:

        flash(
            "Please create Gym Profile first.",
            "warning"
        )

        return redirect(
            url_for("main.gym_profile")
        )

    # =====================================
    # YEAR / MONTH
    # =====================================

    year = int(
        request.args.get(
            "year",
            now.year
        )
    )

    month = int(
        request.args.get(
            "month",
            now.month
        )
    )

    if month < 1:

        month = 12
        year -= 1

    elif month > 12:

        month = 1
        year += 1

    month_name = calendar.month_name[month]

    # =====================================
    # PROFILE START DATE
    # =====================================

    journey_start = profile.get("created_at")

    if not journey_start:

        streak = mongo.db.gym_streak.find_one({
            "user_id": user_id
        })

        if streak:

            journey_start = streak.get("start_date")

    if not journey_start:

        journey_start = now

    journey_start = journey_start.date()

    # =====================================
    # MONTH RANGE
    # =====================================

    first_day = datetime(
        year,
        month,
        1
    )

    if month == 12:

        next_month = datetime(
            year + 1,
            1,
            1
        )

    else:

        next_month = datetime(
            year,
            month + 1,
            1
        )

    # =====================================
    # CHECKINS
    # =====================================

    checkins = list(

        mongo.db.gym_checkins.find({

            "user_id": user_id,

            "date": {

                "$gte": first_day,

                "$lt": next_month

            }

        })

    )

    checked_days = {}

    for item in checkins:

        d = item.get("date")

        if d:

            checked_days[d.day] = d.strftime("%H:%M")

    # =====================================
    # BUILD CALENDAR
    # =====================================

    calendar_data = []

    for week in calendar.monthcalendar(year, month):

        row = []

        for day in week:

            if day == 0:

                row.append(None)

                continue

            current_date = datetime(
                year,
                month,
                day
            ).date()

            checked = day in checked_days

            if current_date < journey_start:

                status = "before"

            elif current_date > today:

                status = "future"

            elif checked:

                status = "checked"

            else:

                status = "absent"

            row.append({

                "day": day,

                "checked": checked,

                "status": status,

                "time": checked_days.get(day)

            })

        calendar_data.append(row)

    # =====================================
    # STREAK
    # =====================================

    streak = mongo.db.gym_streak.find_one({

        "user_id": user_id

    })

    if not streak:

        streak = {

            "current_streak": 0,

            "best_streak": 0,

            "total_checkins": 0

        }

    # =====================================
    # SUMMARY
    # =====================================

    total_gym_days = len(checked_days)

    total_month_days = 0
    absent_days = 0

    for week in calendar_data:

        for cell in week:

            if not cell:

                continue

            if cell["status"] in ["checked", "absent"]:

                total_month_days += 1

            if cell["status"] == "absent":

                absent_days += 1

    # =====================================
    # TEMPLATE
    # =====================================

    return render_template(

        "backend/pages/components/gym/gym_calendar.html",

        calendar_data=calendar_data,

        month_name=month_name,

        month=month,

        year=year,

        today=today,

        journey_start=journey_start,

        total_month_days=total_month_days,

        total_gym_days=total_gym_days,

        absent_days=absent_days,

        current_streak=streak.get(
            "current_streak",
            0
        ),

        best_streak=streak.get(
            "best_streak",
            0
        ),

        total_checkins=streak.get(
            "total_checkins",
            0
        )

    )


# ==========================================
# GYM PROFILE
# ==========================================

@bp.route("/gym/profile", methods=["GET","POST"])
@login_required
def gym_profile():

    user_id = ObjectId(current_user.id)


    profile = mongo.db.gym_profile.find_one({

        "user_id": user_id

    })



    # DEFAULT VALUES

    bmi = 0

    bmi_status_text = ""

    goal_weight = 0

    healthy_min = 0

    healthy_max = 0

    age = 0

    height_m = 0



    # ==========================
    # POST SAVE
    # ==========================

    if request.method == "POST":


        height_cm = float(
            request.form.get("height")
        )


        start_weight = float(
            request.form.get("start_weight")
        )


        birth_year = int(
            request.form.get("birth_year")
        )



        # ==========================
        # HEIGHT CONVERSION
        # ==========================

        height_m = round(

            height_cm / 100,

            2

        )



        # AUTO AGE

        current_year = datetime.utcnow().year

        age = current_year - birth_year



        gender = request.form.get(
            "gender"
        )


        activity_level = request.form.get(
            "activity_level"
        )



        # ==========================
        # GOAL WEIGHT
        # ==========================

        goal_weight, healthy_min, healthy_max = calculate_goal_weight(

            height_cm

        )



        # ==========================
        # BMI
        # ==========================

        bmi = calculate_bmi(

            start_weight,

            height_cm

        )


        bmi_status_text = bmi_status(

            bmi

        )



        data = {


            "user_id": user_id,


            # CM VALUE

            "height_cm": height_cm,


            # METER VALUE

            "height_m": height_m,


            "start_weight": start_weight,


            "goal_weight": goal_weight,


            "healthy_min": healthy_min,


            "healthy_max": healthy_max,


            "birth_year": birth_year,


            "age": age,


            "gender": gender,


            "activity_level": activity_level,


            "updated_at": datetime.utcnow()

        }




        if profile:


            mongo.db.gym_profile.update_one(

                {
                    "_id": profile["_id"]
                },

                {
                    "$set": data
                }

            )


        else:


            data["created_at"] = datetime.utcnow()


            mongo.db.gym_profile.insert_one(

                data

            )




        flash(

            "Gym profile saved successfully.",

            "success"

        )


        return redirect(

            url_for(

                "main.gym_profile"

            )

        )





    # ==========================
    # GET DISPLAY
    # ==========================


    if profile:



        height_cm = float(

            profile.get(

                "height_cm",

                0

            )

        )



        height_m = round(

            height_cm / 100,

            2

        )




        weight = float(

            profile.get(

                "start_weight",

                0

            )

        )



        if height_cm and weight:



            bmi = calculate_bmi(

                weight,

                height_cm

            )


            bmi_status_text = bmi_status(

                bmi

            )




        goal_weight, healthy_min, healthy_max = calculate_goal_weight(

            height_cm

        )




        birth_year = profile.get(

            "birth_year"

        )



        if birth_year:


            age = datetime.utcnow().year - birth_year







    return render_template(


        "backend/pages/components/gym/profile.html",


        profile=profile,


        bmi=bmi,


        bmi_status=bmi_status_text,


        goal_weight=goal_weight,


        healthy_min=healthy_min,


        healthy_max=healthy_max,


        age=age,


        height_cm=height_cm if profile else 0,


        height_m=height_m


    )


@bp.route("/gym/weight/add", methods=["GET", "POST"])
@login_required
def add_gym_weight():

    user_id = ObjectId(current_user.id)


    # =====================================
    # GET PROFILE
    # =====================================

    profile = mongo.db.gym_profile.find_one({
        "user_id": user_id
    })


    if not profile:

        flash(
            "Please create your Gym Profile first.",
            "warning"
        )

        return redirect(
            url_for("main.gym_profile")
        )



    # =====================================
    # LAST WEIGHT
    # =====================================

    last_weight = mongo.db.weight_progress.find_one(
        {
            "user_id": user_id
        },
        sort=[
            ("date",-1)
        ]
    )


    if last_weight:

        previous_weight = float(
            last_weight.get(
                "weight",
                0
            )
        )

    else:

        previous_weight = float(
            profile.get(
                "start_weight",
                0
            )
        )




    # =====================================
    # HEIGHT CONVERSION
    # =====================================

    height_cm = float(
        profile.get(
            "height_cm",
            0
        )
    )


    height_m = round(
        height_cm / 100,
        2
    ) if height_cm else 0




    # =====================================
    # BMI FUNCTION
    # =====================================

    def calculate_bmi_value(weight):

        if height_m <= 0:

            return 0

        return round(
            weight /
            (height_m ** 2),
            2
        )



    def get_bmi_status(value):

        if value < 18.5:

            return "Underweight"


        elif value < 25:

            return "Normal Weight"


        elif value < 30:

            return "Overweight"


        else:

            return "Obese"




    # =====================================
    # CURRENT BMI
    # =====================================

    bmi = calculate_bmi_value(
        previous_weight
    )


    bmi_status = get_bmi_status(
        bmi
    )




    # =====================================
    # POST SAVE
    # =====================================

    if request.method == "POST":


        weight = request.form.get(
            "weight",
            ""
        ).strip()


        date = request.form.get(
            "date"
        )


        note = request.form.get(
            "note",
            ""
        ).strip()


        record_type = request.form.get(
            "type",
            "weight_record"
        )



        if not weight:


            flash(
                "Weight is required.",
                "danger"
            )

            return redirect(
                request.url
            )



        try:

            weight = float(weight)

        except:


            flash(
                "Invalid weight.",
                "danger"
            )

            return redirect(
                request.url
            )



        if weight <= 0:


            flash(
                "Weight must be greater than zero.",
                "danger"
            )

            return redirect(
                request.url
            )




        # DATE

        if date:

            date = datetime.strptime(
                date,
                "%Y-%m-%d"
            )

        else:

            date = datetime.utcnow()




        # DUPLICATE CHECK

        exists = mongo.db.weight_progress.find_one({

            "user_id": user_id,

            "date": date

        })


        if exists:


            flash(
                "Weight already exists for this date.",
                "warning"
            )

            return redirect(
                request.url
            )




        # WEIGHT CHANGE

        difference = round(
            weight - previous_weight,
            2
        )


        if difference > 0:

            change_type = "Weight Gain"


        elif difference < 0:

            change_type = "Weight Loss"


        else:

            change_type = "No Change"




        # NEW BMI

        new_bmi = calculate_bmi_value(
            weight
        )


        new_bmi_status = get_bmi_status(
            new_bmi
        )




        # SAVE


        mongo.db.weight_progress.insert_one({

            "user_id":user_id,

            "type":record_type,

            "weight":weight,

            "previous_weight":previous_weight,

            "difference":difference,

            "change_type":change_type,


            "height_cm":height_cm,

            "height_m":height_m,


            "bmi":new_bmi,

            "bmi_status":new_bmi_status,


            "date":date,

            "note":note,


            "created_at":datetime.utcnow(),

            "updated_at":datetime.utcnow()

        })




        flash(
            "Weight record saved successfully.",
            "success"
        )


        return redirect(
            url_for(
                "main.weight_history"
            )
        )




    return render_template(

        "backend/pages/components/gym/add_weight.html",

        profile=profile,

        previous_weight=previous_weight,


        height_cm=height_cm,

        height_m=height_m,


        bmi=bmi,

        bmi_status=bmi_status,


        datetime=datetime

    )




@bp.route("/gym/weight/edit/<id>", methods=["GET","POST"])
@login_required
def edit_gym_weight(id):

    try:

        weight_id = ObjectId(id)

    except:

        abort(404)



    user_id = ObjectId(current_user.id)



    # ==========================
    # GET OLD RECORD
    # ==========================

    weight = mongo.db.weight_progress.find_one({

        "_id": weight_id,

        "user_id": user_id

    })



    if not weight:

        abort(404)





    if request.method == "POST":


        new_weight = request.form.get(
            "weight"
        )


        date = request.form.get(
            "date"
        )


        note = request.form.get(
            "note",
            ""
        )



        # ==========================
        # VALIDATION
        # ==========================


        if not new_weight:


            flash(
                "Weight is required.",
                "danger"
            )

            return redirect(
                request.url
            )



        try:

            new_weight = float(
                new_weight
            )


        except:


            flash(
                "Invalid weight.",
                "danger"
            )

            return redirect(
                request.url
            )




        if new_weight <= 0:


            flash(
                "Weight must be greater than zero.",
                "danger"
            )

            return redirect(
                request.url
            )





        # ==========================
        # DATE
        # ==========================


        if date:


            date = datetime.strptime(
                date,
                "%Y-%m-%d"
            )


        else:

            date = weight.get(
                "date",
                datetime.utcnow()
            )







        # ==========================
        # CALCULATE CHANGE
        # ==========================


        old_weight = float(
            weight.get(
                "weight",
                0
            )
        )



        difference = round(

            new_weight - old_weight,

            2

        )



        if difference > 0:


            change_type = "Weight Gain"



        elif difference < 0:


            change_type = "Weight Loss"



        else:


            change_type = "No Change"







        # ==========================
        # UPDATE
        # ==========================


        mongo.db.weight_progress.update_one(

            {

                "_id":weight_id,

                "user_id":user_id

            },


            {

                "$set":{


                    "weight":
                    new_weight,


                    "difference":
                    difference,


                    "change_type":
                    change_type,


                    "date":
                    date,


                    "note":
                    note,


                    "updated_at":
                    datetime.utcnow()


                }

            }

        )






        flash(

            "Weight updated successfully.",

            "success"

        )



        return redirect(

            url_for(
                "main.weight_history"
            )

        )







    return render_template(

        "backend/pages/components/gym/edit_weight.html",

        weight=weight

    )

@bp.route("/gym/weight/delete/<id>")
@login_required
def delete_gym_weight(id):

    mongo.db.weight_progress.delete_one({

        "_id":ObjectId(id),

        "user_id":ObjectId(current_user.id)

    })


    flash(
        "Weight deleted successfully.",
        "success"
    )


    return redirect(
        url_for("main.weight_history")
    )



@bp.route("/gym/weight-history")
@login_required
def weight_history():

    user_id = ObjectId(current_user.id)

    # ==========================
    # GET GYM PROFILE
    # ==========================
    profile = mongo.db.gym_profile.find_one({
        "user_id": user_id
    })

    if not profile:
        flash(
            "Please create your Gym Profile first.",
            "warning"
        )
        return redirect(
            url_for("main.gym_profile")
        )

    # ==========================
    # GET ALL WEIGHTS
    # Latest first
    # ==========================
    weights = list(
        mongo.db.weight_progress.find({
            "user_id": user_id
        }).sort("date", -1)
    )

    # ==========================
    # PROFILE DATA
    # ==========================
    start_weight = float(
        profile.get(
            "start_weight",
            0
        )
    )

    goal_weight = float(
        profile.get(
            "goal_weight",
            0
        )
    )

    # ==========================
    # CURRENT WEIGHT
    # ==========================
    if weights:

        current_weight = float(
            weights[0]["weight"]
        )

    else:

        current_weight = start_weight

    # ==========================
    # GAIN / LOSS
    # ==========================
    weight_difference = round(
        current_weight - start_weight,
        2
    )

    if weight_difference > 0:

        weight_status = "gain"

    elif weight_difference < 0:

        weight_status = "loss"

    else:

        weight_status = "same"

    # ==========================
    # PROGRESS %
    # ==========================

    progress_percent = 0

    if start_weight > goal_weight:

        # User wants to LOSE weight

        total = start_weight - goal_weight

        done = start_weight - current_weight

        if total > 0:

            progress_percent = round(
                (done / total) * 100,
                1
            )

    elif goal_weight > start_weight:

        # User wants to GAIN weight

        total = goal_weight - start_weight

        done = current_weight - start_weight

        if total > 0:

            progress_percent = round(
                (done / total) * 100,
                1
            )

    # Prevent bad values
    progress_percent = max(
        0,
        min(progress_percent, 100)
    )

    # ==========================
    # FIX OBJECTID
    # ==========================
    for item in weights:

        item["_id"] = str(item["_id"])

    # ==========================
    # TEMPLATE
    # ==========================
    return render_template(

        "backend/pages/components/gym/weight_history.html",

        profile=profile,

        weights=weights,

        current_weight=current_weight,

        start_weight=start_weight,

        goal_weight=goal_weight,

        weight_difference=abs(weight_difference),

        weight_status=weight_status,

        progress_percent=progress_percent

    )



@bp.route("/gym/dashboard")
@login_required
def gym_dashboard():

    profile = mongo.db.gym_profile.find_one({
        "user_id": ObjectId(current_user.id)
    })

    if not profile:

        flash("Please complete your gym profile first.", "warning")
        return redirect(url_for("main.gym_profile"))

    weights = list(
        mongo.db.weight_progress.find(
            {
                "user_id": ObjectId(current_user.id)
            }
        ).sort("date", 1)
    )

    current_weight = profile["start_weight"]

    if weights:
        current_weight = weights[-1]["weight"]

    start_weight = profile["start_weight"]
    goal_weight = profile["goal_weight"]
    height = profile["height_cm"]

    bmi = round(
        current_weight /
        ((height / 100) ** 2),
        1
    )

    lost = round(
        start_weight - current_weight,
        2
    )

    remaining = round(
        current_weight - goal_weight,
        2
    )

    total = start_weight - goal_weight

    progress = 0

    if total > 0:

        progress = round(
            (lost / total) * 100,
            2
        )

    progress = max(0, min(progress, 100))

    return render_template(
        "backend/pages/components/gym/dashboard.html",

        profile=profile,

        weights=weights,

        current_weight=current_weight,

        bmi=bmi,

        lost=lost,

        remaining=remaining,

        progress=progress
    )


@bp.route('/add-category', methods=['GET', 'POST'])
@login_required
def add_category():

    if request.method == "POST":

        # 🧼 CLEAN INPUT PROPERLY
        name = request.form.get("name", "").strip()
        category_type = request.form.get("type", "expense").strip()

        if not name:
            flash("Category name is required", "danger")
            return redirect(url_for("main.add_category"))

        # 🔥 NORMALIZE NAME (VERY IMPORTANT)
        normalized_name = re.sub(r'\s+', ' ', name).lower()

        # 🔒 HARD DUPLICATE CHECK
        existing = mongo.db.categories.find_one({
            "user_id": ObjectId(current_user.id),
            "type": category_type,
            "name_normalized": normalized_name
        })

        if existing:
            flash("Category already exists.", "danger")
            return redirect(url_for("main.add_category"))

        # 🧠 ITEMS SAFE
        items_raw = request.form.get("items", "").strip()

        items = []
        if items_raw and items_raw.lower() not in ["no items", "none", "-"]:
            items = [
                i.strip()
                for i in items_raw.split(",")
                if i.strip()
            ]

        items = list(dict.fromkeys(items))

        # 💾 SAVE WITH NORMALIZED FIELD
        data = {
            "user_id": ObjectId(current_user.id),
            "name": name,
            "name_normalized": normalized_name,   # 🔥 KEY FIX
            "slug": name.lower().replace(" ", "-"),
            "items": items,
            "type": category_type,
            "status": True,
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow()
        }

        mongo.db.categories.insert_one(data)

        flash("Category added successfully.", "success")
        return redirect(url_for("main.category_list"))

    return render_template("backend/pages/components/categories/add_category.html")



def normalize_name(name):
    return re.sub(r"\s+", " ", str(name)).strip().lower()


@bp.route("/categories")
@login_required
def category_list():

    # Get categories
    if current_user.role == UserRole.superadmin.value:
        raw_categories = list(mongo.db.categories.find())
    else:
        raw_categories = list(
            mongo.db.categories.find({
                "user_id": ObjectId(current_user.id)
            })
        )

    categories = []

    for category in raw_categories:

        items = category.get("items", [])

        fixed_items = []
        seen = set()

        if isinstance(items, list):

            for item in items:

                if item is None:
                    continue

                item = str(item).replace('"', "").strip()

                if not item:
                    continue

                # haddii uu database-ku string ahaan u kaydiyey
                if item.startswith("[") and item.endswith("]"):
                    item = item[1:-1].strip()

                key = normalize_name(item)

                if key not in seen:
                    seen.add(key)
                    fixed_items.append(item)

        category["items"] = fixed_items

        categories.append(Category(category))

    return render_template(
        "backend/pages/components/categories/all_categories.html",
        categories=categories
    )


@bp.route("/edit-category/<id>", methods=["GET", "POST"])
@login_required
def edit_category(id):

    # 🔒 Validate ObjectId
    try:
        category_id = ObjectId(id)
    except:
        flash("Invalid category ID", "danger")
        return redirect(url_for("main.category_list"))

    # 🔎 Get category
    category = mongo.db.categories.find_one({"_id": category_id})

    if not category:
        flash("Category not found", "danger")
        return redirect(url_for("main.category_list"))

    # 🔐 SECURITY CHECK
    if current_user.role != "superadmin":
        if str(category.get("user_id")) != str(current_user.id):
            flash("Not allowed", "danger")
            return redirect(url_for("main.category_list"))

    # 🧹 CLEAN ITEMS FOR DISPLAY
    raw_items = category.get("items", [])
    cleaned_items = []

    if isinstance(raw_items, list):
        for i in raw_items:
            if isinstance(i, str):
                try:
                    decoded = json.loads(i)
                    if isinstance(decoded, list):
                        cleaned_items.extend(decoded)
                    else:
                        cleaned_items.append(str(decoded))
                except:
                    cleaned_items.append(i)
            else:
                cleaned_items.append(str(i))

    category["items"] = list(dict.fromkeys(cleaned_items))

    # POST UPDATE
    if request.method == "POST":

        name = request.form.get("name", "").strip()
        category_type = request.form.get("type", "").strip()

        if not name:
            flash("Category name is required", "danger")
            return redirect(request.url)

        # 🔥 NORMALIZE NAME
        normalized_name = normalize_name(name)

        # 🔒 DUPLICATE CHECK (IMPORTANT FIX)
        existing = mongo.db.categories.find_one({
            "_id": {"$ne": category_id},
            "user_id": ObjectId(current_user.id),
            "type": category_type,
            "name_normalized": normalized_name
        })

        if existing:
            flash("Category already exists.", "danger")
            return redirect(request.url)

        # 🧠 SAFE ITEMS PARSING
        items_raw = request.form.get("items", "")

        items = []

        if items_raw and items_raw.lower() not in ["no items", "none", "-"]:
            items = [
                i.strip()
                for i in items_raw.split(",")
                if i.strip() and i.lower() != "no items"
            ]

        # 🧼 CLEAN + REMOVE DUPLICATES
        items = list(dict.fromkeys(items))

        # 💾 UPDATE DB
        mongo.db.categories.update_one(
            {"_id": category_id},
            {"$set": {
                "name": name,
                "name_normalized": normalized_name,  # 🔥 IMPORTANT FIX
                "slug": name.lower().replace(" ", "-"),
                "type": category_type,
                "items": items,
                "updated_at": datetime.utcnow()
            }}
        )

        flash("Category updated successfully", "success")
        return redirect(url_for("main.category_list"))

    return render_template(
        "backend/pages/components/categories/edit_category.html",
        category=category
    )


@bp.route("/delete-category/<id>", methods=["GET", "POST"])
@login_required
def delete_category(id):

    category = mongo.db.categories.find_one({"_id": ObjectId(id)})

    if not category:
        flash("Category not found", "danger")
        return redirect(url_for("main.category_list"))

    # ADMIN SECURITY CHECK
    if current_user.role != "superadmin" and str(category["user_id"]) != str(current_user.id):
        flash("Not allowed", "danger")
        return redirect(url_for("main.category_list"))

    mongo.db.categories.delete_one({"_id": ObjectId(id)})

    flash("Category deleted successfully", "success")
    return redirect(url_for("main.category_list"))


@bp.route("/export-categories")
@login_required
def export_categories():

    # 🔎 GET DATA BASED ON ROLE
    if current_user.role == UserRole.superadmin.value:
        categories = list(mongo.db.categories.find())
    else:
        categories = list(mongo.db.categories.find({
            "user_id": ObjectId(current_user.id)
        }))

    # 🧼 CLEAN FOR JSON EXPORT
    clean_data = []

    for c in categories:
        clean_data.append({
            "name": c.get("name"),
            "name_normalized": c.get("name_normalized"),
            "slug": c.get("slug"),
            "type": c.get("type"),
            "items": c.get("items", []),
            "status": c.get("status"),
            "created_at": str(c.get("created_at")),
            "updated_at": str(c.get("updated_at")),
        })

    return Response(
        json.dumps(clean_data, indent=4),
        mimetype="application/json",
        headers={
            "Content-Disposition": "attachment; filename=categories.json"
        }
    )



@bp.route("/import-categories", methods=["POST"])
@login_required
def import_categories():

    file = request.files.get("file")

    if not file:
        flash("Please upload a file", "danger")
        return redirect(url_for("main.category_list"))

    try:
        data = json.load(file)
    except:
        flash("Invalid JSON file", "danger")
        return redirect(url_for("main.category_list"))

    if not isinstance(data, list):
        flash("Invalid data format (must be list)", "danger")
        return redirect(url_for("main.category_list"))

    imported = 0
    skipped = 0

    for item in data:

        if not isinstance(item, dict):
            continue

        name = (item.get("name") or "").strip()
        category_type = (item.get("type") or "expense").strip()

        if not name:
            continue

        normalized = normalize_name(name)

        # 🔒 DUPLICATE CHECK (USER + TYPE + NAME)
        existing = mongo.db.categories.find_one({
            "user_id": ObjectId(current_user.id),
            "type": category_type,
            "name_normalized": normalized
        })

        if existing:
            skipped += 1
            continue

        # 🧠 SAFE ITEMS PARSE
        items = item.get("items") or []

        if not isinstance(items, list):
            items = []

        # remove duplicates inside items
        items = list(dict.fromkeys([i.strip() for i in items if isinstance(i, str) and i.strip()]))

        # 💾 INSERT
        mongo.db.categories.insert_one({
            "user_id": ObjectId(current_user.id),
            "name": name,
            "name_normalized": normalized,
            "slug": name.lower().replace(" ", "-"),
            "items": items,
            "type": category_type,
            "status": bool(item.get("status", True)),
            "created_at": datetime.utcnow(),
            "updated_at": datetime.utcnow()
        })

        imported += 1

    flash(f"Imported: {imported}, Skipped: {skipped}", "success")
    return redirect(url_for("main.category_list"))


# ==============================
# ACCOUNT LIST
# ==============================
@bp.route("/accounts")
@login_required
def account_list():

    from bson import ObjectId
    from datetime import datetime

    # ========================================================
    # SAFE NUMBER
    # ========================================================

    def to_float(value):

        try:
            return float(value or 0)

        except (TypeError, ValueError):

            return 0.0


    # ========================================================
    # SAFE OBJECT ID
    # ========================================================

    def safe_object_id(value):

        try:

            return ObjectId(str(value))

        except Exception:

            return None


    # ========================================================
    # SERIALIZE MONGO VALUES
    # ========================================================

    def serialize_value(value):

        if isinstance(value, ObjectId):

            return str(value)

        if isinstance(value, datetime):

            return value.strftime(
                "%Y-%m-%d %H:%M:%S"
            )

        if isinstance(value, dict):

            return {
                str(k): serialize_value(v)
                for k, v in value.items()
            }

        if isinstance(value, list):

            return [
                serialize_value(v)
                for v in value
            ]

        return value


    # ========================================================
    # USER ID
    # ========================================================

    current_user_id = safe_object_id(
        current_user.id
    )


    # ========================================================
    # USER QUERY
    # ========================================================

    query = {}

    if current_user.role != UserRole.superadmin.value:

        if current_user_id is None:

            return render_template(
                "backend/pages/components/accounts/all_accounts.html",
                accounts=[],
                savings=[],
                account_reports=[],
                all_accounts_report={
                    "total_accounts": 0,
                    "opening_balance": 0,
                    "total_income": 0,
                    "total_expense": 0,
                    "transfer_in": 0,
                    "transfer_out": 0,
                    "net_movement": 0,
                    "current_balance": 0,
                    "transaction_count": 0,
                    "income_categories": {},
                    "expense_categories": {}
                },
                income_categories={},
                expense_categories={}
            )

        query["user_id"] = current_user_id


    # ========================================================
    # GET ACCOUNTS
    # ========================================================

    accounts = list(
        mongo.db.accounts.find(query)
        .sort("created_at", -1)
    )


    # ========================================================
    # GET TRANSACTIONS
    # ========================================================

    transactions = list(
        mongo.db.transactions.find(query)
        .sort("date", -1)
    )


    # ========================================================
    # GET SAVINGS
    # ========================================================

    saving_query = {}

    if current_user.role != UserRole.superadmin.value:

        saving_query["user_id"] = current_user_id


    savings = list(
        mongo.db.savings.find(
            saving_query
        )
    )


    # ========================================================
    # SAVING ID
    # ========================================================

    active_saving = (
        savings[0]
        if savings
        else None
    )


    saving_id = ""

    if active_saving:

        saving_id = str(
            active_saving.get("_id")
        )


    # ========================================================
    # GRAND TOTALS
    # ========================================================

    grand_opening = 0.0
    grand_income = 0.0
    grand_expense = 0.0
    grand_transfer_in = 0.0
    grand_transfer_out = 0.0
    grand_balance = 0.0
    grand_transactions = 0

    income_categories = {}
    expense_categories = {}


    # ========================================================
    # ACCOUNT REPORTS
    # ========================================================

    account_reports = []


    # ========================================================
    # ACCOUNT LOOP
    # ========================================================

    for account in accounts:

        account_id = str(
            account.get("_id")
        )


        account_name = (
            account.get("name")
            or "Unnamed Account"
        )


        account_type = (
            account.get("type")
            or "Other"
        )


        currency = (
            account.get("currency")
            or "USD"
        )


        # ====================================================
        # BALANCE
        # ====================================================

        current_balance = to_float(
            account.get("balance")
        )

        if (
            "balance" not in account
            and "current_balance" in account
        ):

            current_balance = to_float(
                account.get(
                    "current_balance"
                )
            )


        # ====================================================
        # OPENING BALANCE
        # ====================================================

        opening_balance = to_float(
            account.get(
                "opening_balance"
            )
        )


        if opening_balance == 0:

            opening_balance = to_float(
                account.get(
                    "initial_balance"
                )
            )


        # ====================================================
        # ACCOUNT TRANSACTIONS
        # ====================================================

        account_transactions = []

        total_income = 0.0
        total_expense = 0.0
        total_transfer_in = 0.0
        total_transfer_out = 0.0

        income_count = 0
        expense_count = 0
        transfer_count = 0

        account_income_categories = {}
        account_expense_categories = {}


        # ====================================================
        # TRANSACTION LOOP
        # ====================================================

        for trx in transactions:

            trx_account_id = trx.get(
                "account_id"
            )


            if trx_account_id is None:

                continue


            if str(
                trx_account_id
            ) != account_id:

                continue


            transaction_type = str(

                trx.get(
                    "transaction_type"
                )

                or

                trx.get(
                    "type"
                )

                or ""

            ).lower().strip()


            amount = to_float(
                trx.get("amount")
            )


            category = str(
                trx.get("category")
                or "Other"
            )


            # =================================================
            # INCOME
            # =================================================

            if transaction_type == "income":

                total_income += amount

                income_count += 1


                account_income_categories[
                    category
                ] = (

                    account_income_categories.get(
                        category,
                        0.0
                    )

                    + amount

                )


                income_categories[
                    category
                ] = (

                    income_categories.get(
                        category,
                        0.0
                    )

                    + amount

                )


            # =================================================
            # EXPENSE
            # =================================================

            elif transaction_type == "expense":

                total_expense += amount

                expense_count += 1


                account_expense_categories[
                    category
                ] = (

                    account_expense_categories.get(
                        category,
                        0.0
                    )

                    + amount

                )


                expense_categories[
                    category
                ] = (

                    expense_categories.get(
                        category,
                        0.0
                    )

                    + amount

                )


            # =================================================
            # TRANSFER IN
            # =================================================

            elif transaction_type in (

                "transfer_in",
                "transfer-in",
                "transfer in",
                "transferin"

            ):

                total_transfer_in += amount

                transfer_count += 1


            # =================================================
            # TRANSFER OUT
            # =================================================

            elif transaction_type in (

                "transfer_out",
                "transfer-out",
                "transfer out",
                "transferout"

            ):

                total_transfer_out += amount

                transfer_count += 1


            # =================================================
            # COPY TRANSACTION
            # =================================================

            transaction_copy = dict(
                trx
            )


            # Serialize ALL Mongo values
            transaction_copy = serialize_value(
                transaction_copy
            )


            # Account information
            transaction_copy[
                "account_name"
            ] = account_name


            transaction_copy[
                "account_type"
            ] = account_type


            account_transactions.append(
                transaction_copy
            )


        # ====================================================
        # NET MOVEMENT
        # ====================================================

        net_movement = (

            total_income

            + total_transfer_in

            - total_expense

            - total_transfer_out

        )


        # ====================================================
        # TOTAL TRANSACTIONS
        # ====================================================

        transaction_count = len(
            account_transactions
        )


        # ====================================================
        # GRAND TOTALS
        # ====================================================

        grand_opening += opening_balance

        grand_income += total_income

        grand_expense += total_expense

        grand_transfer_in += (
            total_transfer_in
        )

        grand_transfer_out += (
            total_transfer_out
        )

        grand_balance += current_balance

        grand_transactions += (
            transaction_count
        )


        # ====================================================
        # REPORT
        # ====================================================

        report = {

            "id":
                account_id,

            "name":
                account_name,

            "type":
                account_type,

            "currency":
                currency,

            "opening_balance":
                opening_balance,

            "balance":
                current_balance,

            # Compatibility
            "current_balance":
                current_balance,

            "total_income":
                total_income,

            "total_expense":
                total_expense,

            "transfer_in":
                total_transfer_in,

            "transfer_out":
                total_transfer_out,

            "net_movement":
                net_movement,

            "transaction_count":
                transaction_count,

            "income_count":
                income_count,

            "expense_count":
                expense_count,

            "transfer_count":
                transfer_count,

            "income_categories":
                account_income_categories,

            "expense_categories":
                account_expense_categories,

            "transactions":
                account_transactions
        }


        # ====================================================
        # ADD REPORT
        # ====================================================

        account_reports.append(
            report
        )


    # ========================================================
    # GRAND NET MOVEMENT
    # ========================================================

    grand_net = (

        grand_income

        + grand_transfer_in

        - grand_expense

        - grand_transfer_out

    )


    # ========================================================
    # ALL ACCOUNTS REPORT
    # ========================================================

    all_accounts_report = {

        "total_accounts":
            len(account_reports),

        "opening_balance":
            grand_opening,

        "total_income":
            grand_income,

        "total_expense":
            grand_expense,

        "transfer_in":
            grand_transfer_in,

        "transfer_out":
            grand_transfer_out,

        "net_movement":
            grand_net,

        "current_balance":
            grand_balance,

        "transaction_count":
            grand_transactions,

        "income_categories":
            income_categories,

        "expense_categories":
            expense_categories
    }


    # ========================================================
    # PREPARE NORMAL ACCOUNTS
    # ========================================================

    prepared_accounts = []


    for account in accounts:

        account_copy = dict(
            account
        )


        # ====================================================
        # ACCOUNT ID
        # ====================================================

        account_copy["_id"] = str(
            account_copy.get("_id")
        )


        # ====================================================
        # USER ID
        # ====================================================

        if account_copy.get(
            "user_id"
        ) is not None:

            account_copy["user_id"] = str(
                account_copy["user_id"]
            )


        # ====================================================
        # SAVING ID
        # ====================================================

        account_copy[
            "saving_id"
        ] = saving_id


        # ====================================================
        # FIND REPORT
        # ====================================================

        matching_report = next(

            (
                report

                for report in account_reports

                if report["id"]
                == account_copy["_id"]

            ),

            None

        )


        # ====================================================
        # ATTACH REPORT
        # ====================================================

        if matching_report:

            account_copy[
                "report"
            ] = matching_report


            # ------------------------------------------------
            # Compatibility fields
            # ------------------------------------------------

            account_copy[
                "current_balance"
            ] = matching_report[
                "balance"
            ]


            account_copy[
                "balance"
            ] = matching_report[
                "balance"
            ]


            account_copy[
                "opening_balance"
            ] = matching_report[
                "opening_balance"
            ]


            account_copy[
                "total_income"
            ] = matching_report[
                "total_income"
            ]


            account_copy[
                "total_expense"
            ] = matching_report[
                "total_expense"
            ]


            account_copy[
                "net_movement"
            ] = matching_report[
                "net_movement"
            ]


            account_copy[
                "transaction_count"
            ] = matching_report[
                "transaction_count"
            ]


            account_copy[
                "income_count"
            ] = matching_report[
                "income_count"
            ]


            account_copy[
                "expense_count"
            ] = matching_report[
                "expense_count"
            ]


            account_copy[
                "income_categories"
            ] = matching_report[
                "income_categories"
            ]


            account_copy[
                "expense_categories"
            ] = matching_report[
                "expense_categories"
            ]


            account_copy[
                "report_transactions"
            ] = matching_report[
                "transactions"
            ]


        else:

            # =================================================
            # EMPTY REPORT
            # =================================================

            empty_report = {

                "id":
                    account_copy["_id"],

                "name":
                    account_copy.get(
                        "name",
                        "Unnamed Account"
                    ),

                "type":
                    account_copy.get(
                        "type",
                        "Other"
                    ),

                "currency":
                    account_copy.get(
                        "currency",
                        "USD"
                    ),

                "opening_balance":
                    0.0,

                "balance":
                    0.0,

                "current_balance":
                    0.0,

                "total_income":
                    0.0,

                "total_expense":
                    0.0,

                "transfer_in":
                    0.0,

                "transfer_out":
                    0.0,

                "net_movement":
                    0.0,

                "transaction_count":
                    0,

                "income_count":
                    0,

                "expense_count":
                    0,

                "transfer_count":
                    0,

                "income_categories":
                    {},

                "expense_categories":
                    {},

                "transactions":
                    []
            }


            account_copy[
                "report"
            ] = empty_report


            account_copy[
                "current_balance"
            ] = 0.0


            account_copy[
                "total_income"
            ] = 0.0


            account_copy[
                "total_expense"
            ] = 0.0


            account_copy[
                "net_movement"
            ] = 0.0


            account_copy[
                "transaction_count"
            ] = 0


        prepared_accounts.append(
            account_copy
        )


    # ========================================================
    # SERIALIZE SAVINGS
    # ========================================================

    prepared_savings = []

    for saving in savings:

        saving_copy = serialize_value(
            dict(saving)
        )

        prepared_savings.append(
            saving_copy
        )


    # ========================================================
    # RENDER
    # ========================================================

    return render_template(

        "backend/pages/components/accounts/all_accounts.html",

        accounts=prepared_accounts,

        savings=prepared_savings,

        account_reports=account_reports,

        all_accounts_report=
            all_accounts_report,

        income_categories=
            income_categories,

        expense_categories=
            expense_categories,

        saving_id=
            saving_id

    )





@bp.route('/add-account', methods=['GET', 'POST'])
@login_required
def add_account():

    if request.method == "POST":

        # Form Data
        name = request.form.get("name", "").strip()
        account_type = request.form.get("type", "cash").strip()
        balance = request.form.get("balance", 0)
        currency = request.form.get("currency", "USD").strip()

        # Validation
        if not name:
            flash("Account name is required.", "danger")
            return redirect(url_for("main.add_account"))

        try:
            balance = float(balance)
        except ValueError:
            flash("Invalid balance amount.", "danger")
            return redirect(url_for("main.add_account"))

        # Duplicate Check
        existing = mongo.db.accounts.find_one({
            "user_id": ObjectId(current_user.id),
            "name": {
                "$regex": f"^{name}$",
                "$options": "i"
            }
        })

        if existing:
            flash("Account already exists.", "danger")
            return redirect(url_for("main.add_account"))

        # Save
        account = Account()

        data = account.add(
            user_id=current_user.id,
            name=name,
            account_type=account_type,
            balance=balance,
            currency=currency,
            status=True
        )

        data["user_id"] = ObjectId(current_user.id)

        mongo.db.accounts.insert_one(data)

        flash("Account created successfully.", "success")
        return redirect(url_for("main.account_list"))

    return render_template(
        "backend/pages/components/accounts/add_account.html"
    )


@bp.route("/edit-account/<id>", methods=["GET", "POST"])
@login_required
def edit_account(id):

    account = mongo.db.accounts.find_one({
        "_id": ObjectId(id),
        "user_id": ObjectId(current_user.id)
    })

    if not account:
        flash("Account not found.", "danger")
        return redirect(url_for("main.account_list"))

    if request.method == "POST":

        name = request.form.get("name", "").strip()
        account_type = request.form.get("type")
        currency = request.form.get("currency")
        balance = request.form.get("balance", 0)

        if not name:
            flash("Account name is required.", "danger")
            return redirect(url_for("main.edit_account", id=id))

        try:
            balance = float(balance)
        except ValueError:
            balance = 0

        mongo.db.accounts.update_one(
            {"_id": ObjectId(id)},
            {
                "$set": {
                    "name": name,
                    "type": account_type,
                    "currency": currency,
                    "balance": balance,
                    "updated_at": datetime.utcnow()
                }
            }
        )

        flash("Account updated successfully.", "success")
        return redirect(url_for("main.account_list"))

    return render_template(
        "backend/pages/components/accounts/edit_account.html",
        account=account
    )


@bp.route("/delete-account/<id>")
@login_required
def delete_account(id):

    account = mongo.db.accounts.find_one({
        "_id": ObjectId(id),
        "user_id": ObjectId(current_user.id)
    })

    if not account:
        flash("Account not found.", "danger")
        return redirect(url_for("main.account_list"))

    # 🔥 SAFETY CHECK (IMPORTANT)
    has_transactions = mongo.db.transactions.find_one({
        "account_id": id
    })

    has_savings = mongo.db.savings.find_one({
        "account_id": id
    })

    if has_transactions or has_savings:
        flash("This account cannot be deleted because it is in use.", "warning")
        return redirect(url_for("main.account_list"))

    mongo.db.accounts.delete_one({
        "_id": ObjectId(id)
    })

    flash("Account deleted successfully.", "success")
    return redirect(url_for("main.account_list"))



# ============================================================
# ACCOUNT TRANSFER
# ============================================================


# ============================================================
# ACCOUNT TRANSFER
# ============================================================
@bp.route("/account/transfer", methods=["POST"])
@login_required
def account_transfer():

    from bson import ObjectId
    from datetime import datetime
    from flask import request, jsonify
    import uuid

    # ============================================================
    # HELPERS
    # ============================================================

    def oid(value):
        if not value:
            return None

        try:
            if isinstance(value, ObjectId):
                return value

            return ObjectId(str(value))

        except Exception:
            return None

    def safe_float(value, default=0.0):
        try:
            return float(value or 0)
        except (TypeError, ValueError):
            return default

    # ============================================================
    # REQUEST DATA
    # ============================================================

    data = request.get_json(silent=True)

    if not data:
        return jsonify({
            "success": False,
            "error": "No data provided."
        }), 400

    transfer_type = (
        data.get("transfer_type")
        or "account_to_account"
    )

    # ------------------------------------------------------------
    # Supported:
    #
    # account_to_account
    # account_to_saving
    # saving_to_account
    # ------------------------------------------------------------

    allowed_types = {
        "account_to_account",
        "account_to_saving",
        "saving_to_account"
    }

    if transfer_type not in allowed_types:
        return jsonify({
            "success": False,
            "error": "Invalid transfer type."
        }), 400

    amount = safe_float(
        data.get("amount")
    )

    if amount <= 0:
        return jsonify({
            "success": False,
            "error": "Transfer amount must be greater than zero."
        }), 400

    # ============================================================
    # CURRENT USER
    # ============================================================

    user_id = str(current_user.id)

    user_object_id = oid(user_id)

    # ============================================================
    # OWNER FILTER
    # ============================================================

    if current_user.role == UserRole.superadmin.value:

        owner_filter = {}

    else:

        owner_conditions = [
            {
                "user_id": user_id
            }
        ]

        if user_object_id:

            owner_conditions.append({
                "user_id": user_object_id
            })

        owner_filter = {
            "$or": owner_conditions
        }

    # ============================================================
    # IDS
    # ============================================================

    from_account_id = data.get(
        "from_account"
    )

    to_account_id = data.get(
        "to_account"
    )

    from_saving_id = data.get(
        "from_saving"
    )

    to_saving_id = data.get(
        "to_saving"
    )

    from_account_obj = oid(
        from_account_id
    )

    to_account_obj = oid(
        to_account_id
    )

    from_saving_obj = oid(
        from_saving_id
    )

    to_saving_obj = oid(
        to_saving_id
    )

    # ============================================================
    # DATE
    # ============================================================

    now = datetime.utcnow()

    # ============================================================
    # REFERENCE
    # ============================================================

    if transfer_type == "saving_to_account":

        reference_no = (
            "SAV-"
            + now.strftime("%Y%m%d%H%M%S")
            + "-"
            + uuid.uuid4().hex[:6].upper()
        )

    elif transfer_type == "account_to_saving":

        reference_no = (
            "SAV-"
            + now.strftime("%Y%m%d%H%M%S")
            + "-"
            + uuid.uuid4().hex[:6].upper()
        )

    else:

        reference_no = (
            "TRF-"
            + uuid.uuid4().hex[:10].upper()
        )

    # ============================================================
    # COMMON VARIABLES
    # ============================================================

    source_name = "Unknown"
    destination_name = "Unknown"

    source_currency = "USD"
    destination_currency = "USD"

    description = (
        data.get("description")
        or ""
    )

    direction = ""

    # ============================================================
    # BALANCE SNAPSHOTS
    #
    # IMPORTANT:
    #
    # These are the values that the REPORT/MODAL will use.
    # ============================================================

    from_account_balance_before = None
    from_account_balance_after = None

    to_account_balance_before = None
    to_account_balance_after = None

    from_saving_balance_before = None
    from_saving_balance_after = None

    to_saving_balance_before = None
    to_saving_balance_after = None

    # ============================================================
    # ACCOUNT -> ACCOUNT
    # ============================================================

    if transfer_type == "account_to_account":

        direction = "transfer"

        if not from_account_obj:
            return jsonify({
                "success": False,
                "error": "Source account is required."
            }), 400

        if not to_account_obj:
            return jsonify({
                "success": False,
                "error": "Destination account is required."
            }), 400

        if from_account_obj == to_account_obj:
            return jsonify({
                "success": False,
                "error": "You cannot transfer to the same account."
            }), 400

        # --------------------------------------------------------
        # GET ACCOUNTS
        # --------------------------------------------------------

        source_query = {
            "_id": from_account_obj,
            **owner_filter
        }

        destination_query = {
            "_id": to_account_obj,
            **owner_filter
        }

        source = mongo.db.accounts.find_one(
            source_query
        )

        destination = mongo.db.accounts.find_one(
            destination_query
        )

        if not source:
            return jsonify({
                "success": False,
                "error": "Source account not found."
            }), 404

        if not destination:
            return jsonify({
                "success": False,
                "error": "Destination account not found."
            }), 404

        # --------------------------------------------------------
        # CURRENCY
        # --------------------------------------------------------

        source_currency = (
            source.get("currency")
            or "USD"
        )

        destination_currency = (
            destination.get("currency")
            or "USD"
        )

        if source_currency != destination_currency:

            return jsonify({
                "success": False,
                "error": (
                    f"Currency mismatch. "
                    f"Source uses {source_currency}, "
                    f"destination uses {destination_currency}."
                )
            }), 400

        # --------------------------------------------------------
        # BEFORE
        # --------------------------------------------------------

        from_account_balance_before = safe_float(
            source.get("balance")
        )

        to_account_balance_before = safe_float(
            destination.get("balance")
        )

        if amount > from_account_balance_before:

            return jsonify({
                "success": False,
                "error": (
                    "Insufficient balance. "
                    f"Available {source_currency} "
                    f"{from_account_balance_before:,.2f}"
                )
            }), 400

        source_name = (
            source.get("name")
            or source.get("account_name")
            or "Source Account"
        )

        destination_name = (
            destination.get("name")
            or destination.get("account_name")
            or "Destination Account"
        )

        description = (
            data.get("description")
            or f"Transfer from {source_name} to {destination_name}"
        )

        # --------------------------------------------------------
        # UPDATE SOURCE
        # --------------------------------------------------------

        source_update = mongo.db.accounts.update_one(

            {
                "_id": from_account_obj,
                **owner_filter,
                "balance": {
                    "$gte": amount
                }
            },

            {
                "$inc": {
                    "balance": -amount
                },

                "$set": {
                    "updated_at": now
                }
            }
        )

        if source_update.modified_count != 1:

            return jsonify({
                "success": False,
                "error": "Source balance changed. Transfer failed."
            }), 400

        # --------------------------------------------------------
        # UPDATE DESTINATION
        # --------------------------------------------------------

        destination_update = mongo.db.accounts.update_one(

            {
                "_id": to_account_obj,
                **owner_filter
            },

            {
                "$inc": {
                    "balance": amount
                },

                "$set": {
                    "updated_at": now
                }
            }
        )

        if destination_update.modified_count != 1:

            mongo.db.accounts.update_one(
                {
                    "_id": from_account_obj
                },
                {
                    "$inc": {
                        "balance": amount
                    }
                }
            )

            return jsonify({
                "success": False,
                "error": "Destination account could not be updated."
            }), 500

        # --------------------------------------------------------
        # AFTER
        # --------------------------------------------------------

        from_account_balance_after = (
            from_account_balance_before - amount
        )

        to_account_balance_after = (
            to_account_balance_before + amount
        )

    # ============================================================
    # ACCOUNT -> SAVING
    # ============================================================

    elif transfer_type == "account_to_saving":

        direction = "deposit"

        if not from_account_obj:
            return jsonify({
                "success": False,
                "error": "Source account is required."
            }), 400

        if not to_saving_obj:
            return jsonify({
                "success": False,
                "error": "Destination saving is required."
            }), 400

        # --------------------------------------------------------
        # GET ACCOUNT
        # --------------------------------------------------------

        account = mongo.db.accounts.find_one({
            "_id": from_account_obj,
            **owner_filter
        })

        # --------------------------------------------------------
        # GET SAVING
        # --------------------------------------------------------

        saving = mongo.db.savings.find_one({
            "_id": to_saving_obj,
            **owner_filter
        })

        if not account:
            return jsonify({
                "success": False,
                "error": "Source account not found."
            }), 404

        if not saving:
            return jsonify({
                "success": False,
                "error": "Destination saving not found."
            }), 404

        # --------------------------------------------------------
        # CURRENCY
        # --------------------------------------------------------

        source_currency = (
            account.get("currency")
            or "USD"
        )

        destination_currency = (
            saving.get("currency")
            or source_currency
        )

        if source_currency != destination_currency:

            return jsonify({
                "success": False,
                "error": (
                    f"Currency mismatch. "
                    f"Account uses {source_currency}, "
                    f"saving uses {destination_currency}."
                )
            }), 400

        # --------------------------------------------------------
        # BEFORE
        # --------------------------------------------------------

        from_account_balance_before = safe_float(
            account.get("balance")
        )

        to_saving_balance_before = safe_float(
            saving.get("current_balance")
        )

        if amount > from_account_balance_before:

            return jsonify({
                "success": False,
                "error": (
                    "Insufficient account balance. "
                    f"Available {source_currency} "
                    f"{from_account_balance_before:,.2f}"
                )
            }), 400

        source_name = (
            account.get("name")
            or account.get("account_name")
            or "Source Account"
        )

        destination_name = (
            saving.get("name")
            or saving.get("saving_name")
            or saving.get("title")
            or "Saving"
        )

        description = (
            data.get("description")
            or "Account deposit to saving"
        )

        # --------------------------------------------------------
        # ACCOUNT - AMOUNT
        # --------------------------------------------------------

        account_update = mongo.db.accounts.update_one(

            {
                "_id": from_account_obj,
                **owner_filter,
                "balance": {
                    "$gte": amount
                }
            },

            {
                "$inc": {
                    "balance": -amount
                },

                "$set": {
                    "updated_at": now
                }
            }
        )

        if account_update.modified_count != 1:

            return jsonify({
                "success": False,
                "error": "Account balance changed. Transfer failed."
            }), 400

        # --------------------------------------------------------
        # SAVING + AMOUNT
        # --------------------------------------------------------

        saving_update = mongo.db.savings.update_one(

            {
                "_id": to_saving_obj,
                **owner_filter
            },

            {
                "$inc": {
                    "current_balance": amount
                },

                "$set": {
                    "updated_at": now
                }
            }
        )

        if saving_update.modified_count != 1:

            mongo.db.accounts.update_one(
                {
                    "_id": from_account_obj
                },
                {
                    "$inc": {
                        "balance": amount
                    }
                }
            )

            return jsonify({
                "success": False,
                "error": "Saving could not be updated."
            }), 500

        # --------------------------------------------------------
        # AFTER
        # --------------------------------------------------------

        from_account_balance_after = (
            from_account_balance_before - amount
        )

        to_saving_balance_after = (
            to_saving_balance_before + amount
        )

    # ============================================================
    # SAVING -> ACCOUNT
    # ============================================================

    elif transfer_type == "saving_to_account":

        direction = "withdraw"

        if not from_saving_obj:
            return jsonify({
                "success": False,
                "error": "Source saving is required."
            }), 400

        if not to_account_obj:
            return jsonify({
                "success": False,
                "error": "Destination account is required."
            }), 400

        # --------------------------------------------------------
        # GET SAVING
        # --------------------------------------------------------

        saving = mongo.db.savings.find_one({
            "_id": from_saving_obj,
            **owner_filter
        })

        # --------------------------------------------------------
        # GET ACCOUNT
        # --------------------------------------------------------

        account = mongo.db.accounts.find_one({
            "_id": to_account_obj,
            **owner_filter
        })

        if not saving:

            return jsonify({
                "success": False,
                "error": "Source saving not found."
            }), 404

        if not account:

            return jsonify({
                "success": False,
                "error": "Destination account not found."
            }), 404

        # --------------------------------------------------------
        # CURRENCY
        # --------------------------------------------------------

        source_currency = (
            saving.get("currency")
            or "USD"
        )

        destination_currency = (
            account.get("currency")
            or source_currency
        )

        if source_currency != destination_currency:

            return jsonify({
                "success": False,
                "error": (
                    f"Currency mismatch. "
                    f"Saving uses {source_currency}, "
                    f"account uses {destination_currency}."
                )
            }), 400

        # --------------------------------------------------------
        # BEFORE
        # --------------------------------------------------------

        from_saving_balance_before = safe_float(
            saving.get("current_balance")
        )

        to_account_balance_before = safe_float(
            account.get("balance")
        )

        # ========================================================
        # IMPORTANT
        # ========================================================
        #
        # Saving before must be the REAL current saving balance.
        #
        # Example:
        #
        # Saving = 212
        # Amount = 2
        #
        # After = 210
        #
        # NOT 0.
        # ========================================================

        if amount > from_saving_balance_before:

            return jsonify({
                "success": False,
                "error": (
                    "Insufficient saving balance. "
                    f"Available {source_currency} "
                    f"{from_saving_balance_before:,.2f}"
                )
            }), 400

        source_name = (
            saving.get("name")
            or saving.get("saving_name")
            or saving.get("title")
            or "Saving"
        )

        destination_name = (
            account.get("name")
            or account.get("account_name")
            or "Destination Account"
        )

        description = (
            data.get("description")
            or "Saving withdrawal to account"
        )

        # --------------------------------------------------------
        # SAVING - AMOUNT
        # --------------------------------------------------------

        saving_update = mongo.db.savings.update_one(

            {
                "_id": from_saving_obj,
                **owner_filter,
                "current_balance": {
                    "$gte": amount
                }
            },

            {
                "$inc": {
                    "current_balance": -amount
                },

                "$set": {
                    "updated_at": now
                }
            }
        )

        if saving_update.modified_count != 1:

            return jsonify({
                "success": False,
                "error": "Saving balance changed. Transfer failed."
            }), 400

        # --------------------------------------------------------
        # ACCOUNT + AMOUNT
        # --------------------------------------------------------

        account_update = mongo.db.accounts.update_one(

            {
                "_id": to_account_obj,
                **owner_filter
            },

            {
                "$inc": {
                    "balance": amount
                },

                "$set": {
                    "updated_at": now
                }
            }
        )

        if account_update.modified_count != 1:

            # Rollback saving
            mongo.db.savings.update_one(
                {
                    "_id": from_saving_obj
                },
                {
                    "$inc": {
                        "current_balance": amount
                    }
                }
            )

            return jsonify({
                "success": False,
                "error": "Destination account could not be updated."
            }), 500

        # --------------------------------------------------------
        # AFTER
        # --------------------------------------------------------

        from_saving_balance_after = (
            from_saving_balance_before - amount
        )

        to_account_balance_after = (
            to_account_balance_before + amount
        )

    # ============================================================
    # MASTER TRANSFER RECORD
    # ============================================================

    transfer_record = {

        "_id": ObjectId(),

        "user_id": str(user_id),

        # --------------------------------------------------------
        # TYPE
        # --------------------------------------------------------

        "transfer_type": transfer_type,

        "direction": direction,

        # --------------------------------------------------------
        # IDS
        # --------------------------------------------------------

        "from_account": from_account_obj,
        "to_account": to_account_obj,

        "from_saving": from_saving_obj,
        "to_saving": to_saving_obj,

        # --------------------------------------------------------
        # NAMES
        # --------------------------------------------------------

        "from_account_name": (
            source_name
            if transfer_type != "saving_to_account"
            else None
        ),

        "to_account_name": (
            destination_name
            if transfer_type != "account_to_saving"
            else None
        ),

        "from_saving_name": (
            source_name
            if transfer_type == "saving_to_account"
            else None
        ),

        "to_saving_name": (
            destination_name
            if transfer_type == "account_to_saving"
            else None
        ),

        # --------------------------------------------------------
        # AMOUNT
        # --------------------------------------------------------

        "amount": amount,

        "currency": source_currency,

        # --------------------------------------------------------
        # STATUS
        # --------------------------------------------------------

        "status": "completed",

        # --------------------------------------------------------
        # REFERENCE
        # --------------------------------------------------------

        "reference": reference_no,

        "reference_no": reference_no,

        # --------------------------------------------------------
        # DESCRIPTION
        # --------------------------------------------------------

        "description": description,

        # ========================================================
        # ACCOUNT BALANCE REPORT
        # ========================================================

        "from_account_balance_before":
            from_account_balance_before,

        "from_account_balance_after":
            from_account_balance_after,

        "to_account_balance_before":
            to_account_balance_before,

        "to_account_balance_after":
            to_account_balance_after,

        # ========================================================
        # SAVING BALANCE REPORT
        # ========================================================

        "from_saving_balance_before":
            from_saving_balance_before,

        "from_saving_balance_after":
            from_saving_balance_after,

        "to_saving_balance_before":
            to_saving_balance_before,

        "to_saving_balance_after":
            to_saving_balance_after,

        # --------------------------------------------------------
        # TIMESTAMPS
        # --------------------------------------------------------

        "created_at": now,

        "updated_at": now
    }

    # ============================================================
    # INSERT MASTER
    # ============================================================

    transfer_result = (
        mongo.db.account_transfers.insert_one(
            transfer_record
        )
    )

    transfer_id = (
        transfer_result.inserted_id
    )

    # ============================================================
    # TRANSACTION RECORDS
    # ============================================================

    transaction_records = []

    # ============================================================
    # ACCOUNT -> ACCOUNT
    # ============================================================

    if transfer_type == "account_to_account":

        # --------------------------------------------------------
        # SOURCE
        # --------------------------------------------------------

        transaction_records.append({

            "user_id": str(user_id),

            "account_id": from_account_obj,

            "transaction_type": "transfer_out",

            "type": "transfer_out",

            "amount": amount,

            "currency": source_currency,

            "category": "Account Transfer",

            "item": (
                f"Transfer to {destination_name}"
            ),

            "description": description,

            "note": (
                f"Money transferred to "
                f"{destination_name}"
            ),

            "reference_no": reference_no,

            "transfer_id": transfer_id,

            "from_account": from_account_obj,

            "to_account": to_account_obj,

            "balance_before":
                from_account_balance_before,

            "balance_after":
                from_account_balance_after,

            "created_at": now,

            "updated_at": now,

            "date": now
        })

        # --------------------------------------------------------
        # DESTINATION
        # --------------------------------------------------------

        transaction_records.append({

            "user_id": str(user_id),

            "account_id": to_account_obj,

            "transaction_type": "transfer_in",

            "type": "transfer_in",

            "amount": amount,

            "currency": destination_currency,

            "category": "Account Transfer",

            "item": (
                f"Transfer from {source_name}"
            ),

            "description": description,

            "note": (
                f"Money received from "
                f"{source_name}"
            ),

            "reference_no": reference_no,

            "transfer_id": transfer_id,

            "from_account": from_account_obj,

            "to_account": to_account_obj,

            "balance_before":
                to_account_balance_before,

            "balance_after":
                to_account_balance_after,

            "created_at": now,

            "updated_at": now,

            "date": now
        })

    # ============================================================
    # ACCOUNT -> SAVING
    # ============================================================

    elif transfer_type == "account_to_saving":

        transaction_records.append({

            "user_id": str(user_id),

            "account_id": from_account_obj,

            "transaction_type": "transfer_out",

            "type": "transfer_out",

            "amount": amount,

            "currency": source_currency,

            "category": "Saving Deposit",

            "item": (
                f"Deposit to {destination_name}"
            ),

            "description": description,

            "note": (
                f"Money transferred from "
                f"{source_name} to saving"
            ),

            "reference_no": reference_no,

            "transfer_id": transfer_id,

            "from_account": from_account_obj,

            "to_saving": to_saving_obj,

            "balance_before":
                from_account_balance_before,

            "balance_after":
                from_account_balance_after,

            "created_at": now,

            "updated_at": now,

            "date": now
        })

        transaction_records.append({

            "user_id": str(user_id),

            "saving_id": to_saving_obj,

            "account_id": from_account_obj,

            "transaction_type": "deposit",

            "type": "deposit",

            "amount": amount,

            "currency": destination_currency,

            "category": "Saving Deposit",

            "item": (
                f"Deposit from {source_name}"
            ),

            "description": description,

            "note": (
                f"Money received from "
                f"{source_name}"
            ),

            "reference_no": reference_no,

            "transfer_id": transfer_id,

            "from_account": from_account_obj,

            "to_saving": to_saving_obj,

            "balance_before":
                to_saving_balance_before,

            "balance_after":
                to_saving_balance_after,

            "saving_balance_before":
                to_saving_balance_before,

            "saving_balance_after":
                to_saving_balance_after,

            "created_at": now,

            "updated_at": now,

            "date": now
        })

    # ============================================================
    # SAVING -> ACCOUNT
    # ============================================================

    elif transfer_type == "saving_to_account":

        # --------------------------------------------------------
        # SAVING TRANSACTION
        # --------------------------------------------------------

        transaction_records.append({

            "user_id": str(user_id),

            "saving_id": from_saving_obj,

            "account_id": to_account_obj,

            "transaction_type": "withdrawal",

            "type": "withdrawal",

            "amount": amount,

            "currency": source_currency,

            "category": "Saving Withdrawal",

            "item": (
                f"Withdrawal to {destination_name}"
            ),

            "description": description,

            "note": (
                f"Money withdrawn from "
                f"{source_name}"
            ),

            "reference_no": reference_no,

            "transfer_id": transfer_id,

            "from_saving": from_saving_obj,

            "to_account": to_account_obj,

            # ====================================================
            # REAL SAVING BALANCE
            # ====================================================

            "balance_before":
                from_saving_balance_before,

            "balance_after":
                from_saving_balance_after,

            "saving_balance_before":
                from_saving_balance_before,

            "saving_balance_after":
                from_saving_balance_after,

            "created_at": now,

            "updated_at": now,

            "date": now
        })

        # --------------------------------------------------------
        # ACCOUNT TRANSACTION
        # --------------------------------------------------------

        transaction_records.append({

            "user_id": str(user_id),

            "account_id": to_account_obj,

            "transaction_type": "transfer_in",

            "type": "transfer_in",

            "amount": amount,

            "currency": destination_currency,

            "category": "Saving Withdrawal",

            "item": (
                f"Received from {source_name}"
            ),

            "description": description,

            "note": (
                f"Money received from "
                f"{source_name}"
            ),

            "reference_no": reference_no,

            "transfer_id": transfer_id,

            "from_saving": from_saving_obj,

            "to_account": to_account_obj,

            # ====================================================
            # REAL ACCOUNT BALANCE
            # ====================================================

            "balance_before":
                to_account_balance_before,

            "balance_after":
                to_account_balance_after,

            "account_balance_before":
                to_account_balance_before,

            "account_balance_after":
                to_account_balance_after,

            "created_at": now,

            "updated_at": now,

            "date": now
        })

    # ============================================================
    # INSERT TRANSACTIONS
    # ============================================================

    if transaction_records:

        mongo.db.transactions.insert_many(
            transaction_records
        )

    # ============================================================
    # ALSO SAVE SAVING TRANSACTION HISTORY
    # ============================================================

    if transfer_type == "account_to_saving":

        mongo.db.saving_transactions.insert_one({

            "user_id": str(user_id),

            "saving_id": to_saving_obj,

            "account_id": from_account_obj,

            "transaction_type": "deposit",

            "amount": amount,

            "description": description,

            "note": (
                f"Deposit from {source_name}"
            ),

            "reference_no": reference_no,

            "transfer_id": transfer_id,

            "balance_before":
                to_saving_balance_before,

            "balance_after":
                to_saving_balance_after,

            "created_at": now,

            "updated_at": now,

            "date": now,

            "status": True
        })

    elif transfer_type == "saving_to_account":

        mongo.db.saving_transactions.insert_one({

            "user_id": str(user_id),

            "saving_id": from_saving_obj,

            "account_id": to_account_obj,

            "transaction_type": "withdrawal",

            "amount": amount,

            "description": description,

            "note": (
                f"Withdrawal to {destination_name}"
            ),

            "reference_no": reference_no,

            "transfer_id": transfer_id,

            "balance_before":
                from_saving_balance_before,

            "balance_after":
                from_saving_balance_after,

            "created_at": now,

            "updated_at": now,

            "date": now,

            "status": True
        })

    # ============================================================
    # SUCCESS RESPONSE
    # ============================================================

    return jsonify({

        "success": True,

        "message": (
            "Money transferred successfully."
        ),

        "transfer_id": str(
            transfer_id
        ),

        "transfer_type":
            transfer_type,

        "direction":
            direction,

        "reference_no":
            reference_no,

        "amount":
            amount,

        "currency":
            source_currency,

        "from_name":
            source_name,

        "to_name":
            destination_name,

        # --------------------------------------------------------
        # ACCOUNT REPORT
        # --------------------------------------------------------

        "from_account_balance_before":
            from_account_balance_before,

        "from_account_balance_after":
            from_account_balance_after,

        "to_account_balance_before":
            to_account_balance_before,

        "to_account_balance_after":
            to_account_balance_after,

        # --------------------------------------------------------
        # SAVING REPORT
        # --------------------------------------------------------

        "from_saving_balance_before":
            from_saving_balance_before,

        "from_saving_balance_after":
            from_saving_balance_after,

        "to_saving_balance_before":
            to_saving_balance_before,

        "to_saving_balance_after":
            to_saving_balance_after

    }), 200

# ============================================================
# DELETE ACCOUNT / SAVING TRANSFER
# ============================================================

@bp.route(
    "/account/transfer/delete/<transfer_id>",
    methods=["POST", "DELETE"]
)
@login_required
def delete_account_transfer(transfer_id):

    from bson import ObjectId
    from datetime import datetime
    from flask import flash, redirect, url_for, jsonify

    # ========================================================
    # HELPER
    # ========================================================

    def redirect_back_success(message):
        flash(message, "success")

        return redirect(
            url_for("main.all_account_transfers")
        )

    def redirect_back_error(message):
        flash(message, "danger")

        return redirect(
            url_for("main.all_account_transfers")
        )

    def json_success(message):
        return jsonify({
            "success": True,
            "message": message
        }), 200

    def json_error(message, status=400):
        return jsonify({
            "success": False,
            "error": message,
            "message": message
        }), status

    # ========================================================
    # VALIDATE TRANSFER ID
    # ========================================================

    try:

        transfer_obj_id = ObjectId(
            str(transfer_id)
        )

    except Exception:

        message = "Invalid transfer ID."

        if request.method == "POST":
            return redirect_back_error(message)

        return json_error(message, 400)

    # ========================================================
    # CURRENT USER
    # ========================================================

    current_user_id = str(
        current_user.id
    )

    try:

        current_user_obj_id = ObjectId(
            current_user_id
        )

    except Exception:

        current_user_obj_id = None

    # ========================================================
    # FIND TRANSFER
    # ========================================================

    try:

        # ----------------------------------------------------
        # SUPERADMIN CAN DELETE ANY TRANSFER
        # ----------------------------------------------------

        if current_user.role == UserRole.superadmin.value:

            transfer = (
                mongo.db.account_transfers.find_one({
                    "_id": transfer_obj_id
                })
            )

        # ----------------------------------------------------
        # NORMAL USER
        # ----------------------------------------------------

        else:

            user_conditions = [
                {
                    "user_id": current_user_id
                }
            ]

            if current_user_obj_id:

                user_conditions.append({
                    "user_id":
                    current_user_obj_id
                })

            transfer = (
                mongo.db.account_transfers.find_one({

                    "_id":
                    transfer_obj_id,

                    "$or":
                    user_conditions

                })
            )

    except Exception as e:

        print(
            "FIND TRANSFER DELETE ERROR:",
            repr(e)
        )

        message = (
            "Unable to find the transfer."
        )

        if request.method == "POST":
            return redirect_back_error(message)

        return json_error(message, 500)

    # ========================================================
    # NOT FOUND
    # ========================================================

    if not transfer:

        message = (
            "Transfer not found or you do not "
            "have permission to delete it."
        )

        if request.method == "POST":
            return redirect_back_error(message)

        return json_error(message, 404)

    # ========================================================
    # TRANSFER TYPE
    # ========================================================

    transfer_type = (
        transfer.get(
            "transfer_type"
        )
        or "account_to_account"
    )

    # ========================================================
    # AMOUNT
    # ========================================================

    try:

        amount = float(
            transfer.get("amount") or 0
        )

    except Exception:

        amount = 0.0

    if amount <= 0:

        message = (
            "Invalid transfer amount."
        )

        if request.method == "POST":
            return redirect_back_error(message)

        return json_error(message, 400)

    # ========================================================
    # IDS
    # ========================================================

    from_account_id = transfer.get(
        "from_account"
    )

    to_account_id = transfer.get(
        "to_account"
    )

    from_saving_id = transfer.get(
        "from_saving"
    )

    to_saving_id = transfer.get(
        "to_saving"
    )

    # ========================================================
    # NORMALIZE OBJECT ID
    # ========================================================

    def normalize_object_id(value):

        if not value:
            return None

        try:

            if isinstance(
                value,
                ObjectId
            ):
                return value

            return ObjectId(
                str(value)
            )

        except Exception:

            return None

    from_account_obj = (
        normalize_object_id(
            from_account_id
        )
    )

    to_account_obj = (
        normalize_object_id(
            to_account_id
        )
    )

    from_saving_obj = (
        normalize_object_id(
            from_saving_id
        )
    )

    to_saving_obj = (
        normalize_object_id(
            to_saving_id
        )
    )

    # ========================================================
    # TIMESTAMP
    # ========================================================

    now = datetime.utcnow()

    # ========================================================
    # KEEP TRACK OF CHANGES
    # ========================================================

    balance_changes = []

    history_deleted = 0
    transactions_deleted = 0

    try:

        # ====================================================
        # ACCOUNT -> SAVING
        # ====================================================
        #
        # Original:
        #
        # Account  - amount
        # Saving   + amount
        #
        # DELETE:
        #
        # Account  + amount
        # Saving   - amount
        # ====================================================

        if transfer_type == "account_to_saving":

            if not from_account_obj:

                message = (
                    "Source account information "
                    "is missing."
                )

                if request.method == "POST":
                    return redirect_back_error(message)

                return json_error(message, 400)

            if not to_saving_obj:

                message = (
                    "Destination saving information "
                    "is missing."
                )

                if request.method == "POST":
                    return redirect_back_error(message)

                return json_error(message, 400)

            # ------------------------------------------------
            # GET CURRENT RECORDS
            # ------------------------------------------------

            account = (
                mongo.db.accounts.find_one({
                    "_id":
                    from_account_obj
                })
            )

            saving = (
                mongo.db.savings.find_one({
                    "_id":
                    to_saving_obj
                })
            )

            if not account:

                message = (
                    "Source account no longer exists."
                )

                if request.method == "POST":
                    return redirect_back_error(message)

                return json_error(message, 400)

            if not saving:

                message = (
                    "Destination saving no longer exists."
                )

                if request.method == "POST":
                    return redirect_back_error(message)

                return json_error(message, 400)

            # ------------------------------------------------
            # REVERSE ACCOUNT
            # ------------------------------------------------

            account_result = (
                mongo.db.accounts.update_one(

                    {
                        "_id":
                        from_account_obj
                    },

                    {
                        "$inc": {
                            "balance":
                            amount
                        },

                        "$set": {
                            "updated_at":
                            now
                        }
                    }
                )
            )

            if account_result.modified_count != 1:

                raise Exception(
                    "Failed to restore source account balance."
                )

            balance_changes.append({
                "collection": "accounts",
                "id": from_account_obj,
                "operation": "restore",
                "amount": amount
            })

            # ------------------------------------------------
            # REVERSE SAVING
            # ------------------------------------------------

            saving_result = (
                mongo.db.savings.update_one(

                    {
                        "_id":
                        to_saving_obj
                    },

                    {
                        "$inc": {
                            "current_balance":
                            -amount
                        },

                        "$set": {
                            "updated_at":
                            now
                        }
                    }
                )
            )

            if saving_result.modified_count != 1:

                # Restore account if saving failed
                mongo.db.accounts.update_one(
                    {
                        "_id":
                        from_account_obj
                    },
                    {
                        "$inc": {
                            "balance":
                            -amount
                        },
                        "$set": {
                            "updated_at":
                            now
                        }
                    }
                )

                raise Exception(
                    "Failed to restore saving balance."
                )

            balance_changes.append({
                "collection": "savings",
                "id": to_saving_obj,
                "operation": "restore",
                "amount": -amount
            })

        # ====================================================
        # SAVING -> ACCOUNT
        # ====================================================
        #
        # Original:
        #
        # Saving  - amount
        # Account + amount
        #
        # DELETE:
        #
        # Saving  + amount
        # Account - amount
        # ====================================================

        elif transfer_type == "saving_to_account":

            if not from_saving_obj:

                message = (
                    "Source saving information "
                    "is missing."
                )

                if request.method == "POST":
                    return redirect_back_error(message)

                return json_error(message, 400)

            if not to_account_obj:

                message = (
                    "Destination account information "
                    "is missing."
                )

                if request.method == "POST":
                    return redirect_back_error(message)

                return json_error(message, 400)

            # ------------------------------------------------
            # GET CURRENT RECORDS
            # ------------------------------------------------

            saving = (
                mongo.db.savings.find_one({
                    "_id":
                    from_saving_obj
                })
            )

            account = (
                mongo.db.accounts.find_one({
                    "_id":
                    to_account_obj
                })
            )

            if not saving:

                message = (
                    "Source saving no longer exists."
                )

                if request.method == "POST":
                    return redirect_back_error(message)

                return json_error(message, 400)

            if not account:

                message = (
                    "Destination account no longer exists."
                )

                if request.method == "POST":
                    return redirect_back_error(message)

                return json_error(message, 400)

            # ------------------------------------------------
            # REVERSE SAVING
            # ------------------------------------------------

            saving_result = (
                mongo.db.savings.update_one(

                    {
                        "_id":
                        from_saving_obj
                    },

                    {
                        "$inc": {
                            "current_balance":
                            amount
                        },

                        "$set": {
                            "updated_at":
                            now
                        }
                    }
                )
            )

            if saving_result.modified_count != 1:

                raise Exception(
                    "Failed to restore source saving balance."
                )

            balance_changes.append({
                "collection": "savings",
                "id": from_saving_obj,
                "operation": "restore",
                "amount": amount
            })

            # ------------------------------------------------
            # REVERSE ACCOUNT
            # ------------------------------------------------

            account_result = (
                mongo.db.accounts.update_one(

                    {
                        "_id":
                        to_account_obj
                    },

                    {
                        "$inc": {
                            "balance":
                            -amount
                        },

                        "$set": {
                            "updated_at":
                            now
                        }
                    }
                )
            )

            if account_result.modified_count != 1:

                # Restore saving if account failed
                mongo.db.savings.update_one(
                    {
                        "_id":
                        from_saving_obj
                    },
                    {
                        "$inc": {
                            "current_balance":
                            -amount
                        },
                        "$set": {
                            "updated_at":
                            now
                        }
                    }
                )

                raise Exception(
                    "Failed to restore destination account balance."
                )

            balance_changes.append({
                "collection": "accounts",
                "id": to_account_obj,
                "operation": "restore",
                "amount": -amount
            })

        # ====================================================
        # ACCOUNT -> ACCOUNT
        # ====================================================
        #
        # Original:
        #
        # From Account - amount
        # To Account   + amount
        #
        # DELETE:
        #
        # From Account + amount
        # To Account   - amount
        # ====================================================

        else:

            if not from_account_obj:

                message = (
                    "Source account information "
                    "is missing."
                )

                if request.method == "POST":
                    return redirect_back_error(message)

                return json_error(message, 400)

            if not to_account_obj:

                message = (
                    "Destination account information "
                    "is missing."
                )

                if request.method == "POST":
                    return redirect_back_error(message)

                return json_error(message, 400)

            # ------------------------------------------------
            # GET ACCOUNTS
            # ------------------------------------------------

            from_account = (
                mongo.db.accounts.find_one({
                    "_id":
                    from_account_obj
                })
            )

            to_account = (
                mongo.db.accounts.find_one({
                    "_id":
                    to_account_obj
                })
            )

            if not from_account:

                message = (
                    "Source account no longer exists."
                )

                if request.method == "POST":
                    return redirect_back_error(message)

                return json_error(message, 400)

            if not to_account:

                message = (
                    "Destination account no longer exists."
                )

                if request.method == "POST":
                    return redirect_back_error(message)

                return json_error(message, 400)

            # ------------------------------------------------
            # REVERSE SOURCE ACCOUNT
            # ------------------------------------------------

            from_result = (
                mongo.db.accounts.update_one(

                    {
                        "_id":
                        from_account_obj
                    },

                    {
                        "$inc": {
                            "balance":
                            amount
                        },

                        "$set": {
                            "updated_at":
                            now
                        }
                    }
                )
            )

            if from_result.modified_count != 1:

                raise Exception(
                    "Failed to restore source account balance."
                )

            balance_changes.append({
                "collection": "accounts",
                "id": from_account_obj,
                "operation": "restore",
                "amount": amount
            })

            # ------------------------------------------------
            # REVERSE DESTINATION ACCOUNT
            # ------------------------------------------------

            to_result = (
                mongo.db.accounts.update_one(

                    {
                        "_id":
                        to_account_obj
                    },

                    {
                        "$inc": {
                            "balance":
                            -amount
                        },

                        "$set": {
                            "updated_at":
                            now
                        }
                    }
                )
            )

            if to_result.modified_count != 1:

                # Roll source back
                mongo.db.accounts.update_one(
                    {
                        "_id":
                        from_account_obj
                    },
                    {
                        "$inc": {
                            "balance":
                            -amount
                        },
                        "$set": {
                            "updated_at":
                            now
                        }
                    }
                )

                raise Exception(
                    "Failed to restore destination account balance."
                )

            balance_changes.append({
                "collection": "accounts",
                "id": to_account_obj,
                "operation": "restore",
                "amount": -amount
            })

        # ====================================================
        # DELETE SAVING TRANSACTION HISTORY
        # ====================================================

        history_query = {
            "$or": [
                {
                    "transfer_id":
                    transfer_obj_id
                },
                {
                    "transfer_id":
                    str(transfer_obj_id)
                }
            ]
        }

        history_result = (
            mongo.db.saving_transactions.delete_many(
                history_query
            )
        )

        history_deleted = (
            history_result.deleted_count
        )

        # ====================================================
        # DELETE GENERAL TRANSACTION HISTORY
        # ====================================================

        transaction_conditions = [
            {
                "transfer_id":
                transfer_obj_id
            },
            {
                "transfer_id":
                str(transfer_obj_id)
            }
        ]

        # ----------------------------------------------------
        # ADD REFERENCE MATCHES ONLY IF THEY EXIST
        # ----------------------------------------------------

        reference_no = transfer.get(
            "reference_no"
        )

        reference = transfer.get(
            "reference"
        )

        if reference_no:

            transaction_conditions.append({
                "reference_no":
                reference_no
            })

        if reference:

            transaction_conditions.append({
                "reference":
                reference
            })

        transaction_result = (
            mongo.db.transactions.delete_many({

                "$or":
                transaction_conditions

            })
        )

        transactions_deleted = (
            transaction_result.deleted_count
        )

        # ====================================================
        # DELETE MAIN TRANSFER
        # ====================================================

        delete_result = (
            mongo.db.account_transfers.delete_one({

                "_id":
                transfer_obj_id

            })
        )

        # ====================================================
        # VERIFY MAIN DELETE
        # ====================================================

        if delete_result.deleted_count != 1:

            # ------------------------------------------------
            # IMPORTANT:
            # Main transfer was not deleted.
            # Restore balances because operation failed.
            # ------------------------------------------------

            if transfer_type == "account_to_saving":

                mongo.db.accounts.update_one(
                    {
                        "_id":
                        from_account_obj
                    },
                    {
                        "$inc": {
                            "balance":
                            -amount
                        }
                    }
                )

                mongo.db.savings.update_one(
                    {
                        "_id":
                        to_saving_obj
                    },
                    {
                        "$inc": {
                            "current_balance":
                            amount
                        }
                    }
                )

            elif transfer_type == "saving_to_account":

                mongo.db.savings.update_one(
                    {
                        "_id":
                        from_saving_obj
                    },
                    {
                        "$inc": {
                            "current_balance":
                            -amount
                        }
                    }
                )

                mongo.db.accounts.update_one(
                    {
                        "_id":
                        to_account_obj
                    },
                    {
                        "$inc": {
                            "balance":
                            amount
                        }
                    }
                )

            else:

                mongo.db.accounts.update_one(
                    {
                        "_id":
                        from_account_obj
                    },
                    {
                        "$inc": {
                            "balance":
                            -amount
                        }
                    }
                )

                mongo.db.accounts.update_one(
                    {
                        "_id":
                        to_account_obj
                    },
                    {
                        "$inc": {
                            "balance":
                            amount
                        }
                    }
                )

            raise Exception(
                "Transfer could not be deleted."
            )

        # ====================================================
        # SUCCESS MESSAGE
        # ====================================================

        success_message = (
            "Transfer deleted successfully. "
            "Balances and financial movements "
            "have been restored."
        )

        # ====================================================
        # POST -> FLASH MESSAGE
        # ====================================================

        if request.method == "POST":

            flash(
                success_message,
                "success"
            )

            return redirect(
                url_for(
                    "main.all_account_transfers"
                )
            )

        # ====================================================
        # DELETE -> JSON
        # ====================================================

        return jsonify({

            "success":
            True,

            "message":
            success_message,

            "transfer_id":
            str(transfer_obj_id),

            "transfer_type":
            transfer_type,

            "amount":
            amount,

            "saving_history_deleted":
            history_deleted,

            "transaction_history_deleted":
            transactions_deleted

        }), 200

    # ========================================================
    # EXCEPTION
    # ========================================================

    except Exception as e:

        print(
            "DELETE ACCOUNT TRANSFER ERROR:",
            repr(e)
        )

        error_message = (
            "Failed to delete transfer. "
            "No complete deletion was performed."
        )

        # ----------------------------------------------------
        # POST -> FLASH
        # ----------------------------------------------------

        if request.method == "POST":

            flash(
                error_message,
                "danger"
            )

            return redirect(
                url_for(
                    "main.all_account_transfers"
                )
            )

        # ----------------------------------------------------
        # DELETE -> JSON
        # ----------------------------------------------------

        return jsonify({

            "success":
            False,

            "error":
            error_message,

            "details":
            str(e)

        }), 500


# ============================================================
# ALL ACCOUNT TRANSFERS
# ============================================================

@bp.route("/account/transfers")
@login_required
def all_account_transfers():

    from bson import ObjectId
    from datetime import datetime
    import json

    # ============================================================
    # CURRENT USER
    # ============================================================

    current_user_id = str(current_user.id)

    try:
        current_user_object_id = ObjectId(current_user_id)
    except Exception:
        current_user_object_id = None

    # ============================================================
    # USER FILTER
    # ============================================================

    if current_user.role == UserRole.superadmin.value:

        transfer_query = {}

    else:

        user_conditions = [
            {"user_id": current_user_id}
        ]

        if current_user_object_id:
            user_conditions.append(
                {"user_id": current_user_object_id}
            )

        transfer_query = {
            "$or": user_conditions
        }

    # ============================================================
    # GET TRANSFERS
    # ============================================================

    transfers_raw = list(
        mongo.db.account_transfers.find(
            transfer_query
        ).sort(
            "created_at",
            -1
        )
    )

    # ============================================================
    # HELPERS
    # ============================================================

    def safe_float(value, default=0.0):

        try:

            if value is None:
                return default

            return float(value)

        except (TypeError, ValueError):

            return default

    def safe_string(value, default=""):

        if value is None:
            return default

        return str(value)

    def get_balance(transfer, *keys):

        for key in keys:

            if key in transfer:

                value = transfer.get(key)

                if value is not None:
                    return safe_float(value)

        return 0.0

    def format_datetime(value):

        if not value:
            return {
                "iso": "",
                "display": "-"
            }

        # Mongo datetime
        if isinstance(value, datetime):

            return {
                "iso": value.isoformat(),
                "display": value.strftime(
                    "%d %b %Y, %I:%M:%S %p"
                )
            }

        # String datetime
        value_string = str(value)

        try:

            parsed = datetime.fromisoformat(
                value_string.replace("Z", "+00:00")
            )

            return {
                "iso": parsed.isoformat(),
                "display": parsed.strftime(
                    "%d %b %Y, %I:%M:%S %p"
                )
            }

        except Exception:

            return {
                "iso": value_string,
                "display": value_string
            }

    # ============================================================
    # COLLECT ACCOUNT IDS
    # ============================================================

    account_ids = set()

    for transfer in transfers_raw:

        for key in [
            "from_account",
            "to_account"
        ]:

            value = transfer.get(key)

            if value:

                account_ids.add(
                    str(value)
                )

    # ============================================================
    # LOAD ACCOUNTS
    # ============================================================

    accounts_map = {}

    for account_id in account_ids:

        try:

            account = mongo.db.accounts.find_one({
                "_id": ObjectId(account_id)
            })

            if account:

                accounts_map[
                    str(account["_id"])
                ] = account

        except Exception:

            continue

    # ============================================================
    # COLLECT SAVING IDS
    # ============================================================

    saving_ids = set()

    for transfer in transfers_raw:

        for key in [
            "from_saving",
            "to_saving"
        ]:

            value = transfer.get(key)

            if value:

                saving_ids.add(
                    str(value)
                )

    # ============================================================
    # LOAD SAVINGS
    # ============================================================

    savings_map = {}

    for saving_id in saving_ids:

        try:

            saving = mongo.db.savings.find_one({
                "_id": ObjectId(saving_id)
            })

            if saving:

                savings_map[
                    str(saving["_id"])
                ] = saving

        except Exception:

            continue

    # ============================================================
    # PREPARE TRANSFERS
    # ============================================================

    transfers = []

    for transfer in transfers_raw:

        # ========================================================
        # BASIC
        # ========================================================

        transfer_id = str(
            transfer.get("_id")
        )

        transfer_type = (
            transfer.get("transfer_type")
            or "account_to_account"
        )

        direction = (
            transfer.get("direction")
            or ""
        )

        amount = safe_float(
            transfer.get("amount")
        )

        currency = (
            transfer.get("currency")
            or "USD"
        )

        status = (
            transfer.get("status")
            or "completed"
        )

        status_display = (
            str(status).replace("_", " ").title()
        )

        reference = (
            transfer.get("reference_no")
            or transfer.get("reference")
            or transfer.get("transfer_no")
            or transfer_id
        )

        description = (
            transfer.get("description")
            or ""
        )

        # ========================================================
        # DATE
        # ========================================================

        date_info = format_datetime(
            transfer.get("created_at")
        )

        # ========================================================
        # DEFAULT
        # ========================================================

        transaction_type = "Transfer"

        from_type = "Account"
        to_type = "Account"

        from_name = "Unknown Account"
        to_name = "Unknown Account"

        # ========================================================
        # BALANCES
        # ========================================================

        from_balance_before = 0.0
        from_balance_after = 0.0

        to_balance_before = 0.0
        to_balance_after = 0.0

        from_account_balance_before = 0.0
        from_account_balance_after = 0.0

        to_account_balance_before = 0.0
        to_account_balance_after = 0.0

        from_saving_balance_before = 0.0
        from_saving_balance_after = 0.0

        to_saving_balance_before = 0.0
        to_saving_balance_after = 0.0

        # ========================================================
        # ACCOUNT → SAVING
        # ========================================================

        if transfer_type == "account_to_saving":

            transaction_type = "Deposit"

            from_type = "Account"
            to_type = "Saving"

            # ----------------------------------------------------
            # FROM ACCOUNT
            # ----------------------------------------------------

            from_account_id = transfer.get(
                "from_account"
            )

            account = None

            if from_account_id:

                account = accounts_map.get(
                    str(from_account_id)
                )

            if account:

                from_name = (
                    transfer.get(
                        "from_account_name"
                    )
                    or account.get("name")
                    or account.get("account_name")
                    or "Unknown Account"
                )

            else:

                from_name = (
                    transfer.get(
                        "from_account_name"
                    )
                    or "Unknown Account"
                )

            # ----------------------------------------------------
            # TO SAVING
            # ----------------------------------------------------

            to_saving_id = transfer.get(
                "to_saving"
            )

            saving = None

            if to_saving_id:

                saving = savings_map.get(
                    str(to_saving_id)
                )

            if saving:

                to_name = (
                    transfer.get(
                        "to_saving_name"
                    )
                    or saving.get("name")
                    or saving.get("saving_name")
                    or saving.get("title")
                    or "Unknown Saving"
                )

            else:

                to_name = (
                    transfer.get(
                        "to_saving_name"
                    )
                    or "Unknown Saving"
                )

            # ----------------------------------------------------
            # CURRENCY
            # ----------------------------------------------------

            currency = (
                transfer.get("currency")
                or (
                    account.get("currency")
                    if account
                    else None
                )
                or "USD"
            )

            # ----------------------------------------------------
            # ACCOUNT BALANCE
            # ----------------------------------------------------

            from_account_balance_before = get_balance(
                transfer,
                "account_balance_before",
                "from_account_balance_before",
                "source_account_balance_before"
            )

            from_account_balance_after = get_balance(
                transfer,
                "account_balance_after",
                "from_account_balance_after",
                "source_account_balance_after"
            )

            # ----------------------------------------------------
            # SAVING BALANCE
            # ----------------------------------------------------

            to_saving_balance_before = get_balance(
                transfer,
                "saving_balance_before",
                "to_saving_balance_before",
                "destination_saving_balance_before"
            )

            to_saving_balance_after = get_balance(
                transfer,
                "saving_balance_after",
                "to_saving_balance_after",
                "destination_saving_balance_after"
            )

            from_balance_before = (
                from_account_balance_before
            )

            from_balance_after = (
                from_account_balance_after
            )

            to_balance_before = (
                to_saving_balance_before
            )

            to_balance_after = (
                to_saving_balance_after
            )

        # ========================================================
        # SAVING → ACCOUNT
        # ========================================================

        elif transfer_type == "saving_to_account":

            transaction_type = "Withdraw"

            from_type = "Saving"
            to_type = "Account"

            # ----------------------------------------------------
            # FROM SAVING
            # ----------------------------------------------------

            from_saving_id = transfer.get(
                "from_saving"
            )

            saving = None

            if from_saving_id:

                saving = savings_map.get(
                    str(from_saving_id)
                )

            if saving:

                from_name = (
                    transfer.get(
                        "from_saving_name"
                    )
                    or saving.get("name")
                    or saving.get("saving_name")
                    or saving.get("title")
                    or "Unknown Saving"
                )

            else:

                from_name = (
                    transfer.get(
                        "from_saving_name"
                    )
                    or "Unknown Saving"
                )

            # ----------------------------------------------------
            # TO ACCOUNT
            # ----------------------------------------------------

            to_account_id = transfer.get(
                "to_account"
            )

            account = None

            if to_account_id:

                account = accounts_map.get(
                    str(to_account_id)
                )

            if account:

                to_name = (
                    transfer.get(
                        "to_account_name"
                    )
                    or account.get("name")
                    or account.get("account_name")
                    or "Unknown Account"
                )

            else:

                to_name = (
                    transfer.get(
                        "to_account_name"
                    )
                    or "Unknown Account"
                )

            # ----------------------------------------------------
            # CURRENCY
            # ----------------------------------------------------

            currency = (
                transfer.get("currency")
                or (
                    saving.get("currency")
                    if saving
                    else None
                )
                or (
                    account.get("currency")
                    if account
                    else None
                )
                or "USD"
            )

            # ----------------------------------------------------
            # SAVING BALANCE
            # ----------------------------------------------------

            from_saving_balance_before = get_balance(
                transfer,
                "saving_balance_before",
                "from_saving_balance_before",
                "source_saving_balance_before"
            )

            from_saving_balance_after = get_balance(
                transfer,
                "saving_balance_after",
                "from_saving_balance_after",
                "source_saving_balance_after"
            )

            # ----------------------------------------------------
            # ACCOUNT BALANCE
            # ----------------------------------------------------

            to_account_balance_before = get_balance(
                transfer,
                "account_balance_before",
                "to_account_balance_before",
                "destination_account_balance_before"
            )

            to_account_balance_after = get_balance(
                transfer,
                "account_balance_after",
                "to_account_balance_after",
                "destination_account_balance_after"
            )

            from_balance_before = (
                from_saving_balance_before
            )

            from_balance_after = (
                from_saving_balance_after
            )

            to_balance_before = (
                to_account_balance_before
            )

            to_balance_after = (
                to_account_balance_after
            )

        # ========================================================
        # ACCOUNT → ACCOUNT
        # ========================================================

        else:

            transaction_type = "Transfer"

            from_type = "Account"
            to_type = "Account"

            # ----------------------------------------------------
            # FROM ACCOUNT
            # ----------------------------------------------------

            from_account_id = transfer.get(
                "from_account"
            )

            from_account = None

            if from_account_id:

                from_account = accounts_map.get(
                    str(from_account_id)
                )

            if from_account:

                from_name = (
                    transfer.get(
                        "from_account_name"
                    )
                    or from_account.get("name")
                    or from_account.get("account_name")
                    or "Unknown Account"
                )

            else:

                from_name = (
                    transfer.get(
                        "from_account_name"
                    )
                    or "Unknown Account"
                )

            # ----------------------------------------------------
            # TO ACCOUNT
            # ----------------------------------------------------

            to_account_id = transfer.get(
                "to_account"
            )

            to_account = None

            if to_account_id:

                to_account = accounts_map.get(
                    str(to_account_id)
                )

            if to_account:

                to_name = (
                    transfer.get(
                        "to_account_name"
                    )
                    or to_account.get("name")
                    or to_account.get("account_name")
                    or "Unknown Account"
                )

            else:

                to_name = (
                    transfer.get(
                        "to_account_name"
                    )
                    or "Unknown Account"
                )

            # ----------------------------------------------------
            # CURRENCY
            # ----------------------------------------------------

            currency = (
                transfer.get("currency")
                or (
                    from_account.get("currency")
                    if from_account
                    else None
                )
                or (
                    to_account.get("currency")
                    if to_account
                    else None
                )
                or "USD"
            )

            # ----------------------------------------------------
            # SOURCE BALANCE
            # ----------------------------------------------------

            from_account_balance_before = get_balance(
                transfer,
                "from_account_balance_before",
                "source_account_balance_before",
                "account_balance_before"
            )

            from_account_balance_after = get_balance(
                transfer,
                "from_account_balance_after",
                "source_account_balance_after",
                "account_balance_after"
            )

            # ----------------------------------------------------
            # DESTINATION BALANCE
            # ----------------------------------------------------

            to_account_balance_before = get_balance(
                transfer,
                "to_account_balance_before",
                "destination_account_balance_before",
                "receiver_account_balance_before"
            )

            to_account_balance_after = get_balance(
                transfer,
                "to_account_balance_after",
                "destination_account_balance_after",
                "receiver_account_balance_after"
            )

            from_balance_before = (
                from_account_balance_before
            )

            from_balance_after = (
                from_account_balance_after
            )

            to_balance_before = (
                to_account_balance_before
            )

            to_balance_after = (
                to_account_balance_after
            )

        # ========================================================
        # ORIGINAL BALANCE FIELDS
        # ========================================================

        saving_balance_before = get_balance(
            transfer,
            "saving_balance_before"
        )

        saving_balance_after = get_balance(
            transfer,
            "saving_balance_after"
        )

        account_balance_before = get_balance(
            transfer,
            "account_balance_before"
        )

        account_balance_after = get_balance(
            transfer,
            "account_balance_after"
        )

        # ========================================================
        # USER
        # ========================================================

        raw_transfer_user_id = transfer.get(
            "user_id"
        )

        transfer_user_id = (
            str(raw_transfer_user_id)
            if raw_transfer_user_id is not None
            else ""
        )

        # ========================================================
        # TRANSFER DATA
        # ========================================================

        transfer_data = {

            # ====================================================
            # IDENTITY
            # ====================================================

            "id": transfer_id,
            "_id": transfer_id,
            "user_id": transfer_user_id,

            # ====================================================
            # TYPE
            # ====================================================

            "transfer_type": transfer_type,
            "transaction_type": transaction_type,
            "direction": direction,

            # ====================================================
            # FROM
            # ====================================================

            "from_name": from_name,
            "from_type": from_type,

            "from_account_name": (
                from_name
                if from_type == "Account"
                else (
                    transfer.get(
                        "from_account_name"
                    )
                    or ""
                )
            ),

            "from_saving_name": (
                from_name
                if from_type == "Saving"
                else (
                    transfer.get(
                        "from_saving_name"
                    )
                    or ""
                )
            ),

            # ====================================================
            # TO
            # ====================================================

            "to_name": to_name,
            "to_type": to_type,

            "to_account_name": (
                to_name
                if to_type == "Account"
                else (
                    transfer.get(
                        "to_account_name"
                    )
                    or ""
                )
            ),

            "to_saving_name": (
                to_name
                if to_type == "Saving"
                else (
                    transfer.get(
                        "to_saving_name"
                    )
                    or ""
                )
            ),

            # ====================================================
            # AMOUNT
            # ====================================================

            "amount": amount,
            "currency": currency,

            # ====================================================
            # STATUS
            # ====================================================

            "status": status_display,

            # ====================================================
            # REFERENCE
            # ====================================================

            "reference_no": safe_string(reference),
            "reference": safe_string(reference),

            # ====================================================
            # DESCRIPTION
            # ====================================================

            "description": description,

            # ====================================================
            # DATE
            # ====================================================

            "created_at": date_info["iso"],

            "created_at_display": date_info["display"],

            # ====================================================
            # GENERIC BALANCES
            # ====================================================

            "from_balance_before":
                from_balance_before,

            "from_balance_after":
                from_balance_after,

            "to_balance_before":
                to_balance_before,

            "to_balance_after":
                to_balance_after,

            # ====================================================
            # ACCOUNT BALANCES
            # ====================================================

            "from_account_balance_before":
                from_account_balance_before,

            "from_account_balance_after":
                from_account_balance_after,

            "to_account_balance_before":
                to_account_balance_before,

            "to_account_balance_after":
                to_account_balance_after,

            # ====================================================
            # SAVING BALANCES
            # ====================================================

            "from_saving_balance_before":
                from_saving_balance_before,

            "from_saving_balance_after":
                from_saving_balance_after,

            "to_saving_balance_before":
                to_saving_balance_before,

            "to_saving_balance_after":
                to_saving_balance_after,

            # ====================================================
            # ORIGINAL DATABASE FIELDS
            # ====================================================

            "saving_balance_before":
                saving_balance_before,

            "saving_balance_after":
                saving_balance_after,

            "account_balance_before":
                account_balance_before,

            "account_balance_after":
                account_balance_after,

            # ====================================================
            # OBJECT IDS AS STRINGS
            # Useful for modal/detail page
            # ====================================================

            "from_account_id": (
                str(transfer.get("from_account"))
                if transfer.get("from_account")
                else ""
            ),

            "to_account_id": (
                str(transfer.get("to_account"))
                if transfer.get("to_account")
                else ""
            ),

            "from_saving_id": (
                str(transfer.get("from_saving"))
                if transfer.get("from_saving")
                else ""
            ),

            "to_saving_id": (
                str(transfer.get("to_saving"))
                if transfer.get("to_saving")
                else ""
            )
        }

        transfers.append(
            transfer_data
        )

    # ============================================================
    # SUMMARY
    # ============================================================

    total_transfers = len(transfers)

    total_amount = sum(
        safe_float(
            item.get("amount")
        )
        for item in transfers
    )

    total_deposits = sum(
        safe_float(item.get("amount"))
        for item in transfers
        if item.get("transaction_type") == "Deposit"
    )

    total_withdrawals = sum(
        safe_float(item.get("amount"))
        for item in transfers
        if item.get("transaction_type") == "Withdraw"
    )

    total_account_transfers = sum(
        safe_float(item.get("amount"))
        for item in transfers
        if item.get("transaction_type") == "Transfer"
    )

    deposit_count = sum(
        1
        for item in transfers
        if item.get("transaction_type") == "Deposit"
    )

    withdrawal_count = sum(
        1
        for item in transfers
        if item.get("transaction_type") == "Withdraw"
    )

    account_transfer_count = sum(
        1
        for item in transfers
        if item.get("transaction_type") == "Transfer"
    )

    # ============================================================
    # JSON SAFE
    # ============================================================

    transfers_json = json.dumps(
        transfers,
        default=str
    )

    # ============================================================
    # RENDER
    # ============================================================

    return render_template(

        "backend/pages/components/accounts/account_transfers.html",

        transfers=transfers,

        transfers_json=transfers_json,

        total_transfers=total_transfers,

        total_amount=total_amount,

        total_deposits=total_deposits,

        total_withdrawals=total_withdrawals,

        total_account_transfers=total_account_transfers,

        deposit_count=deposit_count,

        withdrawal_count=withdrawal_count,

        account_transfer_count=account_transfer_count
    )



@bp.route("/saving-topup", methods=["POST"])
@login_required
def saving_topup():

    from bson import ObjectId
    from datetime import datetime
    from flask import request, jsonify
    import math
    import secrets

    # ============================================================
    # RESPONSE HELPERS
    # ============================================================

    def error_response(message, status=400):
        return jsonify({
            "success": False,
            "error": message,
            "message": message
        }), status

    # ============================================================
    # GET JSON DATA
    # ============================================================

    data = request.get_json(silent=True)

    if not data:
        return error_response(
            "No data provided.",
            400
        )

    saving_id = data.get("saving_id")
    account_id = data.get("account_id")
    amount = data.get("amount")

    # ============================================================
    # REQUIRED FIELDS
    # ============================================================

    if not saving_id:
        return error_response(
            "Saving account is required.",
            400
        )

    if not account_id:
        return error_response(
            "Source account is required.",
            400
        )

    if amount is None or amount == "":
        return error_response(
            "Amount is required.",
            400
        )

    # ============================================================
    # VALIDATE AMOUNT
    # ============================================================

    try:
        amount = float(amount)

    except (TypeError, ValueError):
        return error_response(
            "Invalid amount.",
            400
        )

    if not math.isfinite(amount):
        return error_response(
            "Invalid amount.",
            400
        )

    if amount <= 0:
        return error_response(
            "Amount must be greater than zero.",
            400
        )

    # ============================================================
    # OBJECT IDS
    # ============================================================

    try:

        saving_obj_id = ObjectId(
            str(saving_id)
        )

        account_obj_id = ObjectId(
            str(account_id)
        )

    except Exception:

        return error_response(
            "Invalid account or saving ID.",
            400
        )

    # ============================================================
    # CURRENT USER
    # ============================================================

    user_id = str(
        current_user.id
    )

    try:

        user_object_id = ObjectId(
            user_id
        )

    except Exception:

        user_object_id = None

    # ============================================================
    # USER FILTER
    # Supports STRING + ObjectId
    # ============================================================

    if current_user.role == UserRole.superadmin.value:

        user_filter = {}

    else:

        owner_conditions = [
            {
                "user_id": user_id
            }
        ]

        if user_object_id:

            owner_conditions.append({
                "user_id": user_object_id
            })

        user_filter = {
            "$or": owner_conditions
        }

    # ============================================================
    # GET SAVING
    # ============================================================

    saving_query = {
        "_id": saving_obj_id
    }

    saving_query.update(
        user_filter
    )

    saving = mongo.db.savings.find_one(
        saving_query
    )

    if not saving:

        return error_response(
            "Saving not found.",
            404
        )

    # ============================================================
    # GET SOURCE ACCOUNT
    # ============================================================

    account_query = {
        "_id": account_obj_id
    }

    account_query.update(
        user_filter
    )

    account = mongo.db.accounts.find_one(
        account_query
    )

    if not account:

        return error_response(
            "Source account not found.",
            404
        )

    # ============================================================
    # ACCOUNT BALANCE BEFORE
    # ============================================================

    try:

        account_balance_before = float(
            account.get("balance", 0) or 0
        )

    except (TypeError, ValueError):

        account_balance_before = 0.0

    # ============================================================
    # SAVING BALANCE BEFORE
    # ============================================================

    try:

        saving_balance_before = float(
            saving.get("current_balance", 0) or 0
        )

    except (TypeError, ValueError):

        saving_balance_before = 0.0

    # ============================================================
    # CURRENCY
    # ============================================================

    account_currency = (
        account.get("currency")
        or "USD"
    )

    saving_currency = (
        saving.get("currency")
        or account_currency
        or "USD"
    )

    # ============================================================
    # CURRENCY CHECK
    # ============================================================

    if (
        account_currency
        and saving_currency
        and account_currency != saving_currency
    ):

        return error_response(
            (
                "Currency mismatch. "
                f"Account uses {account_currency}, "
                f"Saving uses {saving_currency}."
            ),
            400
        )

    currency = (
        account_currency
        or saving_currency
        or "USD"
    )

    # ============================================================
    # CHECK ACCOUNT BALANCE
    # ============================================================

    if amount > account_balance_before:

        return error_response(
            (
                "Insufficient account balance. "
                f"Available "
                f"{currency} "
                f"{account_balance_before:,.2f}"
            ),
            400
        )

    # ============================================================
    # NAMES
    # ============================================================

    account_name = (
        account.get("name")
        or account.get("account_name")
        or "Unknown Account"
    )

    saving_name = (
        saving.get("name")
        or saving.get("saving_name")
        or saving.get("title")
        or "Unknown Saving"
    )

    # ============================================================
    # NEW BALANCES
    # ============================================================

    account_balance_after = (
        account_balance_before - amount
    )

    saving_balance_after = (
        saving_balance_before + amount
    )

    # ============================================================
    # REFERENCE
    # ============================================================

    reference_no = (
        "SAV-"
        + datetime.utcnow().strftime(
            "%Y%m%d%H%M%S"
        )
        + "-"
        + secrets.token_hex(3).upper()
    )

    # ============================================================
    # TIMESTAMP
    # ============================================================

    now = datetime.utcnow()

    # ============================================================
    # DEBUG BEFORE TRANSACTION
    # ============================================================

    print(
        "\n=================================================="
    )

    print(
        "SAVING TOPUP START"
    )

    print(
        "USER:",
        user_id
    )

    print(
        "REFERENCE:",
        reference_no
    )

    print(
        "FROM ACCOUNT:",
        account_name,
        str(account_obj_id)
    )

    print(
        "TO SAVING:",
        saving_name,
        str(saving_obj_id)
    )

    print(
        "AMOUNT:",
        amount,
        currency
    )

    print(
        "ACCOUNT BEFORE:",
        account_balance_before
    )

    print(
        "ACCOUNT AFTER:",
        account_balance_after
    )

    print(
        "SAVING BEFORE:",
        saving_balance_before
    )

    print(
        "SAVING AFTER:",
        saving_balance_after
    )

    print(
        "==================================================\n"
    )

    # ============================================================
    # STEP 1
    # ACCOUNT → REMOVE MONEY
    #
    # IMPORTANT:
    # Atomic balance check
    # ============================================================

    account_result = mongo.db.accounts.update_one(

        {
            "_id": account_obj_id,

            **user_filter,

            "balance": {
                "$gte": amount
            }
        },

        {
            "$inc": {
                "balance": -amount
            },

            "$set": {
                "updated_at": now
            }
        }

    )

    if account_result.modified_count != 1:

        return error_response(
            (
                "Failed to remove money from source "
                "account. The balance may have changed."
            ),
            400
        )

    # ============================================================
    # STEP 2
    # SAVING → ADD MONEY
    # ============================================================

    saving_result = mongo.db.savings.update_one(

        {
            "_id": saving_obj_id,

            **user_filter
        },

        {
            "$inc": {
                "current_balance": amount
            },

            "$set": {
                "updated_at": now
            }
        }

    )

    # ============================================================
    # ROLLBACK IF SAVING UPDATE FAILED
    # ============================================================

    if saving_result.modified_count != 1:

        mongo.db.accounts.update_one(

            {
                "_id": account_obj_id
            },

            {
                "$inc": {
                    "balance": amount
                },

                "$set": {
                    "updated_at": now
                }
            }

        )

        return error_response(
            (
                "Failed to add money to saving. "
                "Account balance has been restored."
            ),
            500
        )

    # ============================================================
    # STEP 3
    # CREATE MASTER TRANSFER RECORD
    # ============================================================

    try:

        transfer_document = {

            # ====================================================
            # IDENTIFICATION
            # ====================================================

            "user_id":
                user_id,

            "transfer_type":
                "account_to_saving",

            "direction":
                "deposit",

            # ====================================================
            # SOURCE ACCOUNT
            # ====================================================

            "from_account":
                account_obj_id,

            "from_account_name":
                account_name,

            "from_type":
                "Account",

            # ====================================================
            # DESTINATION SAVING
            # ====================================================

            "to_saving":
                saving_obj_id,

            "to_saving_name":
                saving_name,

            "to_type":
                "Saving",

            # ====================================================
            # AMOUNT
            # ====================================================

            "amount":
                amount,

            "currency":
                currency,

            # ====================================================
            # STATUS
            # ====================================================

            "status":
                "completed",

            # ====================================================
            # REFERENCE
            # ====================================================

            "reference":
                reference_no,

            "reference_no":
                reference_no,

            "transfer_no":
                reference_no,

            # ====================================================
            # DESCRIPTION
            # ====================================================

            "description":
                "Account to Saving deposit",

            "note":
                (
                    f"Deposit from "
                    f"{account_name} "
                    f"to "
                    f"{saving_name}"
                ),

            # ====================================================
            # ACCOUNT BALANCE AUDIT
            # ====================================================

            "account_balance_before":
                account_balance_before,

            "account_balance_after":
                account_balance_after,

            # ====================================================
            # SAVING BALANCE AUDIT
            # ====================================================

            "saving_balance_before":
                saving_balance_before,

            "saving_balance_after":
                saving_balance_after,

            # ====================================================
            # CREATED / UPDATED
            # ====================================================

            "created_at":
                now,

            "updated_at":
                now
        }

        # --------------------------------------------------------
        # INSERT MASTER
        # --------------------------------------------------------

        transfer_result = (
            mongo.db.account_transfers.insert_one(
                transfer_document
            )
        )

        transfer_id = (
            transfer_result.inserted_id
        )

    except Exception as transfer_error:

        print(
            "ACCOUNT TRANSFER INSERT ERROR:",
            repr(transfer_error)
        )

        # ========================================================
        # ROLLBACK SAVING
        # ========================================================

        mongo.db.savings.update_one(

            {
                "_id": saving_obj_id
            },

            {
                "$inc": {
                    "current_balance": -amount
                },

                "$set": {
                    "updated_at": now
                }
            }

        )

        # ========================================================
        # ROLLBACK ACCOUNT
        # ========================================================

        mongo.db.accounts.update_one(

            {
                "_id": account_obj_id
            },

            {
                "$inc": {
                    "balance": amount
                },

                "$set": {
                    "updated_at": now
                }
            }

        )

        return error_response(
            (
                "Transfer history could not be saved. "
                "The transaction was rolled back."
            ),
            500
        )

    # ============================================================
    # STEP 4
    # GENERAL TRANSACTION RECORD
    #
    # This allows the normal transactions report
    # to also understand the movement.
    # ============================================================

    try:

        transaction_document = {

            # ====================================================
            # USER
            # ====================================================

            "user_id":
                user_id,

            # ====================================================
            # ACCOUNT
            # ====================================================

            "account_id":
                account_obj_id,

            # ====================================================
            # TRANSACTION TYPE
            # ====================================================

            "transaction_type":
                "transfer_out",

            "type":
                "transfer_out",

            # ====================================================
            # AMOUNT
            # ====================================================

            "amount":
                amount,

            "currency":
                currency,

            # ====================================================
            # CATEGORY
            # ====================================================

            "category":
                "Saving Transfer",

            "item":
                f"Saving Deposit - {saving_name}",

            # ====================================================
            # DESCRIPTION
            # ====================================================

            "description":
                (
                    f"Transfer from "
                    f"{account_name} "
                    f"to "
                    f"{saving_name}"
                ),

            "note":
                (
                    f"Money moved from "
                    f"{account_name} "
                    f"to "
                    f"{saving_name}"
                ),

            # ====================================================
            # REFERENCE
            # ====================================================

            "reference_no":
                reference_no,

            "reference":
                reference_no,

            # ====================================================
            # TRANSFER ID
            # ====================================================

            "transfer_id":
                transfer_id,

            # ====================================================
            # ROUTE
            # ====================================================

            "from_account":
                account_obj_id,

            "from_account_name":
                account_name,

            "to_saving":
                saving_obj_id,

            "to_saving_name":
                saving_name,

            # ====================================================
            # BALANCE AUDIT
            # ====================================================

            "account_balance_before":
                account_balance_before,

            "account_balance_after":
                account_balance_after,

            "saving_balance_before":
                saving_balance_before,

            "saving_balance_after":
                saving_balance_after,

            # ====================================================
            # DATE
            # ====================================================

            "date":
                now,

            "created_at":
                now,

            "updated_at":
                now,

            "status":
                True
        }

        mongo.db.transactions.insert_one(
            transaction_document
        )

    except Exception as transaction_error:

        print(
            "GENERAL TRANSACTION INSERT ERROR:",
            repr(transaction_error)
        )

        # ========================================================
        # IMPORTANT
        #
        # Master transfer already exists.
        # Do NOT rollback balances here.
        #
        # The master account_transfers record is the
        # source of truth.
        # ========================================================

    # ============================================================
    # STEP 5
    # SAVING TRANSACTION HISTORY
    # ============================================================

    try:

        saving_transaction_document = {

            "user_id":
                user_id,

            "saving_id":
                saving_obj_id,

            "account_id":
                account_obj_id,

            "transaction_type":
                "deposit",

            "amount":
                amount,

            "currency":
                currency,

            "description":
                (
                    f"Deposit from "
                    f"{account_name}"
                ),

            "note":
                (
                    f"Saving deposit from "
                    f"{account_name}"
                ),

            "reference_no":
                reference_no,

            "reference":
                reference_no,

            "transfer_id":
                transfer_id,

            "from_account":
                account_obj_id,

            "from_account_name":
                account_name,

            "saving_balance_before":
                saving_balance_before,

            "saving_balance_after":
                saving_balance_after,

            "account_balance_before":
                account_balance_before,

            "account_balance_after":
                account_balance_after,

            "date":
                now,

            "created_at":
                now,

            "updated_at":
                now,

            "status":
                True
        }

        mongo.db.saving_transactions.insert_one(
            saving_transaction_document
        )

    except Exception as saving_history_error:

        print(
            "SAVING TRANSACTION HISTORY ERROR:",
            repr(saving_history_error)
        )

    # ============================================================
    # FINAL RESPONSE
    # ============================================================

    response_data = {

        "success":
            True,

        "message":
            (
                "Money moved from Account → Saving "
                "successfully."
            ),

        # ========================================================
        # MASTER TRANSFER
        # ========================================================

        "transfer_id":
            str(transfer_id),

        "transfer_type":
            "account_to_saving",

        "direction":
            "deposit",

        "status":
            "completed",

        # ========================================================
        # REFERENCE
        # ========================================================

        "reference":
            reference_no,

        "reference_no":
            reference_no,

        # ========================================================
        # SOURCE
        # ========================================================

        "from_account":
            account_name,

        "from_account_id":
            str(account_obj_id),

        "from_type":
            "Account",

        # ========================================================
        # DESTINATION
        # ========================================================

        "to_saving":
            saving_name,

        "to_saving_id":
            str(saving_obj_id),

        "to_type":
            "Saving",

        # ========================================================
        # MONEY
        # ========================================================

        "amount":
            amount,

        "currency":
            currency,

        # ========================================================
        # ACCOUNT AUDIT
        # ========================================================

        "account_balance_before":
            account_balance_before,

        "account_balance_after":
            account_balance_after,

        # ========================================================
        # SAVING AUDIT
        # ========================================================

        "saving_balance_before":
            saving_balance_before,

        "saving_balance_after":
            saving_balance_after,

        # ========================================================
        # DATE
        # ========================================================

        "created_at":
            now.isoformat()
    }

    # ============================================================
    # DEBUG FINAL REPORT
    # ============================================================

    print(
        "\n=================================================="
    )

    print(
        "SAVING TOPUP COMPLETED"
    )

    print(
        "Transfer ID:",
        str(transfer_id)
    )

    print(
        "Reference:",
        reference_no
    )

    print(
        "Account:",
        account_name
    )

    print(
        "Saving:",
        saving_name
    )

    print(
        "Amount:",
        f"{currency} {amount:,.2f}"
    )

    print(
        "Account:",
        f"{account_balance_before:,.2f}",
        "→",
        f"{account_balance_after:,.2f}"
    )

    print(
        "Saving:",
        f"{saving_balance_before:,.2f}",
        "→",
        f"{saving_balance_after:,.2f}"
    )

    print(
        "==================================================\n"
    )

    return jsonify(
        response_data
    ), 200




@bp.route("/saving-withdraw", methods=["POST"])
@login_required
def saving_withdraw():

    # ============================================================
    # GET JSON DATA
    # ============================================================

    data = request.get_json(silent=True)

    if not data:

        return jsonify({
            "success": False,
            "error": "No data provided"
        }), 400


    saving_id = data.get("saving_id")
    account_id = data.get("account_id")
    amount = data.get("amount")


    # ============================================================
    # REQUIRED FIELDS
    # ============================================================

    if not saving_id:

        return jsonify({
            "success": False,
            "error": "Saving account is required."
        }), 400


    if not account_id:

        return jsonify({
            "success": False,
            "error": "Destination account is required."
        }), 400


    if amount is None or amount == "":

        return jsonify({
            "success": False,
            "error": "Amount is required."
        }), 400


    # ============================================================
    # AMOUNT VALIDATION
    # ============================================================

    try:

        amount = float(amount)

    except (TypeError, ValueError):

        return jsonify({
            "success": False,
            "error": "Invalid amount."
        }), 400


    if amount <= 0:

        return jsonify({
            "success": False,
            "error": "Amount must be greater than zero."
        }), 400


    # ============================================================
    # PREVENT NaN / INFINITY
    # ============================================================

    import math

    if not math.isfinite(amount):

        return jsonify({
            "success": False,
            "error": "Invalid amount."
        }), 400


    # ============================================================
    # OBJECT ID VALIDATION
    # ============================================================

    try:

        saving_obj_id = ObjectId(
            str(saving_id)
        )

        account_obj_id = ObjectId(
            str(account_id)
        )

    except Exception:

        return jsonify({
            "success": False,
            "error": "Invalid saving or account ID."
        }), 400


    # ============================================================
    # CURRENT USER
    # ============================================================

    user_id = str(
        current_user.id
    )


    # ============================================================
    # USER FILTER
    # Supports STRING + ObjectId
    # ============================================================

    try:

        user_object_id = ObjectId(
            user_id
        )

        user_filter = {

            "$or": [

                {
                    "user_id": user_id
                },

                {
                    "user_id": user_object_id
                }

            ]

        }

    except Exception:

        user_filter = {

            "user_id": user_id

        }


    # ============================================================
    # GET SAVING
    # ============================================================

    saving = mongo.db.savings.find_one({

        "_id": saving_obj_id,

        **user_filter

    })


    if not saving:

        return jsonify({

            "success": False,
            "error": "Saving not found."

        }), 404


    # ============================================================
    # GET DESTINATION ACCOUNT
    # ============================================================

    account = mongo.db.accounts.find_one({

        "_id": account_obj_id,

        **user_filter

    })


    if not account:

        return jsonify({

            "success": False,
            "error": "Destination account not found."

        }), 404


    # ============================================================
    # CURRENT BALANCES
    # ============================================================

    saving_balance = float(

        saving.get(
            "current_balance",
            0
        ) or 0

    )


    account_balance = float(

        account.get(
            "balance",
            0
        ) or 0

    )


    # ============================================================
    # CHECK SAVING BALANCE
    # ============================================================

    if amount > saving_balance:

        return jsonify({

            "success": False,

            "error":
                "Insufficient saving balance. "
                f"Available "
                f"{saving.get('currency', 'USD')} "
                f"{saving_balance:,.2f}"

        }), 400


    # ============================================================
    # NAMES
    # ============================================================

    saving_name = (

        saving.get("name")

        or saving.get("saving_name")

        or saving.get("title")

        or "Saving"

    )


    account_name = (

        account.get("name")

        or account.get("account_name")

        or "Unknown Account"

    )


    # ============================================================
    # CURRENCY
    # ============================================================

    currency = (

        saving.get("currency")

        or account.get("currency")

        or "USD"

    )


    # ============================================================
    # REFERENCE NUMBER
    # ============================================================

    reference_no = (

        "SAV-"

        +

        datetime.utcnow().strftime(
            "%Y%m%d%H%M%S"
        )

        +

        "-"

        +

        secrets.token_hex(3).upper()

    )


    # ============================================================
    # BALANCES BEFORE
    # ============================================================

    saving_balance_before = saving_balance

    account_balance_before = account_balance


    # ============================================================
    # BALANCES AFTER
    # ============================================================

    new_saving_balance = (

        saving_balance
        -
        amount

    )


    new_account_balance = (

        account_balance
        +
        amount

    )


    # ============================================================
    # STEP 1
    # REMOVE MONEY FROM SAVING
    # ============================================================

    saving_result = mongo.db.savings.update_one(

        {

            "_id": saving_obj_id,

            **user_filter,

            "current_balance": {
                "$gte": amount
            }

        },

        {

            "$inc": {

                "current_balance":
                    -amount

            },

            "$set": {

                "updated_at":
                    datetime.utcnow()

            }

        }

    )


    if saving_result.modified_count != 1:

        return jsonify({

            "success": False,

            "error":
                "Failed to remove money from saving."

        }), 500


    # ============================================================
    # STEP 2
    # ADD MONEY TO ACCOUNT
    # ============================================================

    account_result = mongo.db.accounts.update_one(

        {

            "_id": account_obj_id,

            **user_filter

        },

        {

            "$inc": {

                "balance":
                    amount

            },

            "$set": {

                "updated_at":
                    datetime.utcnow()

            }

        }

    )


    # ============================================================
    # ACCOUNT UPDATE FAILED → ROLLBACK SAVING
    # ============================================================

    if account_result.modified_count != 1:

        mongo.db.savings.update_one(

            {

                "_id":
                    saving_obj_id

            },

            {

                "$inc": {

                    "current_balance":
                        amount

                }

            }

        )


        return jsonify({

            "success": False,

            "error":
                "Failed to add money to destination account. "
                "Saving balance has been restored."

        }), 500


    # ============================================================
    # STEP 3
    # SAVE ONLY IN account_transfers
    # ============================================================

    try:

        transfer_result = mongo.db.account_transfers.insert_one({

            # ----------------------------------------------------
            # USER
            # ----------------------------------------------------

            "user_id":
                user_id,


            # ----------------------------------------------------
            # TRANSFER TYPE
            # ----------------------------------------------------

            "transfer_type":
                "saving_to_account",


            # ----------------------------------------------------
            # DIRECTION
            # ----------------------------------------------------

            "direction":
                "withdraw",


            # ----------------------------------------------------
            # SOURCE SAVING
            # ----------------------------------------------------

            "from_saving":
                saving_obj_id,

            "from_saving_name":
                saving_name,


            # ----------------------------------------------------
            # DESTINATION ACCOUNT
            # ----------------------------------------------------

            "to_account":
                account_obj_id,

            "to_account_name":
                account_name,


            # ----------------------------------------------------
            # AMOUNT
            # ----------------------------------------------------

            "amount":
                amount,


            # ----------------------------------------------------
            # CURRENCY
            # ----------------------------------------------------

            "currency":
                currency,


            # ----------------------------------------------------
            # STATUS
            # ----------------------------------------------------

            "status":
                "completed",


            # ----------------------------------------------------
            # REFERENCE
            # ----------------------------------------------------

            "reference":
                reference_no,

            "reference_no":
                reference_no,


            # ----------------------------------------------------
            # DESCRIPTION
            # ----------------------------------------------------

            "description":
                "Saving withdrawal to account",


            # ----------------------------------------------------
            # BALANCES
            # ----------------------------------------------------

            "saving_balance_before":
                saving_balance_before,

            "saving_balance_after":
                new_saving_balance,

            "account_balance_before":
                account_balance_before,

            "account_balance_after":
                new_account_balance,


            # ----------------------------------------------------
            # DATE
            # ----------------------------------------------------

            "created_at":
                datetime.utcnow(),

            "updated_at":
                datetime.utcnow()

        })


    except Exception as transfer_error:

        print(
            "Saving withdrawal history error:",
            transfer_error
        )


        # ========================================================
        # ROLLBACK ACCOUNT
        # ========================================================

        mongo.db.accounts.update_one(

            {
                "_id":
                    account_obj_id
            },

            {

                "$inc": {

                    "balance":
                        -amount

                }

            }

        )


        # ========================================================
        # ROLLBACK SAVING
        # ========================================================

        mongo.db.savings.update_one(

            {
                "_id":
                    saving_obj_id
            },

            {

                "$inc": {

                    "current_balance":
                        amount

                }

            }

        )


        return jsonify({

            "success": False,

            "error":
                "Withdrawal could not be recorded. "
                "The transaction was rolled back."

        }), 500


    # ============================================================
    # SUCCESS
    # ============================================================

    return jsonify({

        "success": True,

        "message":
            "Money moved from Saving → Account successfully.",

        "reference":
            reference_no,

        "amount":
            amount,

        "currency":
            currency,

        "saving":
            saving_name,

        "account":
            account_name,

        "saving_balance":
            new_saving_balance,

        "account_balance":
            new_account_balance

    }), 200



def clean_category(cat):

    return {

        "_id": str(
            cat.get("_id")
        ),

        "name": cat.get(
            "name",
            ""
        ),

        "type": cat.get(
            "type",
            ""
        ),

        "items":[

            str(i)

            for i in (
                cat.get("items") or []
            )

            if i

        ]

    }






def get_month_range(year=None, month=None):

    # ==========================
    # CURRENT DATE DEFAULT
    # ==========================

    now = datetime.utcnow()


    if not year:

        year = now.year


    if not month:

        month = now.month



    # ==========================
    # MONTH START
    # ==========================

    start = datetime(
        int(year),
        int(month),
        1
    )



    # ==========================
    # NEXT MONTH
    # ==========================

    if int(month) == 12:

        end = datetime(
            int(year) + 1,
            1,
            1
        )

    else:

        end = datetime(
            int(year),
            int(month) + 1,
            1
        )


    return start, end



def calculate_savings_balance(records):


    total=0

    deposits=0

    withdrawals=0



    for item in records:


        amount=float(
            item.get(
                "amount",
                0
            )
        )


        if item.get(
            "transaction_type"
        )=="deposit":


            deposits += amount

            total += amount



        elif item.get(
            "transaction_type"
        )=="withdrawal":


            withdrawals += amount

            total -= amount



    return {

        "balance":total,

        "deposit":deposits,

        "withdrawal":withdrawals

    }





def calculate_transaction_summary(transactions):


    result={

        "income":0,

        "expense":0,

        "income_categories":defaultdict(float),

        "expense_categories":defaultdict(float),

        "largest_income":None,

        "largest_expense":None

    }



    for trx in transactions:


        amount=float(
            trx.get(
                "amount",
                0
            )
        )


        trx_type=trx.get(
            "transaction_type"
        )


        category=trx.get(
            "category",
            "Unknown"
        )



        if trx_type=="income":


            result["income"] += amount


            result[
                "income_categories"
            ][category]+=amount



            if (

                result["largest_income"] is None

                or amount >

                result["largest_income"]["amount"]

            ):

                result["largest_income"]={

                    "amount":amount,

                    "data":trx

                }




        elif trx_type=="expense":


            result["expense"] += amount


            result[
                "expense_categories"
            ][category]+=amount



            if (

                result["largest_expense"] is None

                or amount >

                result["largest_expense"]["amount"]

            ):


                result["largest_expense"]={

                    "amount":amount,

                    "data":trx

                }


    return result






def get_biggest_category(data):


    if not data:

        return None,0



    key=max(

        data,

        key=data.get

    )


    return key,data[key]



@bp.route("/transactions")
@login_required
def transaction_list():

    from datetime import datetime
    from bson import ObjectId


    # ==================================================
    # USER
    # ==================================================

    user_object_id = ObjectId(
        current_user.id
    )

    user_string_id = str(
        current_user.id
    )




    # ==================================================
    # MONTH / YEAR FILTER
    # ==================================================

    now = datetime.utcnow()


    # ==========================================
    # FIRST TRANSACTION OF USER
    # ==========================================

    first_transaction = mongo.db.transactions.find_one(

        {
            "user_id": ObjectId(current_user.id),
            "status": True
        },

        sort=[
            ("date", 1)
        ]

    )



    # Default start
    SYSTEM_START_MONTH = 1

    FIRST_YEAR = now.year



    if first_transaction:


        first_date = first_transaction.get("date")


        if first_date:


            SYSTEM_START_MONTH = first_date.month

            FIRST_YEAR = first_date.year






    # ==========================================
    # CURRENT SELECTED FILTER
    # ==========================================

    year = int(

        request.args.get(

            "year",

            now.year

        )

    )



    month = int(

        request.args.get(

            "month",

            now.month

        )

    )



    # ==========================================
    # PREVENT FUTURE DATE
    # ==========================================

    if year > now.year:

        year = now.year



    if year == now.year and month > now.month:

        month = now.month






    # ==================================================
    # MONTH RANGE
    # ==================================================

    month_start, month_end = get_month_range(

        year,

        month

    )






    # ==================================================
    # MONTH LIST
    # ==================================================

    month_names = [

        "January",
        "February",
        "March",
        "April",
        "May",
        "June",
        "July",
        "August",
        "September",
        "October",
        "November",
        "December"

    ]



    months = []



    for i in range(12):


        month_number = (

            SYSTEM_START_MONTH + i - 1

        ) % 12 + 1



        # ===============================
        # REMOVE FUTURE MONTHS
        # ===============================

        if year == now.year and month_number > now.month:

            continue



        # ===============================
        # REMOVE BEFORE FIRST TRANSACTION
        # ===============================

        if year == FIRST_YEAR and month_number < SYSTEM_START_MONTH:

            continue



        months.append({

            "number": month_number,

            "name": month_names[month_number - 1]

        })






    # ==================================================
    # YEAR LIST
    # ==================================================

    years = list(

        range(

            FIRST_YEAR,

            now.year + 1

        )

    )


    # ==================================================
    # FILTERS
    # ==================================================

    transaction_type = request.args.get(
        "type"
    )


    account_id = request.args.get(
        "account_id"
    )


    category = request.args.get(
        "category"
    )


    item = request.args.get(
        "item"
    )



    # ==================================================
    # TRANSACTION QUERY
    # ==================================================

    transaction_query = {

        "user_id": user_object_id,

        "status": True,

        "date": {

            "$gte": month_start,

            "$lt": month_end

        }

    }



    filters = {

        "transaction_type":
        transaction_type,

        "category":
        category,

        "item":
        item

    }



    for key,value in filters.items():

        if value:

            transaction_query[key]=value



    if account_id:

        try:

            transaction_query["account_id"] = ObjectId(
                account_id
            )

        except:

            pass




    # ==================================================
    # TRANSACTIONS
    # ==================================================

    transactions = list(

        mongo.db.transactions.find(
            transaction_query
        )
        .sort(
            "date",
            -1
        )

    )



    # ==================================================
    # ACCOUNTS
    # ==================================================

    accounts = list(

        mongo.db.accounts.find({

            "user_id": user_object_id,

            "status": True

        })

    )



    account_map = {

        str(account["_id"]):

        account.get(
            "name",
            "Unknown"
        )

        for account in accounts

    }



    for transaction in transactions:

        transaction["account_name"] = (

            account_map.get(

                str(
                    transaction.get(
                        "account_id"
                    )
                ),

                "Unknown"

            )

        )





    # ==================================================
    # CATEGORIES
    # ==================================================

    categories = [

        clean_category(category)

        for category in mongo.db.categories.find({

            "$or":[

                {
                    "user_id":user_object_id
                },

                {
                    "user_id":user_string_id
                }

            ],

            "status":True

        })

    ]





    # ==================================================
    # SAVINGS
    # ==================================================

    saving_transactions=list(

        mongo.db.saving_transactions.find({

            "$or":[

                {
                    "user_id":user_object_id
                },

                {
                    "user_id":user_string_id
                }

            ],

            "status":True

        })

    )



    saving_summary = calculate_savings_balance(

        saving_transactions

    )





    # ==================================================
    # ANALYTICS
    # ==================================================

    summary = calculate_transaction_summary(

        transactions

    )

 


    #waa gii hore

    total_income = summary["income"]

    total_expense = summary["expense"]


    net_profit = (

        total_income

        -

        total_expense

    )



    total_balance = sum(

        float(

            account.get(
                "balance",
                0
            )

        )

        for account in accounts

    )



    income_category, income_amount = get_biggest_category(

        summary["income_categories"]

    )



    expense_category, expense_amount = get_biggest_category(

        summary["expense_categories"]

    )





    # ==================================================
    # ALERTS
    # ==================================================

    warnings=[]



    if total_balance < 10:

        warnings.append({

            "type":"danger",

            "title":"Low Balance",

            "message":

            f"Available balance ${total_balance:.2f}"

        })




    if total_expense > total_income:

        warnings.append({

            "type":"warning",

            "title":"Overspending",

            "message":

            "Expense is higher than income"

        })







    # ==================================================
    # MONTH CARD
    # ==================================================

    monthly_card={

        "month":
        month_start.strftime(
            "%B %Y"
        ),

        "income":
        total_income,

        "expense":
        total_expense,

        "profit":
        net_profit,

        "transactions":
        len(transactions),

        "saving":
        saving_summary["balance"],

        "deposit":
        saving_summary["deposit"],

        "withdrawal":
        saving_summary["withdrawal"]

    }


    # ==================================================
    # ITEM ANALYTICS
    # ==================================================

    expense_items = {}
    income_items = {}


    for transaction in transactions:

        item = (
            transaction.get("item")
            or "Unknown"
        ).strip()


        amount = float(
            transaction.get(
                "amount",
                0
            )
        )


        transaction_type = (
            transaction.get(
                "transaction_type",
                ""
            )
            .lower()
            .strip()
        )


        if transaction_type == "expense":

            expense_items[item] = (
                expense_items.get(item, 0)
                + amount
            )


        elif transaction_type == "income":

            income_items[item] = (
                income_items.get(item, 0)
                + amount
            )



    # ==================================================
    # SORT ITEMS (OUTSIDE LOOP)
    # ==================================================

    expense_item_list = sorted(

        expense_items.items(),

        key=lambda x: x[1],

        reverse=True

    )



    income_item_list = sorted(

        income_items.items(),

        key=lambda x: x[1],

        reverse=True

    )




    # ==================================================
    # BIGGEST EXPENSE ITEM
    # ==================================================

    biggest_expense_item = "-"

    biggest_expense_item_amount = 0



    if expense_item_list:


        biggest_expense_item = (
            expense_item_list[0][0]
        )


        biggest_expense_item_amount = (
            expense_item_list[0][1]
        )





    # ==================================================
    # BIGGEST INCOME ITEM
    # ==================================================

    biggest_income_item = "-"

    biggest_income_item_amount = 0



    if income_item_list:


        biggest_income_item = (
            income_item_list[0][0]
        )


        biggest_income_item_amount = (
            income_item_list[0][1]
        )





    # ==================================================
    # ITEM RECOMMENDATIONS
    # ==================================================

    item_recommendations = []



    # ===============================
    # EXPENSE
    # ===============================

    if biggest_expense_item_amount > 0:


        expense_percent = (

            biggest_expense_item_amount

            /

            max(total_expense, 1)

        ) * 100



        if expense_percent >= 50:


            item_recommendations.append({

                "type": "danger",

                "title": "High Spending Item",

                "message":
                    f"{biggest_expense_item} consumed "
                    f"{expense_percent:.1f}% "
                    "of your expenses."

            })



        elif expense_percent >= 30:


            item_recommendations.append({

                "type": "warning",

                "title": "Monitor Spending",

                "message":
                    f"{biggest_expense_item} accounts for "
                    f"{expense_percent:.1f}% "
                    "of total expenses."

            })





    # ===============================
    # INCOME
    # ===============================

    if biggest_income_item_amount > 0:


        income_percent = (

            biggest_income_item_amount

            /

            max(total_income,1)

        ) * 100



        if income_percent >= 50:


            item_recommendations.append({

                "type": "success",

                "title": "Main Income Source",

                "message":
                    f"{biggest_income_item} generates "
                    f"{income_percent:.1f}% "
                    "of your income."

            })


    # ===============================
    # INCOME
    # ===============================

    if biggest_income_item_amount > 0:

        income_percent = (
            biggest_income_item_amount
            / max(total_income, 1)
        ) * 100

        if income_percent >= 60:

            item_recommendations.append({

                "type": "info",

                "title": "Income Concentration",

                "message":
                    f"{biggest_income_item} generates "
                    f"{income_percent:.1f}% "
                    "of your income."

            })


    # wareeg 5



    # ==================================================
    # RENDER
    # ==================================================

    return render_template(

        "backend/pages/components/transactions/all_transactions.html",


        transactions=transactions,


        accounts=accounts,


        categories=categories,



        year=year,

        month=month,

        months=months,
years=years,


        monthly_card=monthly_card,



        total_balance=total_balance,

        current_balance=total_balance,


        total_income=total_income,

        total_expense=total_expense,


        net_profit=net_profit,



        total_savings=saving_summary["balance"],


        total_deposit=saving_summary["deposit"],


        total_withdrawal=saving_summary["withdrawal"],



        total_transactions=len(transactions),



        biggest_income_category=income_category,

        biggest_income_amount=income_amount,



        biggest_expense_category=expense_category,

        biggest_expense_amount=expense_amount,



        largest_income=summary["largest_income"],

        largest_expense=summary["largest_expense"],



        warnings=warnings,



        selected_type=transaction_type,

        selected_account=account_id,

        selected_category=category,

        selected_item=item,

        expense_item_list=expense_item_list,

income_item_list=income_item_list,

biggest_expense_item=biggest_expense_item,

biggest_expense_item_amount=biggest_expense_item_amount,

biggest_income_item=biggest_income_item,

biggest_income_item_amount=biggest_income_item_amount,

item_recommendations=item_recommendations,


        month_name=month_start.strftime(
            "%B %Y"
        )

    )





@bp.route("/persons")
@login_required
def persons():

    persons_dict = defaultdict(lambda: {
        "_id": "",
        "name": "",
        "total_income": 0.0,
        "total_expense": 0.0,
        "balance": 0.0,
        "transactions": 0
    })


    # ==================================================
    # 1. NORMAL TRANSACTIONS
    # ==================================================

    transactions = mongo.db.transactions.find({
        "$or": [
            {
                "user_id": str(current_user.id)
            },
            {
                "user_id": ObjectId(current_user.id)
            }
        ]
    }).sort(
        "date",
        -1
    )


    for trx in transactions:


        raw_name = (
            trx.get("person_name")
            or trx.get("description")
            or trx.get("note")
            or ""
        )


        raw_name = str(raw_name).strip()


        if not raw_name:
            continue



        # Remove By / by
        raw_name = re.sub(
            r"^by\s*",
            "",
            raw_name,
            flags=re.IGNORECASE
        )


        person_name = " ".join(
            raw_name.split()
        ).title()



        if not person_name:
            continue



        amount = float(
            trx.get("amount", 0) or 0
        )


        persons_dict[person_name]["_id"] = person_name

        persons_dict[person_name]["name"] = person_name

        persons_dict[person_name]["transactions"] += 1



        if trx.get("transaction_type") == "income":

            persons_dict[person_name]["total_income"] += amount

        else:

            persons_dict[person_name]["total_expense"] += amount





    # ==================================================
    # 2. OLD EXCEL / MANUAL TRANSACTIONS
    # ==================================================

    old_transactions = mongo.db.person_opening_transactions.find({

        "user_id": ObjectId(current_user.id)

    }).sort(
        "date",
        -1
    )


    for old in old_transactions:


        raw_name = str(
            old.get("person_name", "")
        ).strip()


        if not raw_name:
            continue



        raw_name = re.sub(
            r"^by\s*",
            "",
            raw_name,
            flags=re.IGNORECASE
        )



        person_name = " ".join(
            raw_name.split()
        ).title()



        amount = float(
            old.get("amount", 0) or 0
        )



        persons_dict[person_name]["_id"] = person_name

        persons_dict[person_name]["name"] = person_name

        persons_dict[person_name]["transactions"] += 1



        if old.get("type") == "income":

            persons_dict[person_name]["total_income"] += amount

        else:

            persons_dict[person_name]["total_expense"] += amount





    # ==================================================
    # FINAL DATA
    # ==================================================

    persons = []


    for person in persons_dict.values():

        person["balance"] = (
            person["total_income"]
            -
            person["total_expense"]
        )

        persons.append(person)



    persons.sort(
        key=lambda x: x["name"].lower()
    )


    return render_template(
        "backend/pages/components/transactions/persons.html",
        persons=persons
    )


@bp.route("/person/<person_name>/invoice")
@login_required
def person_invoice(person_name):

    person_name = person_name.strip()


    user_query = {
        "$or": [
            {
                "user_id": str(current_user.id)
            },
            {
                "user_id": ObjectId(current_user.id)
            }
        ]
    }


    total_income = 0
    total_expense = 0

    invoices = []



    # ================================
    # NORMAL TRANSACTIONS
    # ================================

    normal_transactions = list(
        mongo.db.transactions.find({

            **user_query,

            "$or": [

                {
                    "description": {
                        "$regex": person_name,
                        "$options": "i"
                    }
                },

                {
                    "person_name": {
                        "$regex": person_name,
                        "$options": "i"
                    }
                },

                {
                    "note": {
                        "$regex": person_name,
                        "$options": "i"
                    }
                }

            ]

        })
    )





    # ================================
    # OPENING / MANUAL TRANSACTIONS
    # ================================

    opening_transactions = list(
        mongo.db.person_opening_transactions.find({

            "user_id": ObjectId(current_user.id),

            "person_name": {
                "$regex": person_name,
                "$options": "i"
            }

        })
    )






    # ================================
    # NORMAL
    # ================================

    for t in normal_transactions:


        amount = float(
            t.get("amount", 0)
        )


        t_type = t.get(
            "transaction_type",
            "expense"
        )


        if t_type == "income":

            total_income += amount

        else:

            total_expense += amount



        invoices.append({

            "date": t.get("date"),

            "item": t.get("item") or "Transaction",

            "reference": "NORMAL",

            "type": t_type,

            "amount": amount,

            "note": t.get("note")

        })







    # ================================
    # OPENING / MANUAL
    # ================================

    for t in opening_transactions:


        amount = float(
            t.get("amount", 0)
        )


        t_type = t.get(
            "type",
            "expense"
        )


        if t_type == "income":

            total_income += amount

        else:

            total_expense += amount




        invoices.append({

            "date": t.get("date"),

            "item": t.get("item") or "Opening Balance",

            "reference": "MANUAL",

            "type": t_type,

            "amount": amount,

            "note": t.get("note")

        })







    # ================================
    # LATEST DATE FIRST
    # ================================

    invoices = sorted(

        invoices,

        key=lambda x: x.get("date") or datetime.min,

        reverse=True

    )






    balance = (
        total_income -
        total_expense
    )





    return render_template(

        "backend/pages/components/transactions/person_invoice.html",

        person_name=person_name,

        invoices=invoices,

        total_income=total_income,

        total_expense=total_expense,

        balance=balance,

        now=now_eat()

    )



@bp.route("/person/manual/add", methods=["GET", "POST"])
@login_required
def add_manual_person_transaction():

    if request.method == "POST":

        # =========================
        # FORM DATA
        # =========================

        person_name = request.form.get(
            "person_name",
            ""
        ).strip()

        transaction_type = request.form.get(
            "transaction_type",
            "expense"
        )

        category = request.form.get(
            "category",
            "Person Payment"
        ).strip()

        item = request.form.get(
            "item",
            ""
        ).strip()

        note = request.form.get(
            "note",
            ""
        ).strip()

        try:
            amount = float(
                request.form.get(
                    "amount",
                    0
                )
            )
        except (ValueError, TypeError):
            amount = 0

        # =========================
        # VALIDATION
        # =========================

        if not person_name or amount <= 0:

            flash(
                "Person and amount required.",
                "danger"
            )

            return redirect(request.url)

        # =========================
        # USER SELECTED DATE & TIME
        # =========================

        date_value = request.form.get(
            "date",
            ""
        ).strip()

        if date_value:

            try:

                transaction_date = datetime.strptime(
                    date_value,
                    "%Y-%m-%dT%H:%M"
                )

            except ValueError:

                flash(
                    "Invalid date and time.",
                    "danger"
                )

                return redirect(request.url)

        else:

            transaction_date = now_eat().replace(
                tzinfo=None
            )

        # =========================
        # CLEAN PERSON NAME
        # =========================

        person_name = (
            person_name
            .replace("By ", "")
            .replace("by ", "")
            .strip()
            .title()
        )

        # =========================
        # INSERT
        # =========================

        mongo.db.person_opening_transactions.insert_one({

            "user_id": ObjectId(current_user.id),

            "person_name": person_name,

            "type": transaction_type,

            "category": category,

            "amount": amount,

            "item": item,

            "note": note,

            "source": "manual",

            # User selected Date & Time
            "date": transaction_date,

            # Record created time
            "created_at": now_eat()

        })

        flash(
            "Person transaction added successfully.",
            "success"
        )

        return redirect(
            url_for("main.persons")
        )

    return render_template(

        "backend/pages/components/transactions/add_person_manual.html",

        today=now_eat().replace(
            tzinfo=None
        ).strftime("%Y-%m-%dT%H:%M")

    )


@bp.route("/person-opening-transactions")
@login_required
def person_opening_transactions():

    try:

        user_id = ObjectId(
            current_user.id
        )

    except Exception:

        abort(404)



    # ======================================
    # GET TRANSACTIONS
    # LATEST FIRST
    # ======================================

    transactions = list(

        mongo.db.person_opening_transactions.find({

            "user_id": user_id

        }).sort(

            [
                ("date", -1),          # Latest date first
                ("created_at", -1)     # Latest created first
            ]

        )

    )



    return render_template(

        "backend/pages/components/transactions/person_opening_transactions.html",

        transactions=transactions,

        now=now_eat()

    )



@bp.route("/person-opening-transactions/edit/<id>", methods=["GET", "POST"])
@login_required
def edit_person_opening_transaction(id):

    try:
        transaction_id = ObjectId(id)
        user_id = ObjectId(current_user.id)

    except Exception:
        abort(404)

    # =========================
    # GET TRANSACTION
    # =========================

    transaction = mongo.db.person_opening_transactions.find_one({

        "_id": transaction_id,
        "user_id": user_id

    })

    if not transaction:
        abort(404)

    # =========================
    # UPDATE
    # =========================

    if request.method == "POST":

        person_name = request.form.get(
            "person_name",
            ""
        ).strip().title()

        trx_type = request.form.get(
            "type",
            "expense"
        )

        category = request.form.get(
            "category",
            "Person Payment"
        ).strip()

        item = request.form.get(
            "item",
            ""
        ).strip()

        note = request.form.get(
            "note",
            ""
        ).strip()

        try:
            amount = float(
                request.form.get(
                    "amount",
                    0
                )
            )
        except (ValueError, TypeError):

            flash(
                "Invalid amount.",
                "danger"
            )

            return redirect(request.url)

        if not person_name or amount <= 0:

            flash(
                "Person and amount required.",
                "danger"
            )

            return redirect(request.url)

        # =========================
        # DATE & TIME
        # =========================

        date_value = request.form.get(
            "date",
            ""
        ).strip()

        if date_value:

            try:

                transaction_datetime = datetime.strptime(
                    date_value,
                    "%Y-%m-%dT%H:%M"
                )

            except ValueError:

                flash(
                    "Invalid date and time.",
                    "danger"
                )

                return redirect(request.url)

        else:

            transaction_datetime = transaction.get(
                "date",
                now_eat().replace(tzinfo=None)
            )

        # =========================
        # UPDATE DATABASE
        # =========================

        mongo.db.person_opening_transactions.update_one(

            {
                "_id": transaction_id,
                "user_id": user_id
            },

            {
                "$set": {

                    "person_name": person_name,

                    "type": trx_type,

                    "category": category,

                    "amount": amount,

                    "item": item,

                    "note": note,

                    # User selected date & time
                    "date": transaction_datetime,

                    # Update timestamp
                    "updated_at": now_eat()

                }
            }

        )

        flash(
            "Person opening transaction updated successfully.",
            "success"
        )

        return redirect(
            url_for("main.person_opening_transactions")
        )

    # =========================
    # FORMAT DATE FOR datetime-local
    # =========================

    if transaction.get("date"):

        if hasattr(transaction["date"], "tzinfo") and transaction["date"].tzinfo:

            transaction["date"] = transaction["date"].replace(
                tzinfo=None
            )

    return render_template(

        "backend/pages/components/transactions/edit_person_opening_transaction.html",

        transaction=transaction

    )




@bp.route("/person-opening-transactions/export-sample")
@login_required
def export_sample_person_opening_transactions():

    wb = Workbook()

    ws = wb.active
    ws.title = "Person Opening Transactions"


    # =========================
    # HEADERS
    # =========================

    ws.append([
        "Person Name",
        "Type",
        "Category",
        "Amount",
        "Item",
        "Note",
        "Date"
    ])


    # =========================
    # SAMPLE DATA
    # =========================

    ws.append([
        "Ahmed Ali",
        "Income",
        "Opening Balance",
        1500,
        "Cash",
        "Initial opening balance",
        "2026-07-26T09:00"
    ])


    ws.append([
        "Mohamed Hassan",
        "Expense",
        "Payment",
        250,
        "Transport",
        "Bus fare payment",
        "2026-07-26T10:30"
    ])


    ws.append([
        "Asha Mohamed",
        "Income",
        "Debt Received",
        800,
        "Cash",
        "Received previous debt",
        "2026-07-27T14:15"
    ])


    # =========================
    # COLUMN WIDTH
    # =========================

    widths = {
        "A": 25,
        "B": 15,
        "C": 20,
        "D": 15,
        "E": 20,
        "F": 30,
        "G": 25
    }


    for col, width in widths.items():

        ws.column_dimensions[col].width = width



    # =========================
    # SAVE FILE
    # =========================

    output = io.BytesIO()

    wb.save(output)

    output.seek(0)



    return send_file(

        output,

        as_attachment=True,

        download_name="Person_Opening_Transactions_Sample.xlsx",

        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )


@bp.route("/person-opening-transactions/export")
@login_required
def export_person_opening_transactions():

    user_id = ObjectId(current_user.id)


    # =========================
    # GET USER TRANSACTIONS
    # =========================

    transactions = list(
        mongo.db.person_opening_transactions.find(
            {
                "user_id": user_id
            }
        ).sort(
            "date",
            -1
        )
    )


    wb = Workbook()

    ws = wb.active
    ws.title = "Person Opening Transactions"



    # =========================
    # HEADERS
    # =========================

    ws.append([
        "Person Name",
        "Type",
        "Category",
        "Amount",
        "Item",
        "Note",
        "Date"
    ])



    # =========================
    # DATA ROWS
    # =========================

    for trx in transactions:


        date_value = trx.get("date")


        if date_value:

            if hasattr(date_value, "strftime"):

                date_value = date_value.strftime(
                    "%Y-%m-%d %H:%M"
                )

            else:

                date_value = str(date_value)

        else:

            date_value = ""



        ws.append([

            trx.get(
                "person_name",
                ""
            ),

            trx.get(
                "type",
                ""
            ),

            trx.get(
                "category",
                ""
            ),

            trx.get(
                "amount",
                0
            ),

            trx.get(
                "item",
                ""
            ),

            trx.get(
                "note",
                ""
            ),

            date_value

        ])



    # =========================
    # AUTO COLUMN WIDTH
    # =========================

    for column in ws.columns:

        max_length = 0

        column_letter = column[0].column_letter


        for cell in column:

            try:

                if len(str(cell.value)) > max_length:

                    max_length = len(
                        str(cell.value)
                    )

            except:

                pass


        ws.column_dimensions[
            column_letter
        ].width = max_length + 5



    # =========================
    # DOWNLOAD
    # =========================

    output = io.BytesIO()

    wb.save(output)

    output.seek(0)



    return send_file(

        output,

        as_attachment=True,

        download_name="Person_Opening_Transactions.xlsx",

        mimetype=
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    )



@bp.route("/person-opening-transactions/import", methods=["POST"])
@login_required
def import_person_opening_transactions():

    user_id = ObjectId(current_user.id)

    file = request.files.get("file")


    if not file:

        flash(
            "Please select Excel file.",
            "danger"
        )

        return redirect(request.referrer)



    try:

        wb = load_workbook(file)

        ws = wb.active


    except Exception as e:

        flash(
            "Invalid Excel file.",
            "danger"
        )

        return redirect(request.referrer)



    imported_count = 0
    skipped_count = 0
    duplicate_count = 0
    error_count = 0



    # SAME TIME FOR THIS IMPORT BATCH

    import_time = now_eat().replace(
        tzinfo=None
    )



    for row in ws.iter_rows(
        min_row=2,
        values_only=True
    ):


        try:

            (
                person_name,
                trx_type,
                category,
                amount,
                item,
                note,
                date_value

            ) = row



            # =========================
            # REQUIRED
            # =========================

            if not person_name or amount is None:

                skipped_count += 1

                continue




            # =========================
            # PERSON NAME
            # =========================

            person_name = (
                str(person_name)
                .strip()
                .title()
            )



            # =========================
            # TYPE FIX
            # =========================

            raw_type = str(
                trx_type or ""
            ).strip().lower()



            if raw_type.startswith("income"):

                trx_type = "income"


            elif raw_type.startswith("expense"):

                trx_type = "ixpense"


            else:

                print(
                    "INVALID TYPE:",
                    repr(trx_type)
                )

                skipped_count += 1

                continue




            # =========================
            # AMOUNT CLEAN
            # =========================

            try:

                amount_text = str(
                    amount
                )


                amount_text = (

                    amount_text
                    .replace("$", "")
                    .replace(",", "")
                    .strip()

                )


                amount = float(
                    amount_text
                )


            except:

                skipped_count += 1

                continue




            if amount == 0:

                skipped_count += 1

                continue




            # =========================
            # AMOUNT BY TYPE
            # =========================
            amount = abs(amount)

            if trx_type == "expense":
                amount = -abs(amount)

            elif trx_type == "income":
                amount = abs(amount)




            # =========================
            # CATEGORY
            # =========================

            category = (

                str(category)
                .strip()

                if category

                else "Person Payment"

            )



            item = (

                str(item)
                .strip()

                if item

                else ""

            )



            note = (

                str(note)
                .strip()

                if note

                else ""

            )




            # =========================
            # DATE
            # =========================

            transaction_date = now_eat().replace(
                tzinfo=None
            )



            if date_value:


                if isinstance(
                    date_value,
                    datetime
                ):

                    transaction_date = date_value



                else:


                    date_string = str(
                        date_value
                    ).strip()



                    formats = [

                        "%Y-%m-%dT%H:%M",

                        "%Y-%m-%d %H:%M",

                        "%Y-%m-%d",

                        "%d/%m/%Y",

                        "%m/%d/%Y"

                    ]



                    for fmt in formats:

                        try:

                            transaction_date = datetime.strptime(
                                date_string,
                                fmt
                            )

                            break


                        except ValueError:

                            pass




            # =========================
            # DUPLICATE CHECK
            # =========================

            exists = mongo.db.person_opening_transactions.find_one({

                "user_id": user_id,

                "person_name": person_name,

                "type": trx_type,

                "category": category,

                "amount": amount,

                "date": transaction_date

            })



            if exists:

                duplicate_count += 1

                continue





            # =========================
            # INSERT
            # =========================

            mongo.db.person_opening_transactions.insert_one({

                "user_id": user_id,

                "person_name": person_name,

                "type": trx_type,

                "category": category,

                "amount": amount,

                "item": item,

                "note": note,

                "date": transaction_date,


                # FOR UNDO LAST IMPORT

                "source": "import",

                "import_batch": import_time,


                "created_at": import_time,

                "updated_at": import_time

            })



            imported_count += 1




        except Exception as e:

            print(
                "IMPORT ERROR:",
                e
            )

            error_count += 1

            continue





    flash(

        f"""
        Import Completed ✅

        Imported: {imported_count}

        Skipped: {skipped_count}

        Duplicate: {duplicate_count}

        Errors: {error_count}
        """,

        "success"

    )



    return redirect(

        url_for(
            "main.person_opening_transactions"
        )

    )


@bp.route("/person-opening-transactions/undo-last-import")
@login_required
def undo_last_person_opening_import():

    try:
        user_id = ObjectId(current_user.id)

    except Exception:
        abort(404)


    # =========================
    # FIND LAST IMPORT RECORD
    # =========================

    last_record = mongo.db.person_opening_transactions.find_one(
        {
            "user_id": user_id
        },
        sort=[
            (
                "created_at",
                -1
            )
        ]
    )


    if not last_record:

        flash(
            "No last import found to undo.",
            "warning"
        )

        return redirect(
            url_for(
                "main.person_opening_transactions"
            )
        )


    created_at = last_record.get("created_at")


    if not created_at:

        flash(
            "Created time missing.",
            "danger"
        )

        return redirect(
            url_for(
                "main.person_opening_transactions"
            )
        )


    # =========================
    # FIND SAME IMPORT TIME RANGE
    # =========================

    start_time = created_at - timedelta(seconds=10)

    end_time = created_at + timedelta(seconds=10)



    # =========================
    # DELETE LAST IMPORT BATCH
    # =========================

    result = mongo.db.person_opening_transactions.delete_many(
        {
            "user_id": user_id,

            "created_at": {
                "$gte": start_time,
                "$lte": end_time
            }
        }
    )


    flash(
        f"Last import undone successfully. Removed {result.deleted_count} records.",
        "success"
    )


    return redirect(
        url_for(
            "main.person_opening_transactions"
        )
    )


@bp.route("/person-opening-transactions/delete/<id>")
@login_required
def delete_person_opening_transaction(id):

    try:
        transaction_id = ObjectId(id)
        user_id = ObjectId(current_user.id)

    except Exception:
        abort(404)


    # =========================
    # CHECK EXISTS
    # =========================

    transaction = mongo.db.person_opening_transactions.find_one({

        "_id": transaction_id,

        "user_id": user_id

    })


    if not transaction:

        flash(
            "Opening transaction not found",
            "danger"
        )

        return redirect(
            url_for(
                "main.person_opening_transactions"
            )
        )



    # =========================
    # DELETE
    # =========================

    mongo.db.person_opening_transactions.delete_one({

        "_id": transaction_id,

        "user_id": user_id

    })



    flash(
        "Person opening transaction deleted successfully",
        "success"
    )


    return redirect(
        url_for(
            "main.person_opening_transactions"
        )
    )



@bp.route("/person/<person_id>/ledger")
@login_required
def person_ledger(person_id):

    # =========================
    # CLEAN PERSON NAME
    # =========================

    person_name = str(person_id).strip()

    person_name = re.sub(
        r"^by\s+",
        "",
        person_name,
        flags=re.IGNORECASE
    )

    person_name = person_name.strip().title()



    person_regex = rf"^\s*(?:by\s+)?{re.escape(person_name)}\s*$"




    user_filter = {

        "$or": [

            {
                "user_id": str(current_user.id)
            },

            {
                "user_id": ObjectId(current_user.id)
            }

        ]

    }




    # =========================
    # TRANSACTIONS
    # =========================

    transactions = list(

        mongo.db.transactions.find({

            **user_filter,

            "$or": [

                {
                    "description": {
                        "$regex": person_regex,
                        "$options": "i"
                    }
                },

                {
                    "person_name": {
                        "$regex": person_regex,
                        "$options": "i"
                    }
                },

                {
                    "note": {
                        "$regex": person_regex,
                        "$options": "i"
                    }
                }

            ]

        })

    )






    # =========================
    # OPENING / MANUAL
    # =========================

    old_transactions = list(

        mongo.db.person_opening_transactions.find({

            "user_id": ObjectId(current_user.id),

            "person_name": {
                "$regex": person_regex,
                "$options": "i"
            }

        })

    )





    ledger = []


    total_income = 0

    total_expense = 0





    # =========================
    # NORMAL TRANSACTIONS
    # =========================

    for t in transactions:


        amount = float(
            t.get("amount", 0) or 0
        )


        trx_type = t.get(
            "transaction_type",
            "expense"
        )



        if trx_type == "income":

            total_income += amount

        else:

            total_expense += amount





        ledger.append({

            "date": t.get("date"),

            "type": trx_type,

            "category": t.get("category"),

            "item": t.get("item"),

            "amount": amount,

            "note": t.get("description") or t.get("note"),

            "source": "Transaction"

        })








    # =========================
    # OPENING / MANUAL
    # =========================

    for t in old_transactions:


        amount = float(
            t.get("amount", 0) or 0
        )



        trx_type = t.get(
            "type",
            "expense"
        )



        if trx_type == "income":

            total_income += amount

        else:

            total_expense += amount





        ledger.append({

            "date": t.get("date"),

            "type": trx_type,

            "category": t.get(
                "category",
                "Opening Balance"
            ),

            "item": t.get("item"),


            "amount": amount,


            "note": t.get("note"),


            "source": "Excel / Manual"

        })







    # =========================
    # DATE CONVERTER
    # =========================

    def convert_date(value):

        if isinstance(value, datetime):

            return value



        if isinstance(value, str):

            formats = [

                "%Y-%m-%dT%H:%M",

                "%Y-%m-%d",

                "%Y-%m-%d %H:%M:%S"

            ]


            for fmt in formats:

                try:

                    return datetime.strptime(
                        value,
                        fmt
                    )

                except:

                    continue



        return datetime.min






    # =========================
    # SORT LATEST FIRST
    # =========================

    ledger.sort(

        key=lambda x: convert_date(
            x.get("date")
        ),

        reverse=True

    )







    balance = (

        total_income -

        total_expense

    )






    return render_template(

        "backend/pages/components/transactions/person_ledger.html",

        person_name=person_name,

        ledger=ledger,

        total_income=total_income,

        total_expense=total_expense,

        balance=balance,

        now=now_eat()

    )



@bp.route("/add-transaction", methods=["GET", "POST"])
@login_required
def add_transaction():
    user_oid = ObjectId(current_user.id)

    if request.method == "POST":
        account_id = request.form.get("account_id")
        transaction_type = request.form.get("transaction_type")
        category_id = request.form.get("category")
        item = request.form.get("item")
        amount = request.form.get("amount")
        description = request.form.get("description", "")
        note = request.form.get("note", "")
        reference_no = request.form.get("reference_no")

        if not account_id or not category_id or not amount:
            flash("Account, Category and Amount are required.", "danger")
            return redirect(url_for("main.add_transaction"))
        
        # Gudaha POST
        cat_doc = mongo.db.categories.find_one({"_id": ObjectId(category_id)})
        if cat_doc.get("type") != transaction_type:
            flash("Category-gan kuma haboona Type-ka aad dooratay!", "danger")
            return redirect(url_for("main.add_transaction"))

        try:
            amount = float(amount)
        except ValueError:
            flash("Invalid amount", "danger")
            return redirect(url_for("main.add_transaction"))

        # 🔥 CATEGORY VALIDATION
        category_doc = mongo.db.categories.find_one({
            "_id": ObjectId(category_id),
            "user_id": user_oid
        })

        if not category_doc:
            flash("Invalid category", "danger")
            return redirect(url_for("main.add_transaction"))

        # 🔥 ITEM VALIDATION
        # 🔥 ITEM VALIDATION
        if item:
            # 1. Soo saar list-ka saxda ah
            raw_items = category_doc.get("items", [])
            
            # Haddii ay tahay JSON string, u beddel list
            if isinstance(raw_items, str):
                try:
                    items = json.loads(raw_items)
                except:
                    items = []
            else:
                items = raw_items

            # 2. Nadiifi dhammaan items-ka si ay u noqdaan strings nadiif ah
            # Tani waxay ka saaraysaa quotes-ka iyo brackets-ka aan loo baahnayn
            clean_items = [str(i).strip('[]"\' ') for i in items]

            # 3. Hadda isbarbar dhig
            if item.strip() not in clean_items:
                flash(f"Invalid item: '{item}' maaha mid ku jira category-ga.", "danger")
                return redirect(url_for("main.add_transaction"))
        # 🔥 BALANCE VALIDATION (Halkan ayaan ku daray)
        if transaction_type == "expense":
            account_doc = mongo.db.accounts.find_one({
                "_id": ObjectId(account_id),
                "user_id": user_oid
            })
            if not account_doc or account_doc.get("balance", 0) < amount:
                flash(f"Digniin: Akaunkan balance-kiisu waa {account_doc.get('balance', 0)}. Ma samayn kartid expense ka badan lacagtaas!", "danger")
                return redirect(url_for("main.add_transaction"))

        # 🔥 SAVE TRANSACTION WITH SESSION (Atomicity)
        try:
            with mongo.db.client.start_session() as session:
                with session.start_transaction():
                    # 1. Diyaarinta Data
                    data = Transaction().add(
                        user_id=current_user.id,
                        account_id=account_id,
                        transaction_type=transaction_type,
                        category=category_doc["name"],
                        item=item,
                        amount=amount,
                        description=description,
                        note=note,
                        reference_no=reference_no
                    )
                    data["user_id"] = user_oid
                    data["account_id"] = ObjectId(account_id)

                    # 2. Insert Transaction
                    mongo.db.transactions.insert_one(data, session=session)

                    # 3. Update Balance
                    inc_val = amount if transaction_type == "income" else -amount
                    mongo.db.accounts.update_one(
                        {"_id": ObjectId(account_id)},
                        {"$inc": {"balance": inc_val}},
                        session=session
                    )
            
            flash("Transaction saved successfully", "success")
            return redirect(url_for("main.transaction_list"))
        
        except Exception as e:
            flash(f"Error saving transaction: {str(e)}", "danger")
            return redirect(url_for("main.add_transaction"))

    # GET Request: Fetch and Clean Categories
    # GET Request: Fetch and Clean Categories
    accounts = list(mongo.db.accounts.find({"user_id": user_oid}))
    categories_raw = list(mongo.db.categories.find({"user_id": user_oid}))

    categories = []
    for cat in categories_raw:
        items = cat.get("items", [])
        
        # Haddii ay tahay JSON string, u beddel list
        if isinstance(items, str):
            try:
                items = json.loads(items)
            except:
                items = []
                
        # Hubi in items ay yihiin list dhab ah (haddii ay ku jiraan brackets dhexda)
        clean_items = []
        if isinstance(items, list):
            for i in items:
                # Ka saar brackets iyo quotes dheeraad ah
                clean_item = str(i).replace('[', '').replace(']', '').replace("'", "").replace('"', "").strip()
                if clean_item:
                    clean_items.append(clean_item)
                    
        cat["items"] = clean_items
        categories.append(clean_category(cat))

        

    return render_template(
        "backend/pages/components/transactions/add_transaction.html",
        accounts=accounts,
        categories=categories
    )


@bp.route("/edit-transaction/<id>", methods=["GET", "POST"])
@login_required
def edit_transaction(id):
    user_oid = ObjectId(current_user.id)

    # =========================
    # GET TRANSACTION
    # =========================
    transaction = mongo.db.transactions.find_one(
        {
            "_id": ObjectId(id),
            "user_id": user_oid
        }
    )

    if not transaction:
        flash("Transaction not found", "danger")
        return redirect(url_for("main.transaction_list"))

    # =========================
    # POST
    # =========================
    if request.method == "POST":

        account_id = request.form.get("account_id")
        transaction_type = request.form.get("transaction_type")
        category_id = request.form.get("category")
        item = request.form.get("item")

        amount = float(request.form.get("amount", 0))

        description = request.form.get(
            "description",
            ""
        )

        note = request.form.get(
            "note",
            ""
        )

        reference_no = request.form.get(
            "reference_no",
            ""
        )


        # Date
        date_str = request.form.get("date")

        if date_str:
            try:
                transaction_date = datetime.strptime(
                    date_str,
                    "%Y-%m-%dT%H:%M"
                )
            except:
                transaction_date = datetime.utcnow()

        else:
            transaction_date = datetime.utcnow()


        # Status
        status = True if request.form.get("status") else False

        # =========================
        # BALANCE CHECK FIX
        # =========================

        if transaction_type == "expense":

            account = mongo.db.accounts.find_one(
                {
                    "_id": ObjectId(account_id)
                }
            )

            if account:


                current_balance = account.get(
                    "balance",
                    0
                )


                available_balance = current_balance



                # Haddii account kii hore yahay isla account-ka
                if (
                    transaction["transaction_type"] == "expense"
                    and str(transaction["account_id"]) == account_id
                ):

                    available_balance += transaction["amount"]



                if amount > available_balance:

                    flash(
                        "Balance-ka akaunkan kuma filna!",
                        "danger"
                    )

                    return redirect(
                        url_for(
                            "main.edit_transaction",
                            id=id
                        )
                    )



        # =========================
        # CATEGORY CHECK
        # =========================

        category_doc = mongo.db.categories.find_one(
            {
                "_id": ObjectId(category_id),
                "user_id": user_oid
            }
        )


        if not category_doc:

            flash(
                "Invalid category",
                "danger"
            )

            return redirect(
                url_for(
                    "main.edit_transaction",
                    id=id
                )
            )



        try:

            with mongo.db.client.start_session() as session:

                with session.start_transaction():


                    old_amount = transaction["amount"]

                    old_type = transaction["transaction_type"]



                    # =========================
                    # RESTORE OLD BALANCE
                    # =========================

                    if old_type == "income":

                        revert = -old_amount

                    else:

                        revert = old_amount



                    mongo.db.accounts.update_one(
                        {
                            "_id": transaction["account_id"]
                        },
                        {
                            "$inc": {
                                "balance": revert
                            }
                        },
                        session=session
                    )



                    # =========================
                    # UPDATE TRANSACTION
                    # =========================

                    mongo.db.transactions.update_one(
                        {
                            "_id": ObjectId(id)
                        },
                        {
                            "$set": {

                                "account_id": ObjectId(account_id),

                                "transaction_type": transaction_type,

                                "category": category_doc["name"],

                                "item": item,

                                "amount": amount,


                                "description": description,

                                "note": note,

                                "reference_no": reference_no,

                                "date": transaction_date,

                                "status": status,


                                "updated_at": datetime.utcnow()
                            }
                        },
                        session=session
                    )



                    # =========================
                    # APPLY NEW BALANCE
                    # =========================

                    if transaction_type == "income":

                        new_value = amount

                    else:

                        new_value = -amount



                    mongo.db.accounts.update_one(
                        {
                            "_id": ObjectId(account_id)
                        },
                        {
                            "$inc": {
                                "balance": new_value
                            }
                        },
                        session=session
                    )



            flash(
                "Transaction updated successfully",
                "success"
            )


            return redirect(
                url_for(
                    "main.transaction_list"
                )
            )



        except Exception as e:

            flash(
                f"Error updating: {e}",
                "danger"
            )


            return redirect(
                url_for(
                    "main.edit_transaction",
                    id=id
                )
            )

    # ===================================================
    # GET DATA
    # ===================================================

    # Accounts
    accounts = list(
        mongo.db.accounts.find(
            {
                "user_id": user_oid
            }
        )
    )

    for acc in accounts:
        acc["_id"] = str(acc["_id"])

        if "user_id" in acc:
            acc["user_id"] = str(acc["user_id"])

    # Categories
    categories_raw = list(
        mongo.db.categories.find(
            {
                "user_id": user_oid
            }
        )
    )

    categories = []

    for cat in categories_raw:

        cat["_id"] = str(cat["_id"])

        if "user_id" in cat:
            cat["user_id"] = str(cat["user_id"])

        items = cat.get("items", [])

        if isinstance(items, str):
            try:
                items = json.loads(items)
            except:
                items = []

        if isinstance(items, list):
            cat["items"] = [
                str(i).strip('[]"\' ')
                for i in items
            ]
        else:
            cat["items"] = []

        categories.append(
            clean_category(cat)
        )

    # Transaction
    transaction["_id"] = str(transaction["_id"])

    if "account_id" in transaction:
        transaction["account_id"] = str(transaction["account_id"])

    if "user_id" in transaction:
        transaction["user_id"] = str(transaction["user_id"])

    return render_template(
        "backend/pages/components/transactions/edit_transaction.html",
        transaction=transaction,
        accounts=accounts,
        categories=categories
    )



@bp.route("/delete-transaction/<id>")
@login_required
def delete_transaction(id):

    trx = mongo.db.transactions.find_one({
        "_id": ObjectId(id),
        "user_id": ObjectId(current_user.id)
    })

    if not trx:
        flash("Transaction not found.", "danger")
        return redirect(url_for("main.transaction_list"))

    amount = float(trx["amount"])

    # 🔥 REVERSE BALANCE
    if trx["transaction_type"] == "income":
        mongo.db.accounts.update_one(
            {"_id": trx["account_id"]},
            {"$inc": {"balance": -amount}}
        )
    else:
        mongo.db.accounts.update_one(
            {"_id": trx["account_id"]},
            {"$inc": {"balance": amount}}
        )

    # DELETE
    mongo.db.transactions.delete_one({
        "_id": ObjectId(id)
    })

    flash("Transaction deleted successfully.", "success")
    return redirect(url_for("main.transaction_list"))


@bp.route("/add-saving", methods=["GET", "POST"])
@login_required
def add_saving():

    # ==========================================
    # USER ID
    # ==========================================
    user_id = ObjectId(current_user.id)

    # ==========================================
    # USER ACCOUNTS
    # ==========================================
    accounts = list(
        mongo.db.accounts.find(
            {
                "user_id": user_id
            }
        )
    )

    if request.method == "POST":

        # ==========================================
        # FORM DATA
        # ==========================================
        title = request.form.get(
            "title",
            ""
        ).strip()

        description = request.form.get(
            "description",
            ""
        ).strip()

        target_amount = request.form.get(
            "target_amount",
            ""
        ).strip()

        account_id = request.form.get(
            "account_id",
            ""
        ).strip()

        start_date = request.form.get(
            "start_date",
            ""
        ).strip()

        maturity_date = request.form.get(
            "maturity_date",
            ""
        ).strip()

        status = request.form.get(
            "status",
            "active"
        ).strip().lower()

        # ==========================================
        # VALIDATION
        # ==========================================
        if not title:

            flash(
                "Saving title is required.",
                "danger"
            )

            return redirect(
                url_for("main.add_saving")
            )

        if not target_amount:

            flash(
                "Target amount is required.",
                "danger"
            )

            return redirect(
                url_for("main.add_saving")
            )

        if not account_id:

            flash(
                "Please select account.",
                "danger"
            )

            return redirect(
                url_for("main.add_saving")
            )

        # ==========================================
        # TARGET AMOUNT
        # ==========================================
        try:

            target_amount = float(
                Decimal(target_amount).quantize(
                    Decimal("0.01")
                )
            )

        except InvalidOperation:

            flash(
                "Invalid target amount.",
                "danger"
            )

            return redirect(
                url_for("main.add_saving")
            )

        if target_amount <= 0:

            flash(
                "Target amount must be greater than zero.",
                "danger"
            )

            return redirect(
                url_for("main.add_saving")
            )

        # ==========================================
        # ACCOUNT VALIDATION
        # ==========================================
        try:

            account = mongo.db.accounts.find_one(
                {
                    "_id": ObjectId(account_id),
                    "user_id": user_id
                }
            )

        except Exception:

            account = None

        if not account:

            flash(
                "Invalid account selected.",
                "danger"
            )

            return redirect(
                url_for("main.add_saving")
            )

        # ==========================================
        # DATE CONVERSION
        # ==========================================
        try:

            if start_date:

                start_date = datetime.strptime(
                    start_date,
                    "%Y-%m-%d"
                )

            else:

                start_date = datetime.utcnow()

        except:

            start_date = datetime.utcnow()

        try:

            if maturity_date:

                maturity_date = datetime.strptime(
                    maturity_date,
                    "%Y-%m-%d"
                )

            else:

                maturity_date = None

        except:

            maturity_date = None

        # ==========================================
        # STATUS VALIDATION
        # ==========================================
        if status not in [
            "active",
            "completed",
            "paused"
        ]:

            status = "active"

        # ==========================================
        # CREATE MODEL
        # ==========================================
        saving = Saving()

        data = saving.add(
            user_id=current_user.id,
            title=title,
            description=description,
            target_amount=target_amount,
            account_id=ObjectId(account_id),
            start_date=start_date,
            maturity_date=maturity_date,
            status=status
        )

        # ==========================================
        # MONGODB TYPES
        # ==========================================
        data["user_id"] = user_id
        data["account_id"] = ObjectId(account_id)
        data["target_amount"] = target_amount
        data["current_balance"] = 0.00
        data["status"] = status
        data["created_at"] = datetime.utcnow()
        data["updated_at"] = datetime.utcnow()

        # ==========================================
        # INSERT
        # ==========================================
        mongo.db.savings.insert_one(
            data
        )

        flash(
            "Saving goal created successfully.",
            "success"
        )

        return redirect(
            url_for("main.saving_list")
        )

    return render_template(
        "backend/pages/components/savings/add_saving.html",
        accounts=accounts
    )


@bp.route("/savings")
@login_required
def saving_list():


    user_id = ObjectId(current_user.id)



    # ============================
    # ACCOUNTS
    # ============================

    accounts = list(
        mongo.db.accounts.find(
            {
                "user_id": user_id,
                "status": True
            }
        )
        .sort(
            "created_at",
            -1
        )
    )




    # ============================
    # ACCOUNT MAP
    # ============================

    account_map = {

        str(account["_id"]):
        account.get("name","Unknown")

        for account in accounts

    }





    # ============================
    # SAVINGS
    # ============================

    savings = list(
        mongo.db.savings.find(
            {
                "user_id": user_id
            }
        )
        .sort(
            "created_at",
            -1
        )
    )





    # ============================
    # SAVING DETAILS
    # ============================

    for saving in savings:



        saving["account_name"] = account_map.get(

            str(
                saving.get("account_id")
            ),

            "No Account"

        )



        target = float(
            saving.get(
                "target_amount",
                0
            )
        )



        current = float(
            saving.get(
                "current_balance",
                0
            )
        )




        if target > 0:


            saving["progress"] = round(

                (current / target) * 100,

                2

            )


        else:

            saving["progress"] = 0





        # convert ObjectId for JS/Jinja

        saving["_id"] = str(
            saving["_id"]
        )





    return render_template(

        "backend/pages/components/savings/all_savings.html",

        savings=savings,

        accounts=accounts

    )



# ===============================
# EDIT SAVING
# ===============================
@bp.route("/edit-saving/<id>", methods=["GET", "POST"])
@login_required
def edit_saving(id):

    user_id = ObjectId(current_user.id)

    # ==========================================
    # GET SAVING
    # ==========================================
    try:

        saving = mongo.db.savings.find_one(
            {
                "_id": ObjectId(id),
                "user_id": user_id
            }
        )

    except Exception:

        saving = None

    if not saving:

        flash(
            "Saving goal not found.",
            "danger"
        )

        return redirect(
            url_for("main.saving_list")
        )

    # ==========================================
    # USER ACCOUNTS
    # ==========================================
    accounts = list(
        mongo.db.accounts.find(
            {
                "user_id": user_id
            }
        )
    )

    # ==========================================
    # POST
    # ==========================================
    if request.method == "POST":

        title = request.form.get(
            "title",
            ""
        ).strip()

        description = request.form.get(
            "description",
            ""
        ).strip()

        target_amount = request.form.get(
            "target_amount",
            ""
        ).strip()

        account_id = request.form.get(
            "account_id",
            ""
        ).strip()

        start_date = request.form.get(
            "start_date",
            ""
        ).strip()

        maturity_date = request.form.get(
            "maturity_date",
            ""
        ).strip()

        status = request.form.get(
            "status",
            "active"
        ).strip().lower()

        # ==========================================
        # VALIDATION
        # ==========================================
        if not title:

            flash(
                "Saving title is required.",
                "danger"
            )

            return redirect(request.url)

        if not target_amount:

            flash(
                "Target amount is required.",
                "danger"
            )

            return redirect(request.url)

        if not account_id:

            flash(
                "Please select account.",
                "danger"
            )

            return redirect(request.url)

        # ==========================================
        # TARGET AMOUNT
        # ==========================================
        try:

            target_amount = float(
                Decimal(target_amount).quantize(
                    Decimal("0.01")
                )
            )

        except InvalidOperation:

            flash(
                "Invalid target amount.",
                "danger"
            )

            return redirect(request.url)

        if target_amount <= 0:

            flash(
                "Target amount must be greater than zero.",
                "danger"
            )

            return redirect(request.url)

        # ==========================================
        # ACCOUNT CHECK
        # ==========================================
        try:

            account = mongo.db.accounts.find_one(
                {
                    "_id": ObjectId(account_id),
                    "user_id": user_id
                }
            )

        except Exception:

            account = None

        if not account:

            flash(
                "Invalid account selected.",
                "danger"
            )

            return redirect(request.url)

        # ==========================================
        # START DATE
        # ==========================================
        try:

            if start_date:

                start_date = datetime.strptime(
                    start_date,
                    "%Y-%m-%d"
                )

            else:

                start_date = saving.get("start_date")

        except:

            start_date = saving.get("start_date")

        # ==========================================
        # MATURITY DATE
        # ==========================================
        try:

            if maturity_date:

                maturity_date = datetime.strptime(
                    maturity_date,
                    "%Y-%m-%d"
                )

            else:

                maturity_date = None

        except:

            maturity_date = None

        # ==========================================
        # STATUS
        # ==========================================
        if status not in [
            "active",
            "paused",
            "completed"
        ]:

            status = "active"

        # ==========================================
        # UPDATE DATA
        # ==========================================
        update_data = {

            "title": title,

            "description": description,

            "target_amount": target_amount,

            "account_id": ObjectId(account_id),

            "start_date": start_date,

            "maturity_date": maturity_date,

            "status": status,

            # current_balance lama taabanayo

            "updated_at": datetime.utcnow()

        }

        # ==========================================
        # UPDATE
        # ==========================================
        mongo.db.savings.update_one(

            {
                "_id": ObjectId(id),
                "user_id": user_id
            },

            {
                "$set": update_data
            }

        )

        flash(
            "Saving goal updated successfully.",
            "success"
        )

        return redirect(
            url_for("main.saving_list")
        )

    # ==========================================
    # DATE FORMAT FOR INPUT TYPE=date
    # ==========================================
    if saving.get("start_date"):

        if isinstance(
            saving["start_date"],
            datetime
        ):

            saving["start_date"] = saving[
                "start_date"
            ].strftime("%Y-%m-%d")

    if saving.get("maturity_date"):

        if isinstance(
            saving["maturity_date"],
            datetime
        ):

            saving["maturity_date"] = saving[
                "maturity_date"
            ].strftime("%Y-%m-%d")

    return render_template(

        "backend/pages/components/savings/edit_saving.html",

        saving=saving,

        accounts=accounts

    )

# ===============================
# DELETE SAVING
# ===============================
@bp.route("/delete-saving/<id>")
@login_required
def delete_saving(id):

    saving = mongo.db.savings.find_one({
        "_id": ObjectId(id),
        "user_id": ObjectId(current_user.id)
    })

    if not saving:
        flash("Saving goal not found.", "danger")
        return redirect(url_for("main.saving_list"))

    mongo.db.savings.delete_one({
        "_id": ObjectId(id)
    })

    flash("Saving goal deleted successfully.", "success")

    return redirect(url_for("main.saving_list"))

@bp.route("/saving/<id>/add-transaction", methods=["GET", "POST"])
@login_required
def add_saving_transaction(id):

    user_id = ObjectId(current_user.id)

    # ==================================
    # VALIDATE SAVING ID
    # ==================================
    try:
        saving_id = ObjectId(id)

    except:
        flash("Invalid saving ID.", "danger")
        return redirect(url_for("main.saving_list"))

    # ==================================
    # GET SAVING
    # ==================================
    saving = mongo.db.savings.find_one({
        "_id": saving_id,
        "user_id": user_id
    })

    if not saving:
        flash("Saving goal not found.", "danger")
        return redirect(url_for("main.saving_list"))

    # ==================================
    # ACCOUNTS
    # ==================================
    accounts = list(
        mongo.db.accounts.find({
            "user_id": user_id,
            "status": True
        })
    )

    # ==================================
    # LAST ACCOUNT
    # ==================================
    last_transaction = mongo.db.saving_transactions.find_one(
        {
            "saving_id": saving_id,
            "user_id": user_id
        },
        sort=[("created_at", -1)]
    )

    selected_account_id = None

    if last_transaction:
        selected_account_id = str(
            last_transaction.get("account_id")
        )

    # ==================================
    # POST
    # ==================================
    if request.method == "POST":

        try:

            account_id = request.form.get("account_id")
            transaction_type = request.form.get("transaction_type")
            amount = request.form.get("amount")
            description = request.form.get("description", "")
            note = request.form.get("note", "")
            reference_no = request.form.get("reference_no", "")

            # ==================================
            # REQUIRED
            # ==================================
            if not account_id:
                flash("Please select account.", "danger")
                return redirect(request.url)

            if not amount:
                flash("Amount is required.", "danger")
                return redirect(request.url)

            # ==================================
            # AMOUNT
            # ==================================
            try:
                amount = float(amount)

                if amount <= 0:
                    flash("Amount must be greater than zero.", "danger")
                    return redirect(request.url)

            except:
                flash("Invalid amount.", "danger")
                return redirect(request.url)

            # ==================================
            # ACCOUNT ID
            # ==================================
            try:
                account_obj_id = ObjectId(account_id)

            except:
                flash("Invalid account selected.", "danger")
                return redirect(request.url)

            # ==================================
            # ACCOUNT
            # ==================================
            account = mongo.db.accounts.find_one({
                "_id": account_obj_id,
                "user_id": user_id
            })

            if not account:
                flash("Account not found.", "danger")
                return redirect(request.url)

            # ==================================
            # CURRENT VALUES
            # ==================================
            account_balance = float(
                account.get("balance", 0)
            )

            saving_balance = float(
                saving.get("current_balance", 0)
            )

            target_amount = float(
                saving.get("target_amount", 0)
            )

            remaining_target = max(
                target_amount - saving_balance,
                0
            )

            # ==================================
            # TARGET REACHED VALIDATION
            # ==================================
            if (
                transaction_type == "deposit"
                and saving_balance >= target_amount
                and target_amount > 0
            ):
                flash(
                    "Saving target already reached.",
                    "warning"
                )
                return redirect(request.url)

            # ==================================
            # MATURITY DATE VALIDATION
            # ==================================
            maturity_date = saving.get("maturity_date")

            if maturity_date:

                try:

                    if isinstance(maturity_date, str):

                        maturity_date = datetime.strptime(
                            maturity_date,
                            "%Y-%m-%d"
                        )

                    if (
                        transaction_type == "deposit"
                        and maturity_date.date() <
                        datetime.utcnow().date()
                    ):

                        flash(
                            "Saving maturity date has expired.",
                            "danger"
                        )

                        return redirect(request.url)

                except:
                    pass

            # ==================================
            # DEPOSIT VALIDATION
            # ==================================
            if transaction_type == "deposit":

                if amount > account_balance:

                    flash(
                        f"Insufficient account balance. "
                        f"Available ${account_balance:,.2f}",
                        "danger"
                    )
                    return redirect(request.url)

                # Deposit cannot exceed target
                if (
                    target_amount > 0 and
                    amount > remaining_target
                ):

                    flash(
                        f"Deposit exceeds target amount. "
                        f"Remaining needed: "
                        f"${remaining_target:,.2f}",
                        "warning"
                    )

                    return redirect(request.url)

            # ==================================
            # WITHDRAW VALIDATION
            # ==================================
            elif transaction_type == "withdrawal":

                if amount > saving_balance:

                    flash(
                        f"Insufficient saving balance. "
                        f"Available ${saving_balance:,.2f}",
                        "danger"
                    )

                    return redirect(request.url)

            else:

                flash(
                    "Invalid transaction type.",
                    "danger"
                )

                return redirect(request.url)

            # ==================================
            # INSERT TRANSACTION
            # ==================================
            trx = SavingTransaction()

            data = trx.add(
                user_id=user_id,
                saving_id=saving_id,
                account_id=account_obj_id,
                transaction_type=transaction_type,
                amount=amount,
                description=description,
                note=note,
                reference_no=reference_no
            )

            data["user_id"] = user_id
            data["saving_id"] = saving_id
            data["account_id"] = account_obj_id

            mongo.db.saving_transactions.insert_one(data)

            # ==================================
            # UPDATE BALANCES
            # ==================================
            if transaction_type == "deposit":

                mongo.db.accounts.update_one(
                    {"_id": account_obj_id},
                    {"$inc": {"balance": -amount}}
                )

                mongo.db.savings.update_one(
                    {"_id": saving_id},
                    {"$inc": {"current_balance": amount}}
                )

            else:

                mongo.db.accounts.update_one(
                    {"_id": account_obj_id},
                    {"$inc": {"balance": amount}}
                )

                mongo.db.savings.update_one(
                    {"_id": saving_id},
                    {"$inc": {"current_balance": -amount}}
                )

            # ==================================
            # REFRESH SAVING
            # ==================================
            updated = mongo.db.savings.find_one({
                "_id": saving_id
            })

            current_balance = float(
                updated.get("current_balance", 0)
            )

            target_amount = float(
                updated.get("target_amount", 0)
            )

            remaining_amount = max(
                target_amount - current_balance,
                0
            )

            progress = 0

            if target_amount > 0:
                progress = (
                    current_balance /
                    target_amount
                ) * 100

            progress = min(progress, 100)

            # ==================================
            # DAYS CALCULATION
            # ==================================
            days_remaining = 0
            daily_required = 0
            weekly_required = 0
            monthly_required = 0

            maturity_date = updated.get("maturity_date")

            try:

                if maturity_date:

                    if isinstance(maturity_date, str):

                        end_date = datetime.strptime(
                            maturity_date,
                            "%Y-%m-%d"
                        )

                    else:
                        end_date = maturity_date

                    days_remaining = (
                        end_date.date()
                        -
                        datetime.utcnow().date()
                    ).days

                    if days_remaining < 0:
                        days_remaining = 0

                    if (
                        days_remaining > 0 and
                        remaining_amount > 0
                    ):

                        daily_required = (
                            remaining_amount /
                            days_remaining
                        )

                        weekly_required = (
                            daily_required * 7
                        )

                        monthly_required = (
                            daily_required * 30
                        )

            except:
                pass

            # ==================================
            # STATUS
            # ==================================
            status = "active"

            if (
                target_amount > 0 and
                current_balance >= target_amount
            ):
                status = "completed"

            # maturity passed
            if (
                maturity_date and
                days_remaining == 0 and
                current_balance < target_amount
            ):
                status = "expired"

            # ==================================
            # UPDATE SAVING
            # ==================================
            mongo.db.savings.update_one(
                {"_id": saving_id},
                {
                    "$set": {
                        "status": status,
                        "progress": round(progress, 2),
                        "remaining_amount": round(remaining_amount, 2),
                        "days_remaining": days_remaining,
                        "daily_required": round(daily_required, 2),
                        "weekly_required": round(weekly_required, 2),
                        "monthly_required": round(monthly_required, 2),
                        "updated_at": datetime.utcnow()
                    }
                }
            )

            flash(
                "Saving transaction added successfully.",
                "success"
            )

            return redirect(
                url_for(
                    "main.saving_transaction_list",
                    saving_id=id
                )
            )

        except Exception as e:

            print("Saving Transaction Error:", e)

            flash(
                f"Unexpected error: {str(e)}",
                "danger"
            )

            return redirect(request.url)

    # ==================================
    # RENDER
    # ==================================
    saving = mongo.db.savings.find_one({
        "_id": saving_id,
        "user_id": user_id
    })

    return render_template(
        "backend/pages/components/savings_transaction/add_saving_transaction.html",
        saving=saving,
        accounts=accounts,
        selected_account_id=selected_account_id,
        days_remaining=saving.get("days_remaining", 0),
        daily_required=saving.get("daily_required", 0),
        weekly_required=saving.get("weekly_required", 0),
        monthly_required=saving.get("monthly_required", 0)
    )


@bp.route("/saving-transactions")
@login_required
def saving_transaction_list():

    transaction_type = request.args.get("type")
    saving_id = request.args.get("saving_id")
    account_id = request.args.get("account_id")

    query = {
        "user_id": ObjectId(current_user.id)
    }

    # Transaction Type Filter
    if transaction_type:
        query["transaction_type"] = transaction_type

    # Saving Filter
    if saving_id:
        query["saving_id"] = ObjectId(saving_id)

    # Account Filter
    if account_id:
        query["account_id"] = ObjectId(account_id)

    # Transactions
    transactions = list(
        mongo.db.saving_transactions.find(query).sort("created_at", -1)
    )

    # Savings
    savings = list(
        mongo.db.savings.find({
            "user_id": ObjectId(current_user.id)
        })
    )

    # Accounts
    accounts = list(
        mongo.db.accounts.find({
            "user_id": ObjectId(current_user.id)
        })
    )

    # Maps
    saving_map = {
        str(s["_id"]): s["title"]
        for s in savings
    }

    account_map = {
        str(a["_id"]): a["name"]
        for a in accounts
    }

    # Display Names
    for trx in transactions:

        trx["saving_name"] = saving_map.get(
            str(trx["saving_id"]),
            "Unknown Saving"
        )

        trx["account_name"] = account_map.get(
            str(trx["account_id"]),
            "Unknown Account"
        )

    return render_template(
        "backend/pages/components/savings_transaction/all_saving_transactions.html",
        transactions=transactions,
        savings=savings,
        accounts=accounts,
        selected_type=transaction_type,
        selected_saving=saving_id,
        selected_account=account_id
    )


# ==========================================================
# EDIT SAVING TRANSACTION
# ==========================================================
@bp.route("/saving-transaction/edit/<id>", methods=["GET", "POST"])
@login_required
def edit_saving_transaction(id):

    # ======================================================
    # VALIDATE ID
    # ======================================================

    try:

        trx_id = ObjectId(id)

    except:

        flash(
            "Invalid transaction ID.",
            "danger"
        )

        return redirect(
            url_for("main.saving_transaction_list")
        )


    user_id = ObjectId(
        current_user.id
    )


    # ======================================================
    # GET TRANSACTION
    # ======================================================

    trx = mongo.db.saving_transactions.find_one({

        "_id": trx_id,

        "user_id": user_id

    })


    if not trx:

        flash(
            "Transaction not found.",
            "danger"
        )

        return redirect(
            url_for("main.saving_transaction_list")
        )



    # ======================================================
    # GET SAVING
    # ======================================================

    saving = mongo.db.savings.find_one({

        "_id": trx["saving_id"],

        "user_id": user_id

    })


    if not saving:

        flash(
            "Saving goal not found.",
            "danger"
        )

        return redirect(
            url_for("main.saving_transaction_list")
        )



    # ======================================================
    # ACCOUNTS
    # ======================================================

    accounts = list(
        mongo.db.accounts.find({

            "user_id": user_id

        })
    )



    # ======================================================
    # OLD DATA
    # ======================================================

    old_account_id = trx["account_id"]

    old_amount = float(
        trx.get(
            "amount",
            0
        )
    )

    old_type = trx.get(
        "transaction_type"
    )


    saving_id = trx["saving_id"]



    # ======================================================
    # POST
    # ======================================================

    if request.method == "POST":


        # ==================================================
        # FORM DATA
        # ==================================================

        new_account = request.form.get(
            "account_id"
        )


        new_type = request.form.get(
            "transaction_type"
        )


        try:

            amount = float(
                request.form.get(
                    "amount",
                    0
                )
            )

        except:

            amount = 0



        description = request.form.get(
            "description",
            ""
        )


        note = request.form.get(
            "note",
            ""
        )


        reference_no = request.form.get(
            "reference_no",
            ""
        )


        transaction_date = request.form.get(
            "date"
        )



        # ==================================================
        # VALIDATION
        # ==================================================

        if not new_account:


            flash(
                "Account required.",
                "danger"
            )

            return redirect(
                request.url
            )



        try:

            new_account_id = ObjectId(
                new_account
            )


        except:


            flash(
                "Invalid account.",
                "danger"
            )


            return redirect(
                request.url
            )



        if amount <= 0:


            flash(
                "Amount must be greater than zero.",
                "danger"
            )


            return redirect(
                request.url
            )



        if new_type not in [

            "deposit",
            "withdrawal"

        ]:


            flash(
                "Invalid transaction type.",
                "danger"
            )


            return redirect(
                request.url
            )



        # ==================================================
        # ACCOUNT CHECK
        # ==================================================

        account = mongo.db.accounts.find_one({

            "_id": new_account_id,

            "user_id": user_id

        })


        if not account:


            flash(
                "Account not found.",
                "danger"
            )


            return redirect(
                request.url
            )



        # ==================================================
        # BALANCE CALCULATION
        # ==================================================

        available_account = float(

            account.get(
                "balance",
                0
            )

        )



        if (

            old_type == "deposit"

            and

            old_account_id == new_account_id

        ):


            available_account += old_amount




        available_saving = float(

            saving.get(
                "current_balance",
                0
            )

        )



        if old_type == "withdrawal":


            available_saving += old_amount




        # ==================================================
        # ACCOUNT BALANCE CHECK
        # ==================================================

        if new_type == "deposit":


            if amount > available_account:


                flash(
                    "Insufficient account balance.",
                    "danger"
                )


                return redirect(
                    request.url
                )



        # ==================================================
        # SAVING BALANCE CHECK
        # ==================================================

        if new_type == "withdrawal":


            if amount > available_saving:


                flash(
                    "Insufficient saving balance.",
                    "danger"
                )


                return redirect(
                    request.url
                )



        # ==================================================
        # TARGET VALIDATION
        # ==================================================

        if new_type == "deposit":


            target_amount = float(

                saving.get(
                    "target_amount",
                    0
                )

            )


            current_balance = float(

                saving.get(
                    "current_balance",
                    0
                )

            )



            # remove old transaction effect

            if old_type == "deposit":


                current_balance -= old_amount



            elif old_type == "withdrawal":


                current_balance += old_amount




            if (

                target_amount > 0

                and

                current_balance + amount > target_amount

            ):


                flash(
                    "Deposit exceeds saving target amount.",
                    "danger"
                )


                return redirect(
                    request.url
                )



        # ==================================================
        # MATURITY DATE VALIDATION
        # ==================================================

        maturity = saving.get(
            "maturity_date"
        )



        if maturity and transaction_date:


            try:


                if isinstance(
                    maturity,
                    str
                ):


                    maturity_date = datetime.strptime(

                        maturity,

                        "%Y-%m-%d"

                    ).date()


                else:


                    maturity_date = maturity.date()




                trx_date = datetime.strptime(

                    transaction_date,

                    "%Y-%m-%d"

                ).date()




                if trx_date > maturity_date:


                    flash(
                        "Transaction date cannot exceed maturity date.",
                        "danger"
                    )


                    return redirect(
                        request.url
                    )


            except Exception as e:


                print(
                    "DATE VALIDATION:",
                    e
                )



        # ==================================================
        # REMOVE OLD EFFECT
        # ==================================================

        if old_type == "deposit":


            mongo.db.accounts.update_one(

                {
                    "_id": old_account_id,
                    "user_id": user_id
                },

                {
                    "$inc":
                    {
                        "balance": old_amount
                    }
                }

            )


            mongo.db.savings.update_one(

                {
                    "_id": saving_id,
                    "user_id": user_id
                },

                {
                    "$inc":
                    {
                        "current_balance":
                        -old_amount
                    }
                }

            )



        else:


            mongo.db.accounts.update_one(

                {
                    "_id": old_account_id,
                    "user_id": user_id
                },

                {
                    "$inc":
                    {
                        "balance":
                        -old_amount
                    }
                }

            )


            mongo.db.savings.update_one(

                {
                    "_id": saving_id,
                    "user_id": user_id
                },

                {
                    "$inc":
                    {
                        "current_balance":
                        old_amount
                    }
                }

            )



        # ==================================================
        # APPLY NEW EFFECT
        # ==================================================

        if new_type == "deposit":


            mongo.db.accounts.update_one(

                {
                    "_id": new_account_id,
                    "user_id": user_id
                },

                {
                    "$inc":
                    {
                        "balance":
                        -amount
                    }
                }

            )


            mongo.db.savings.update_one(

                {
                    "_id": saving_id,
                    "user_id": user_id
                },

                {
                    "$inc":
                    {
                        "current_balance":
                        amount
                    }
                }

            )


        else:


            mongo.db.accounts.update_one(

                {
                    "_id": new_account_id,
                    "user_id": user_id
                },

                {
                    "$inc":
                    {
                        "balance":
                        amount
                    }
                }

            )


            mongo.db.savings.update_one(

                {
                    "_id": saving_id,
                    "user_id": user_id
                },

                {
                    "$inc":
                    {
                        "current_balance":
                        -amount
                    }
                }

            )



        # ==================================================
        # UPDATE TRANSACTION
        # ==================================================

        mongo.db.saving_transactions.update_one(

            {
                "_id": trx_id,
                "user_id": user_id
            },

            {

                "$set":
                {

                    "account_id":
                    new_account_id,

                    "transaction_type":
                    new_type,

                    "amount":
                    amount,

                    "description":
                    description,

                    "note":
                    note,

                    "reference_no":
                    reference_no,

                    "date":
                    datetime.strptime(
                        transaction_date,
                        "%Y-%m-%d"
                    ) if transaction_date else datetime.utcnow(),


                    "updated_at":
                    datetime.utcnow()

                }

            }

        )



        # ==================================================
        # ANALYTICS UPDATE
        # ==================================================

        updated = mongo.db.savings.find_one({

            "_id": saving_id,
            "user_id": user_id

        })



        current_balance = float(

            updated.get(
                "current_balance",
                0
            )

        )


        target_amount = float(

            updated.get(
                "target_amount",
                0
            )

        )



        remaining_amount = max(

            target_amount - current_balance,

            0

        )



        progress = 0


        if target_amount > 0:


            progress = (

                current_balance /

                target_amount

            ) * 100



        progress = min(

            round(progress,2),

            100

        )



        # STATUS

        if (

            target_amount > 0

            and

            current_balance >= target_amount

        ):


            status = "completed"



        else:


            status = "active"



        mongo.db.savings.update_one(

            {
                "_id": saving_id,
                "user_id": user_id
            },

            {

                "$set":
                {

                    "status": status,

                    "progress": progress,

                    "remaining_amount":
                    round(
                        remaining_amount,
                        2
                    ),

                    "updated_at":
                    datetime.utcnow()

                }

            }

        )



        flash(
            "Saving transaction updated successfully.",
            "success"
        )


        return redirect(

            url_for(
                "main.saving_transaction_list"
            )

        )

    # ======================================================
    # GET ANALYTICS
    # ======================================================


    current_balance = float(
        saving.get(
            "current_balance",
            0
        )
    )


    target_amount = float(
        saving.get(
            "target_amount",
            0
        )
    )


    remaining_amount = max(

        target_amount - current_balance,

        0

    )



    progress = 0


    if target_amount > 0:

        progress = (

            current_balance /

            target_amount

        ) * 100



    # ==============================
    # DATE CALCULATION
    # ==============================

    days_remaining = 0

    daily_required = 0

    weekly_required = 0

    monthly_required = 0



    maturity_date = saving.get(
        "maturity_date"
    )



    try:


        if maturity_date:


            if isinstance(
                maturity_date,
                str
            ):


                maturity_date = datetime.strptime(

                    maturity_date,

                    "%Y-%m-%d"

                )


            today = datetime.utcnow()



            days_remaining = (

                maturity_date.date()

                -

                today.date()

            ).days



            if days_remaining < 0:

                days_remaining = 0



            if days_remaining > 0:


                daily_required = (

                    remaining_amount /

                    days_remaining

                )


                weekly_required = (

                    daily_required * 7

                )


                monthly_required = (

                    daily_required * 30

                )



    except Exception as e:

        print(
            "DATE ERROR:",
            e
        )



    return render_template(

        "backend/pages/components/savings_transaction/edit_saving_transaction.html",

        trx=trx,

        saving=saving,

        accounts=accounts,

        progress=round(progress,2),

        remaining_amount=remaining_amount,

        days_remaining=days_remaining,

        daily_required=daily_required,

        weekly_required=weekly_required,

        monthly_required=monthly_required

    )


@bp.route("/saving-transaction/delete/<id>")
@login_required
def delete_saving_transaction(id):

    trx = mongo.db.saving_transactions.find_one({
        "_id": ObjectId(id),
        "user_id": ObjectId(current_user.id)
    })

    if not trx:
        flash("Transaction not found.", "danger")
        return redirect(url_for("main.saving_transaction_list"))

    amount = float(trx["amount"])
    account_id = trx["account_id"]
    saving_id = trx["saving_id"]

    # Reverse balances
    if trx["transaction_type"] == "deposit":

        mongo.db.accounts.update_one(
            {"_id": account_id},
            {"$inc": {"balance": amount}}
        )

        mongo.db.savings.update_one(
            {"_id": saving_id},
            {"$inc": {"current_balance": -amount}}
        )

    else:

        mongo.db.accounts.update_one(
            {"_id": account_id},
            {"$inc": {"balance": -amount}}
        )

        mongo.db.savings.update_one(
            {"_id": saving_id},
            {"$inc": {"current_balance": amount}}
        )

    mongo.db.saving_transactions.delete_one({
        "_id": ObjectId(id)
    })

    flash("Transaction deleted successfully.", "success")

    return redirect(url_for("main.saving_transaction_list"))



@bp.route("/reports/weekly")
@login_required
def weekly_report():
    from datetime import datetime, timedelta
    from bson import ObjectId

    # ------------------------------------
    # User Filter
    # ------------------------------------
    if current_user.role == "superadmin":
        query = {}
        category_filter = {}
    else:
        try:
            user_id = ObjectId(current_user.id)
        except:
            user_id = current_user.id

        query = {"user_id": user_id}
        category_filter = {"user_id": user_id}

    # ------------------------------------
    # Categories
    # ------------------------------------
    categories = list(
        mongo.db.categories.find(category_filter).sort("name", 1)
    )

    # ------------------------------------
    # GET FILTERS
    # ------------------------------------
    start_date = request.args.get("start_date", "").strip()
    end_date = request.args.get("end_date", "").strip()
    trans_type = request.args.get("type", "").strip()
    category = request.args.get("category", "").strip()
    search = request.args.get("search", "").strip()

    # ------------------------------------
    # Date Range
    # ------------------------------------
    if start_date and end_date:
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)

    else:
        end = datetime.utcnow()
        start = end - timedelta(days=7)

    query["date"] = {
        "$gte": start,
        "$lt": end
    }

    # ------------------------------------
    # Transaction Type
    # ------------------------------------
    if trans_type:
        query["transaction_type"] = trans_type

    # ------------------------------------
    # Category
    # ------------------------------------
    if category:
        query["category"] = category

    # ------------------------------------
    # Search
    # ------------------------------------
    if search:
        query["$or"] = [
            {"reference_no": {"$regex": search, "$options": "i"}},
            {"description": {"$regex": search, "$options": "i"}},
            {"item": {"$regex": search, "$options": "i"}},
            {"category": {"$regex": search, "$options": "i"}}
        ]

    # ------------------------------------
    # Transactions
    # ------------------------------------
    transactions = list(
        mongo.db.transactions.find(query).sort("date", -1)
    )

    # ------------------------------------
    # Totals
    # ------------------------------------
    total_income = sum(
        float(t.get("amount", 0))
        for t in transactions
        if t.get("transaction_type") == "income"
    )

    total_expense = sum(
        float(t.get("amount", 0))
        for t in transactions
        if t.get("transaction_type") == "expense"
    )

    net_balance = total_income - total_expense

    total_transactions = len(transactions)

    # ------------------------------------
    # Render
    # ------------------------------------
    return render_template(
        "backend/pages/components/reports/weekly_report.html",
        transactions=transactions,
        categories=categories,
        total_income=total_income,
        total_expense=total_expense,
        net_balance=net_balance,
        total_transactions=total_transactions,
        start_date=start,
        end_date=end
    )



@bp.route("/reports/monthly")
@login_required
def monthly_report():
    from datetime import datetime, timedelta
    from bson import ObjectId
    from calendar import monthrange

    today = datetime.utcnow()

    # ------------------------------------
    # USER FILTER
    # ------------------------------------
    if current_user.role == "superadmin":
        base_query = {}
    else:
        try:
            user_id = ObjectId(current_user.id)
        except:
            user_id = current_user.id
        base_query = {"user_id": user_id}

    # ------------------------------------
    # GET FILTERS
    # ------------------------------------
    year = request.args.get("year", type=int) or today.year
    month = request.args.get("month", type=int) or today.month

    start = datetime(year, month, 1)
    last_day = monthrange(year, month)[1]
    end = datetime(year, month, last_day, 23, 59, 59)

    query = {
        **base_query,
        "date": {"$gte": start, "$lte": end}
    }

    # ------------------------------------
    # TRANSACTIONS
    # ------------------------------------
    transactions = list(
        mongo.db.transactions.find(query).sort("date", 1)
    )

    # ------------------------------------
    # CATEGORY BREAKDOWN (LIKE YOUR EXCEL)
    # ------------------------------------
    category_data = {}

    for t in transactions:
        cat = t.get("category", "Unknown")
        t_type = t.get("transaction_type")
        amount = float(t.get("amount", 0))
        date = t.get("date")

        if cat not in category_data:
            category_data[cat] = {
                "income": 0,
                "expense": 0,
                "total": 0,
                "date": date   # ✔ ADD THIS
            }

        if t_type == "income":
            category_data[cat]["income"] += amount
        else:
            category_data[cat]["expense"] += amount

        category_data[cat]["total"] += amount

    # ------------------------------------
    # BALANCE CALCULATION (OPENING / CLOSING)
    # ------------------------------------
    before_query = {
        **base_query,
        "date": {"$lt": start}
    }

    before_transactions = list(mongo.db.transactions.find(before_query))

    opening_balance = sum(
        float(t.get("amount", 0)) if t.get("transaction_type") == "income"
        else -float(t.get("amount", 0))
        for t in before_transactions
    )

    income = sum(float(t.get("amount", 0)) for t in transactions if t.get("transaction_type") == "income")
    expense = sum(float(t.get("amount", 0)) for t in transactions if t.get("transaction_type") == "expense")

    closing_balance = opening_balance + income - expense

    # ------------------------------------
    # MONTHLY LIST (FOR TABLE)
    # ------------------------------------
    report_rows = []

    for cat, val in category_data.items():
        report_rows.append({
            "category": cat,
            "income": val["income"],
            "expense": val["expense"],
            "total": val["total"],
            "date": val.get("date")   # ✔ ADD THIS
        })

    # ------------------------------------
    # RENDER
    # ------------------------------------
    return render_template(
        "backend/pages/components/reports/monthly_report.html",
        transactions=transactions,
        report_rows=report_rows,
        opening_balance=opening_balance,
        closing_balance=closing_balance,
        total_income=income,
        total_expense=expense,
        year=year,
        month=month
    )



@bp.route("/reports/yearly")
@login_required
def yearly_report():

    from datetime import datetime
    from bson import ObjectId

    today = datetime.utcnow()

    start = datetime(today.year, 1, 1)

    if current_user.role == "superadmin":
        query = {
            "date": {
                "$gte": start,
                "$lte": today
            }
        }
    else:
        query = {
            "user_id": ObjectId(current_user.id),
            "date": {
                "$gte": start,
                "$lte": today
            }
        }

    transactions = list(
        mongo.db.transactions.find(query).sort("date", -1)
    )

    return render_template(
        "backend/pages/components/reports/weekly_report.html",
        transactions=transactions
    )



@bp.route("/reports/general")
@login_required
def general_report():

    if current_user.role not in ["superadmin", "admin"]:
        abort(403)

    # ------------------------------------
    # User Filter
    # ------------------------------------
    if current_user.role == "superadmin":
        query = {}
    else:
        try:
            query = {"user_id": ObjectId(current_user.id)}
        except:
            query = {"user_id": current_user.id}

    # ------------------------------------
    # Filters
    # ------------------------------------
    transaction_type = request.args.get("type")
    category = request.args.get("category")
    item = request.args.get("item")
    account_id = request.args.get("account")
    search = request.args.get("search")

    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    # ------------------------------------
    # Type
    # ------------------------------------
    if transaction_type:
        query["transaction_type"] = transaction_type

    # ------------------------------------
    # Category
    # ------------------------------------
    if category:
        query["category"] = category

    # ------------------------------------
    # Item
    # ------------------------------------
    if item:
        query["item"] = item

    # ------------------------------------
    # Account
    # ------------------------------------
    if account_id:
        query["account_id"] = account_id

    # ------------------------------------
    # Date Range
    # ------------------------------------
    if start_date or end_date:

        date_filter = {}

        if start_date:
            date_filter["$gte"] = datetime.strptime(
                start_date,
                "%Y-%m-%d"
            )

        if end_date:
            date_filter["$lte"] = datetime.strptime(
                end_date,
                "%Y-%m-%d"
            )

        query["date"] = date_filter

    # ------------------------------------
    # Search
    # ------------------------------------
    if search:

        query["$or"] = [
            {"description": {"$regex": search, "$options": "i"}},
            {"note": {"$regex": search, "$options": "i"}},
            {"reference_no": {"$regex": search, "$options": "i"}},
        ]

    # ------------------------------------
    # Transactions
    # ------------------------------------
    transactions = list(
        mongo.db.transactions
        .find(query)
        .sort("date", -1)
    )

    # ------------------------------------
    # Totals
    # ------------------------------------
    total_income = sum(
        float(t.get("amount", 0))
        for t in transactions
        if t.get("transaction_type") == "income"
    )

    total_expense = sum(
        float(t.get("amount", 0))
        for t in transactions
        if t.get("transaction_type") == "expense"
    )

    net_balance = total_income - total_expense

    # ------------------------------------
    # Filters Data
    # ------------------------------------
    if current_user.role == "superadmin":
        user_filter = {}
    else:
        user_filter = query.copy()
        user_filter.pop("$or", None)
        user_filter.pop("date", None)
        user_filter.pop("transaction_type", None)
        user_filter.pop("category", None)
        user_filter.pop("item", None)
        user_filter.pop("account_id", None)

    categories = list(
        mongo.db.categories.find(user_filter)
    )

    accounts = list(
        mongo.db.accounts.find(user_filter)
    )

    return render_template(
        "backend/reports/general.html",

        transactions=transactions,

        categories=categories,
        accounts=accounts,

        total_income=total_income,
        total_expense=total_expense,
        net_balance=net_balance,

        filters={
            "type": transaction_type,
            "category": category,
            "item": item,
            "account": account_id,
            "search": search,
            "start_date": start_date,
            "end_date": end_date,
        }
    )





#---------------------------------------------------
#---- Route: 70 | Dashboard - Backend Template -----
#---------------------------------------------------
@bp.route("/logout")
def logout():
    if current_user.is_authenticated:

        # Log the logout action
       

        # Only log out from Flask-Login
        logout_user()

        # ✅ Do NOT clear session or delete DB session yet
        # session.clear()  <-- remove this
        # db.session.delete(user_session)  <-- remove this

        # Flash message
        flash("You have been logged out! Your session record remains for inspection.", "success")

    # Clear remember_token cookie to prevent auto-login
    resp = make_response(redirect(url_for("main.index")))
    resp.set_cookie("remember_token", "", expires=0)
    return resp








